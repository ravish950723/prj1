"""buy_updated_FINAL_GENERATED.py

Final stabilized BUY pipeline entrypoint.

Fixes included
- Correct function signatures vs rank_long_short.py:
  - detect_*_setup(common, df, cfg)
  - score_long(common, cfg)
  - score_short(common, cfg)
  - build_*_plan(common, df, setup, cfg)
  - verdicts(..., cfg)
- Safe-unpack for fetching.fetch_data_daily_with_fallback (supports 2+ return values)
- OHLCV normalization (date index vs column; numeric conversion)
- On-the-fly indicator computation if missing: RSI_14, MACD, MACD_signal, MACD_hist, ATR_14
- FORCE creation + population of SRS columns (never missing; writes N/A/0 defaults)
- Helpful logs per symbol.

Usage
  python buy_updated_FINAL_GENERATED.py

Expected files in working dir
  - predictions_summary.xlsx (template)
  - config.py (symbols list)
  - fetching.py
  - rank_long_short.py
"""

from __future__ import annotations

import os
import sys
import math
import time
import json
import logging
from datetime import datetime
from typing import Dict, Any, Tuple, Optional, List

import numpy as np
import pandas as pd

# ---------------- MAX-ACCURACY BACKFILL HELPERS ----------------
DEFAULT_TEXT = "NONE"   # avoid Excel interpreting N/A as missing
DEFAULT_NUM = 0.0

def _safe_float(x, default=DEFAULT_NUM):
    try:
        if x is None:
            return default
        if isinstance(x, (int, float, np.floating, np.integer)):
            if np.isnan(x):
                return default
            return float(x)
        s = str(x).strip()
        if s == "" or s.lower() in ("nan","none","na","n/a"):
            return default
        return float(s)
    except Exception:
        return default

def _safe_text(x, default=DEFAULT_TEXT):
    try:
        if x is None:
            return default
        s = str(x).strip()
        if s == "" or s.lower() in ("nan","none","na","n/a"):
            return default
        return s
    except Exception:
        return default

def _backfill_legacy_columns(row_out: dict, common: dict, long_plan: dict, short_plan: dict):
    """
    Fill frequently-blank legacy columns using current IBKR/derived values.
    This does NOT invent ML model probabilities; if such columns exist, we fill with
    a technical proxy and label VALUE_ORIGIN.
    """
    price = _safe_float(common.get("PRICE"), DEFAULT_NUM)
    prev_close = _safe_float(common.get("PREV_CLOSE"), DEFAULT_NUM)
    today_open = _safe_float(common.get("TODAY_OPEN"), DEFAULT_NUM)
    today_high = _safe_float(common.get("TODAY_HIGH"), DEFAULT_NUM)
    today_low  = _safe_float(common.get("TODAY_LOW"), DEFAULT_NUM)

    # Common OHLC
    row_out.setdefault("CURRENT_PRICE", price)
    row_out.setdefault("PREV_CLOSE", prev_close)
    row_out.setdefault("TODAY_OPEN", today_open)
    row_out.setdefault("TODAY_HIGH", today_high)
    row_out.setdefault("TODAY_LOW", today_low)

    # Refined Buy Price: use long entry zone low if available, else DMA20 pullback proxy
    refined_buy = _safe_float(row_out.get("LONG_ENTRY_ZONE_LOW"), 0.0)
    if refined_buy <= 0:
        refined_buy = _safe_float(common.get("DMA20"), 0.0)
    if refined_buy <= 0:
        refined_buy = price
    row_out.setdefault("REFINED_BUY_PRICE", refined_buy)

    # Trend / Recommendation proxies
    row_out.setdefault("TREND", _safe_text(row_out.get("DMA_STACK"), DEFAULT_TEXT))
    row_out.setdefault("RECOMMENDATION", _safe_text(row_out.get("FINAL_ACTION"), DEFAULT_TEXT))

    # Confidence grade proxy
    conf = _safe_float(row_out.get("CONFIDENCE_SCORE"), 0.0)
    if conf >= 80: grade="A"
    elif conf >= 65: grade="B"
    elif conf >= 50: grade="C"
    elif conf > 0: grade="D"
    else: grade=DEFAULT_TEXT
    row_out.setdefault("CONFIDENCE_GRADE", grade)

    # Candle entry horizons: simple ATR-based bands around price (not ML, but actionable)
    atr = _safe_float(common.get("ATR14"), 0.0)
    for w, mult in [(2,1.0),(4,1.5),(8,2.0),(12,2.5),(18,3.0),(30,4.0)]:
        key = f"CANDLE_ENTRY_{w}W"
        if key not in row_out or _safe_text(row_out.get(key), "") in ("",DEFAULT_TEXT):
            if atr > 0 and price > 0:
                row_out[key] = max(price - atr*mult, 0.0)
            else:
                row_out[key] = DEFAULT_NUM

    # Model probability: technical proxy (sigmoid of normalized long-short score)
    lp = _safe_float(row_out.get("LONG_SCORE"), 0.0)
    sp = _safe_float(row_out.get("SHORT_SCORE"), 0.0)
    x = (lp - sp) / 25.0
    prob = float(1.0/(1.0+np.exp(-x))) if (lp>0 or sp>0) else DEFAULT_NUM
    row_out.setdefault("MODEL_PROB_PROXY", prob)

    # Mark origin
    row_out.setdefault("VALUE_ORIGIN", "IBKR_DERIVED_TECHNICAL")
from openpyxl import load_workbook

from fetching import fetch_data_daily_with_fallback
from rank_long_short import (
    RankConfig,
    compute_common_indicators_rls,
    detect_long_setup,
    detect_short_setup,
    score_long,
    score_short,
    build_long_plan,
    build_short_plan,
    explain_spike_drop,
    verdicts,
)


# ---------------- Logging ----------------

def _setup_logger() -> logging.Logger:
    os.makedirs("logs", exist_ok=True)
    logger = logging.getLogger("BUYPIPE")
    logger.setLevel(logging.INFO)
    logger.handlers = []

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(os.path.join("logs", "buy_pipeline.log"), mode="a", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


LOG = _setup_logger()


# ---------------- Excel helpers ----------------

def _excel_safe(v: Any) -> Any:
    """Convert non-scalar values so openpyxl doesn't crash; keep blanks as None."""
    if v is None:
        return None
    if isinstance(v, (tuple, list, dict)):
        return json.dumps(v, default=str)
    try:
        if isinstance(v, (np.generic,)):
            v = v.item()
    except Exception:
        pass
    try:
        if isinstance(v, float) and math.isnan(v):
            return None
    except Exception:
        pass
    return v


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u2013", "-").replace("\u2014", "-")
    return " ".join(s.strip().split())


def _ensure_header(ws, header_name: str) -> int:
    """Ensure header exists in row 1, return column index (1-based)."""
    header_name = _norm_header(header_name)
    headers = [_norm_header(c.value) for c in ws[1]]
    if header_name in headers:
        return headers.index(header_name) + 1
    col = len(headers) + 1
    ws.cell(row=1, column=col, value=header_name)
    return col


def _get_symbol_col(ws) -> int:
    headers = [_norm_header(c.value) for c in ws[1]]
    for i, h in enumerate(headers):
        if h.lower() == "symbol":
            return i + 1
    return _ensure_header(ws, "Symbol")


def _build_symbol_row_map(ws) -> Tuple[Dict[str, int], int]:
    sym_col = _get_symbol_col(ws)
    m: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=sym_col).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s:
            m[s] = r
    return m, sym_col


# ---------------- Required columns ----------------

REQUIRED_COLS: List[str] = [
    "DMA20",
    "DMA50",
    "DMA200",
    "PCT_FROM_DMA50",
    "PCT_FROM_DMA200",
    "VOL_TODAY",
    "AVG_VOL_20D",
    "VOL_SURGE_RATIO",
    "VWAP",
    "VWAP_DISTANCE_PCT",
    "ATR14_PCT",
    "DISTRIBUTION_DAYS_20D",
    "ACCUMULATION_DAYS_20D",
    "LONG_SETUP_TAG",
    "LONG_SCORE",
    "LONG_VERDICT",
    "SHORT_SETUP_TAG",
    "SHORT_SCORE",
    "SHORT_VERDICT",
    "FINAL_ACTION",
    "CONFIDENCE_SCORE",
    "Candle Entry 2w",
    "Candle Entry 4w",
    "Candle Entry 6w",
    "Candle Entry 8w",
    "Candle Entry 12w",
    "Candle Entry 18w",
    "Candle Entry 30w",
    "LONG_ENTRY_ZONE_LOW",
    "LONG_ENTRY_ZONE_HIGH",
    "LONG_INVALIDATION",
    "LONG_TARGET_1",
    "LONG_TARGET_2",
    "LONG_RR_RATIO",
    "SHORT_ENTRY_ZONE_LOW",
    "SHORT_ENTRY_ZONE_HIGH",
    "SHORT_INVALIDATION",
    "SHORT_TARGET_1",
    "SHORT_TARGET_2",
    "SHORT_RR_RATIO",]

DEFAULTS: Dict[str, Any] = {
    "DATA_SOURCE": "ERROR",
    "DMA20": "NONE",
    "DMA50": "NONE",
    "DMA200": "NONE",
    "PCT_FROM_DMA50": "NONE",
    "PCT_FROM_DMA200": "NONE",
    "VOL_TODAY": "NONE",
    "AVG_VOL_20D": "NONE",
    "VOL_SURGE_RATIO": "NONE",
    "VWAP": "NONE",
    "VWAP_DISTANCE_PCT": "NONE",
    "ATR14_PCT": "NONE",
    "DISTRIBUTION_DAYS_20D": 0,
    "ACCUMULATION_DAYS_20D": 0,
    "LONG_SETUP_TAG": "NONE",
    "LONG_SCORE": 0.0,
    "LONG_VERDICT": "AVOID",
    "SHORT_SETUP_TAG": "NONE",
    "SHORT_SCORE": 0.0,
    "SHORT_VERDICT": "AVOID",
    "FINAL_ACTION": "AVOID",
    "CONFIDENCE_SCORE": 0.0,
    "LONG_ENTRY_ZONE_LOW": "NONE",
    "LONG_ENTRY_ZONE_HIGH": "NONE",
    "LONG_INVALIDATION": "NONE",
    "LONG_TARGET_1": "NONE",
    "LONG_TARGET_2": "NONE",
    "LONG_RR_RATIO": "NONE",
    "SHORT_ENTRY_ZONE_LOW": "NONE",
    "SHORT_ENTRY_ZONE_HIGH": "NONE",
    "SHORT_INVALIDATION": "NONE",
    "SHORT_TARGET_1": "NONE",
    "SHORT_TARGET_2": "NONE",
    "SHORT_RR_RATIO": "NONE",

    "Candle Entry 2w": "NONE",
    "Candle Entry 4w": "NONE",
    "Candle Entry 6w": "NONE",
    "Candle Entry 8w": "NONE",
    "Candle Entry 12w": "NONE",
    "Candle Entry 18w": "NONE",
    "Candle Entry 30w": "NONE",}


# ---------------- Data helpers ----------------

def _safe_fetch(sym: str, ttl_minutes: int, refresh_cache: bool) -> Tuple[Optional[pd.DataFrame], str]:
    """Handle old/new return signatures from fetch_data_daily_with_fallback."""
    ret = fetch_data_daily_with_fallback(
        sym,
        bar_spec="10 Y",
        bar_size="1 day",
        ttl_minutes=ttl_minutes,
        require_today=False,
        force_refresh=refresh_cache,
    )

    if isinstance(ret, (list, tuple)):
        df = ret[0] if len(ret) > 0 else None
        src = ret[1] if len(ret) > 1 else "ERROR"
        return df, src

    # unexpected
    return None, "ERROR"


def _normalize_ohlcv(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure df has columns: date, open, high, low, close, volume and are numeric."""
    if df is None or df.empty:
        return df

    d = df.copy()

    # If date is the index, bring it back as a column
    if "date" not in d.columns:
        try:
            if d.index is not None:
                d = d.reset_index()
        except Exception:
            pass

    # Standardize common IBKR/AV variations
    rename_map = {
        "Open": "open",
        "High": "high",
        "Low": "low",
        "Close": "close",
        "Volume": "volume",
        "OPEN": "open",
        "HIGH": "high",
        "LOW": "low",
        "CLOSE": "close",
        "VOLUME": "volume",
    }
    for k, v in rename_map.items():
        if k in d.columns and v not in d.columns:
            d.rename(columns={k: v}, inplace=True)

    # Ensure required columns exist
    for c in ["open", "high", "low", "close", "volume"]:
        if c not in d.columns:
            d[c] = np.nan

    # Force numeric
    for c in ["open", "high", "low", "close", "volume"]:
        d[c] = pd.to_numeric(d[c], errors="coerce")

    d = d.dropna(subset=["close"]).reset_index(drop=True)
    return d


def _ema(s: pd.Series, span: int) -> pd.Series:
    return s.ewm(span=span, adjust=False).mean()


def _compute_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi


def _compute_atr(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14) -> pd.Series:
    prev_close = close.shift(1)
    tr = pd.concat([(high - low).abs(), (high - prev_close).abs(), (low - prev_close).abs()], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def _ensure_indicators(df: pd.DataFrame) -> pd.DataFrame:
    """Compute RSI_14, MACD*, ATR_14 if they are missing."""
    d = df.copy()
    close = d["close"]

    if "RSI_14" not in d.columns or d["RSI_14"].isna().all():
        d["RSI_14"] = _compute_rsi(close, 14)

    if "MACD" not in d.columns or d["MACD"].isna().all():
        ema12 = _ema(close, 12)
        ema26 = _ema(close, 26)
        macd = ema12 - ema26
        signal = _ema(macd, 9)
        hist = macd - signal
        d["MACD"] = macd
        d["MACD_signal"] = signal
        d["MACD_hist"] = hist

    if "ATR_14" not in d.columns or d["ATR_14"].isna().all():
        d["ATR_14"] = _compute_atr(d["high"], d["low"], close, 14)

    return d


def _compute_vwap_proxy(df: pd.DataFrame, lookback: int = 20) -> float:
    d = df.tail(lookback).copy()
    if d.empty:
        return np.nan
    tp = (d["high"] + d["low"] + d["close"]) / 3.0
    vol = d["volume"].fillna(0.0)
    denom = float(np.nansum(vol))
    if denom <= 0:
        return np.nan
    return float(np.nansum(tp * vol) / denom)


def _dist_acc_days(df: pd.DataFrame, lookback: int = 20) -> Tuple[int, int]:
    d = df.tail(lookback + 1).copy()
    if len(d) < lookback + 1:
        return 0, 0
    d["prev_close"] = d["close"].shift(1)
    avg = float(d["volume"].tail(lookback).mean())
    red = (d["close"] < d["prev_close"]) & (d["volume"] > avg)
    green = (d["close"] > d["prev_close"]) & (d["volume"] > avg)
    return int(red.tail(lookback).sum()), int(green.tail(lookback).sum())


def load_symbols_from_config_and_excel(template_path: str) -> List[str]:
    # config.py symbols
    try:
        from config import symbols as CONFIG_SYMBOLS
        cfg_syms = [str(x).strip().upper() for x in (CONFIG_SYMBOLS or []) if str(x).strip()]
    except Exception as e:
        cfg_syms = []
        LOG.warning(f"config.py symbols not loaded: {e}")

    # Excel symbols
    xl_syms: List[str] = []
    try:
        df_t = pd.read_excel(template_path, sheet_name=0)
        if "Symbol" in df_t.columns:
            xl_syms = [str(x).strip().upper() for x in df_t["Symbol"].dropna().tolist() if str(x).strip()]
    except Exception:
        xl_syms = []

    out = list(dict.fromkeys(cfg_syms + xl_syms))
    LOG.info(f"✅ Symbols loaded: config={len(cfg_syms)} excel={len(xl_syms)} total={len(out)}")
    return out


def _load_cfg_from_json(path: str) -> RankConfig:
    cfg = RankConfig()
    if not os.path.exists(path):
        return cfg
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for k, v in data.items():
            if hasattr(cfg, k):
                setattr(cfg, k, v)
        return cfg
    except Exception as e:
        LOG.warning(f"rank_config.json not loaded; using defaults. err={e}")
        return cfg


def update_excel(template_xlsx: str, out_xlsx: str, refresh_cache: bool = False, ttl_minutes: int = 360):
    cfg = _load_cfg_from_json("rank_config.json")

    wb = load_workbook(template_xlsx)
    ws = wb.active

    header_cols: Dict[str, int] = {}
    for c in ["DATA_SOURCE"] + REQUIRED_COLS + ["SPIKE_DRIVER", "DROP_DRIVER"]:
        header_cols[c] = _ensure_header(ws, c)

    symbol_row_map, sym_col = _build_symbol_row_map(ws)
    symbols = load_symbols_from_config_and_excel(template_xlsx)

    # Best-effort QQQ for relative strength
    try:
        qqq_df, _qqq_src = _safe_fetch("QQQ", ttl_minutes=ttl_minutes, refresh_cache=False)
        qqq_df = _ensure_indicators(_normalize_ohlcv(qqq_df)) if qqq_df is not None else None
    except Exception:
        qqq_df = None

    for i, sym in enumerate(symbols, start=1):
        t0 = time.time()
        res: Dict[str, Any] = {"Symbol": sym}
        res.update(DEFAULTS)

        df, src = _safe_fetch(sym, ttl_minutes=ttl_minutes, refresh_cache=refresh_cache)
        res["DATA_SOURCE"] = src

        if df is None or df.empty:
            LOG.warning(f"{sym} | src={src} | NO DATA -> writing defaults")
        else:
            try:
                df = _ensure_indicators(_normalize_ohlcv(df))

                # Need enough rows for DMA50 / indicators
                if len(df) < 60:
                    raise ValueError(f"insufficient rows={len(df)}")

                common = compute_common_indicators_rls(df, qqq_df=qqq_df)

                vwap = _compute_vwap_proxy(df, 20)
                price = float(common.get("PRICE")) if common.get("PRICE") is not None else float(df["close"].iloc[-1])
                vwap_dist = (price - vwap) / vwap * 100.0 if np.isfinite(vwap) and vwap != 0 else np.nan

                dist_days, acc_days = _dist_acc_days(df, 20)

                long_setup = detect_long_setup(common, df, cfg)
                short_setup = detect_short_setup(common, df, cfg)

                long_score = float(score_long(common, cfg)) if long_setup else 0.0
                short_score = float(score_short(common, cfg)) if short_setup else 0.0

                long_plan = build_long_plan(common, df, long_setup, cfg) if long_setup else {}
                short_plan = build_short_plan(common, df, short_setup, cfg) if short_setup else {}

                spike_driver, drop_driver = explain_spike_drop(common, long_setup, short_setup)

                long_verdict, short_verdict = verdicts(
                    long_score,
                    short_score,
                    long_setup,
                    short_setup,
                    common,
                    cfg,
                )

                # Final action/confidence
                if long_setup and long_score >= max(55.0, short_score) and long_verdict != "AVOID":
                    final_action = "BUY" if long_verdict == "BUY_NOW" else "WATCH"
                    confidence = min(100.0, max(0.0, long_score))
                elif short_setup and short_score > max(55.0, long_score) and short_verdict != "AVOID":
                    final_action = "SHORT" if short_verdict == "SHORT_NOW" else "WATCH"
                    confidence = min(100.0, max(0.0, short_score))
                else:
                    final_action = "AVOID"
                    confidence = 0.0

                # Fill SRS columns
                res.update({
                    "DMA20": common.get("DMA20", "NONE"),
                    "DMA50": common.get("DMA50", "NONE"),
                    "DMA200": common.get("DMA200", "NONE"),
                    "PCT_FROM_DMA50": common.get("PCT_FROM_DMA50", "NONE"),
                    "PCT_FROM_DMA200": common.get("PCT_FROM_DMA200", "NONE"),
                    "VOL_TODAY": common.get("VOL_TODAY", "NONE"),
                    "AVG_VOL_20D": common.get("AVG_VOL_20D", "NONE"),
                    "VOL_SURGE_RATIO": common.get("VOL_SURGE_RATIO", "NONE"),
                    "VWAP": float(vwap) if np.isfinite(vwap) else "NONE",
                    "VWAP_DISTANCE_PCT": float(vwap_dist) if np.isfinite(vwap_dist) else "NONE",
                    "ATR14_PCT": common.get("ATR14_PCT", "NONE"),
                    "DISTRIBUTION_DAYS_20D": dist_days,
                    "ACCUMULATION_DAYS_20D": acc_days,
                    "LONG_SETUP_TAG": long_setup or "NONE",
                    "LONG_SCORE": round(long_score, 2),
                    "LONG_VERDICT": long_verdict,
                    "SHORT_SETUP_TAG": short_setup or "NONE",
                    "SHORT_SCORE": round(short_score, 2),
                    "SHORT_VERDICT": short_verdict,
                    "FINAL_ACTION": final_action,
                    "CONFIDENCE_SCORE": round(confidence, 2),
                    "LONG_ENTRY_ZONE_LOW": long_plan.get("LONG_ENTRY_ZONE_LOW", "NONE"),
                    "LONG_ENTRY_ZONE_HIGH": long_plan.get("LONG_ENTRY_ZONE_HIGH", "NONE"),
                    "LONG_INVALIDATION": long_plan.get("LONG_INVALIDATION", "NONE"),
                    "LONG_TARGET_1": long_plan.get("LONG_TARGET_1", "NONE"),
                    "LONG_TARGET_2": long_plan.get("LONG_TARGET_2", "NONE"),
                    "LONG_RR_RATIO": long_plan.get("LONG_RR_RATIO", "NONE"),
                    "SHORT_ENTRY_ZONE_LOW": short_plan.get("SHORT_ENTRY_ZONE_LOW", "NONE"),
                    "SHORT_ENTRY_ZONE_HIGH": short_plan.get("SHORT_ENTRY_ZONE_HIGH", "NONE"),
                    "SHORT_INVALIDATION": short_plan.get("SHORT_INVALIDATION", "NONE"),
                    "SHORT_TARGET_1": short_plan.get("SHORT_TARGET_1", "NONE"),
                    "SHORT_TARGET_2": short_plan.get("SHORT_TARGET_2", "NONE"),
                    "SHORT_RR_RATIO": short_plan.get("SHORT_RR_RATIO", "NONE"),
                    "SPIKE_DRIVER": spike_driver,
                    "DROP_DRIVER": drop_driver,
                })


                # Candle Entry horizons (2w..30w)
                # These are action levels: ATR-based pullback bands around current price.
                # If ATR is missing, fall back to "NONE".
                try:
                    atr_val = float(df["ATR_14"].iloc[-1]) if ("ATR_14" in df.columns and len(df) > 0) else float("nan")
                except Exception:
                    atr_val = float("nan")

                if np.isfinite(atr_val) and atr_val > 0 and np.isfinite(price) and price > 0:
                    candle_map = {
                        "Candle Entry 2w":  max(price - 1.0 * atr_val, 0.0),
                        "Candle Entry 4w":  max(price - 1.5 * atr_val, 0.0),
                        "Candle Entry 6w":  max(price - 1.75 * atr_val, 0.0),
                        "Candle Entry 8w":  max(price - 2.0 * atr_val, 0.0),
                        "Candle Entry 12w": max(price - 2.5 * atr_val, 0.0),
                        "Candle Entry 18w": max(price - 3.0 * atr_val, 0.0),
                        "Candle Entry 30w": max(price - 4.0 * atr_val, 0.0),
                    }
                else:
                    candle_map = {k: "NONE" for k in [
                        "Candle Entry 2w","Candle Entry 4w","Candle Entry 6w","Candle Entry 8w",
                        "Candle Entry 12w","Candle Entry 18w","Candle Entry 30w"
                    ]}

                res.update(candle_map)

                LOG.info(
                    f"{sym} | src={src} | DMA20={res['DMA20']} DMA50={res['DMA50']} DMA200={res['DMA200']} | "
                    f"VOLsurge={res['VOL_SURGE_RATIO']} | ATR%={res['ATR14_PCT']} | "
                    f"L={res['LONG_SETUP_TAG']} {res['LONG_SCORE']} {res['LONG_VERDICT']} | "
                    f"S={res['SHORT_SETUP_TAG']} {res['SHORT_SCORE']} {res['SHORT_VERDICT']} | "
                    f"FINAL={res['FINAL_ACTION']} conf={res['CONFIDENCE_SCORE']}"
                )

            except Exception as e:
                LOG.exception(f"{sym} compute failed -> defaults used. err={e}")

        # Write row (append if missing)
        row = symbol_row_map.get(sym)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=sym_col, value=sym)
            symbol_row_map[sym] = row

        for col_name, col_idx in header_cols.items():
            ws.cell(row=row, column=col_idx, value=_excel_safe(res.get(col_name)))

        if i % 10 == 0:
            LOG.info(f"Progress: {i}/{len(symbols)} symbols processed. last={sym} ({time.time()-t0:.2f}s)")

    # Save (avoid permission crash by timestamping if locked)
    try:
        wb.save(out_xlsx)
        LOG.info(f"✅ Saved: {out_xlsx}")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = out_xlsx.replace(".xlsx", f"_{ts}.xlsx")
        wb.save(alt)
        LOG.warning(f"⚠️ Output locked; saved to: {alt}")


def main() -> None:
    template = "predictions_summary.xlsx"
    out = "predictions_summary_out.xlsx"
    refresh_cache = "--refresh" in sys.argv
    update_excel(template, out, refresh_cache=refresh_cache, ttl_minutes=360)


if __name__ == "__main__":
    main()