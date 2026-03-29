
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Optional
from upward import compute_signal_score as _tech_fallback_score
from buy_rl import _predict_proba_for_last_row, BUY_THRESH, get_confidence_band , _pattern_list

import numpy as np
import pandas as pd
import openpyxl
import joblib

from config import symbols as CONFIG_SYMBOLS
from fetching import fetch_data_cached, fetch_data_daily_with_fallback
from compute import compute_indicators, analyze_symbol_all, compute_market_stage_substage, compute_patterns
from backtest import evaluate_backtest_accuracy
from eps_features import fetch_quarterly_eps, eps_growth_flags, fetch_market_sentiment
from exit_signals import compute_exit_signals
from symbol_analysis import analyze_symbol
from upward import (
    detect_smc_accumulation_breakout,
    detect_mean_reversion_buy,
    detect_bullish_engulfing,
    detect_hammer,
    compute_upward_trend,
    compute_signal_score
)

# Long/Short ranking (new)
from rank_long_short import (
    RankConfig,
    compute_common_indicators_rls,
    detect_long_setup,
    detect_short_setup,
    score_long,
    score_short,
    build_long_plan,
    build_short_plan,
    feasibility_label,
    liquidity_score,
)

# -----------------------------
# Thresholds / Bands (sync with ddddd_buy.py)
# -----------------------------
WATCH_PROB = 0.40
DEFAULT_BUY_PROB = 0.62
STRONG_BUY_PROB = 0.80
TECH_BUY_FALLBACK = 0.90

HERE = Path(__file__).resolve().parent
TEMPLATE_FILENAME = "predictions_summary.xlsx"

import pandas as pd
import numpy as np



STRONG_BUY_THRESH = float(STRONG_BUY_PROB)
# ================================
# Helper functions (safe inserts)
# ================================

def _compute_indicator_snap(df):
    """
    Best-effort indicator snapshot for display columns.
    Never raises — returns dict with safe defaults and expected legacy keys.
    """
    try:
        if df is None or len(df) == 0:
            return {}
        last = df.iloc[-1]
        snap = {}

        def _num(v):
            try:
                return float(v) if v is not None and not pd.isna(v) else None
            except Exception:
                return None

        def _pick(*names):
            for n in names:
                if n in df.columns:
                    return last.get(n)
            return None

        # Numeric aliases from both compute.py and rank_long_short common schema
        snap["RSI_14"] = _num(_pick("RSI14", "RSI_14", "RSI"))
        snap["ADX14"] = _num(_pick("ADX14", "ADX_14", "ADX"))
        snap["MACD"] = _num(_pick("MACD"))
        snap["MACD_SIGNAL"] = _num(_pick("MACD_SIGNAL", "MACD_signal"))
        snap["VWAP"] = _num(_pick("VWAP", "vwap_support"))
        snap["ATR14_PCT"] = _num(_pick("ATR14_PCT"))
        snap["EMA21"] = _num(_pick("EMA21", "EMA_21"))
        snap["EMA50"] = _num(_pick("EMA50", "EMA_50"))
        snap["EMA200"] = _num(_pick("EMA200", "EMA_200"))
        snap["VWAP_DISTANCE_PCT"] = _num(_pick("VWAP_DISTANCE_PCT"))
        snap["VOL_SURGE_RATIO"] = _num(_pick("VOL_SURGE_RATIO"))

        closef = _num(_pick("close", "CLOSE", "Close"))
        ema21f = snap["EMA21"]
        ema50f = snap["EMA50"]
        ema200f = snap["EMA200"]

        # EMA21 slope from either schema
        ema21_slope = None
        if "EMA21" in df.columns and len(df) >= 5:
            try:
                ema21_slope = float(df["EMA21"].iloc[-1] - df["EMA21"].iloc[-5])
            except Exception:
                ema21_slope = None
        elif "EMA_21" in df.columns and len(df) >= 5:
            try:
                ema21_slope = float(df["EMA_21"].iloc[-1] - df["EMA_21"].iloc[-5])
            except Exception:
                ema21_slope = None
        snap["EMA21_Slope"] = ema21_slope

        adx = snap["ADX14"]
        if adx is not None and np.isfinite(adx):
            if adx >= 25:
                adx_strength = "STRONG"
            elif adx >= 18:
                adx_strength = "MEDIUM"
            else:
                adx_strength = "WEAK"
        else:
            adx_strength = "N/A"
        snap["ADX_Strength"] = adx_strength

        ema_stack = all(v is not None and np.isfinite(v) for v in (ema21f, ema50f, ema200f)) and (ema21f > ema50f > ema200f)
        close_above = closef is not None and ema21f is not None and np.isfinite(closef) and np.isfinite(ema21f) and closef >= ema21f
        ema_rising = ema21_slope is not None and np.isfinite(ema21_slope) and ema21_slope > 0
        snap["EMA_Uptrend"] = bool(ema_stack or (close_above and ema_rising))

        # Legacy keys expected later in pipeline
        macd_cross = bool(last.get("MACD_Crossover", last.get("MACD_crossover", last.get("MACD_CROSSOVER", False))))
        snap["MACD_Crossover"] = macd_cross
        snap["RSI_State"] = (
            "OVERBOUGHT" if snap["RSI_14"] is not None and snap["RSI_14"] >= 70 else
            "OVERSOLD" if snap["RSI_14"] is not None and snap["RSI_14"] <= 30 else
            "BULLISH" if snap["RSI_14"] is not None and snap["RSI_14"] >= 55 else
            "BEARISH" if snap["RSI_14"] is not None and snap["RSI_14"] <= 45 else
            "NEUTRAL"
        ) if snap["RSI_14"] is not None else "N/A"

        # OBV trend based on recent delta
        obv_series = None
        if "OBV" in df.columns and len(df) >= 6:
            try:
                obv_series = pd.to_numeric(df["OBV"], errors="coerce")
                snap["OBV_Trend"] = "UP" if float(obv_series.iloc[-1] - obv_series.iloc[-6]) > 0 else "DOWN"
            except Exception:
                snap["OBV_Trend"] = "N/A"
        else:
            snap["OBV_Trend"] = "N/A"

        bb_lower = _num(_pick("BB_lower"))
        snap["At_BB_Lower"] = bool(closef is not None and bb_lower is not None and np.isfinite(closef) and np.isfinite(bb_lower) and closef <= bb_lower)

        return snap
    except Exception:
        return {}



def ensure_enrichment_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Guarantee presence of enrichment columns so downstream code never KeyErrors."""
    if df is None:
        return pd.DataFrame()
    df = df.copy()
    bool_defaults = {
        "MACD_Crossover": False,
        "MACD_crossover": False,
        "MACD_CROSSOVER": False,
        "EMA_uptrend": False,
        "near_support": False,
        "volume_surge": False,
        "tight_range": False,
        "refined_buy_signal": False,
    }
    num_defaults = {
        "MACD": 0.0,
        "MACD_signal": 0.0,
        "MACD_hist": 0.0,
        "MACD_hist_slope": 0.0,
        "ADX_14": 0.0,
        "RSI_14": 0.0,
        "ATR_14": 0.0,
        "EMA_20": 0.0,
        "EMA_21": 0.0,
        "EMA_50": 0.0,
        "EMA_200": 0.0,
        "OBV": 0.0,
        "vwap_support": 0.0,
        "darvas_signal": 0.0,
        "darvas_breakout_pct": 0.0,
        "institutional_score": 0.0,
        "volume_weight": 0.0,
        "confidence_score": 0.0,
        "signal_score": 0.0,
        "sym_vol_regime": 0.0,
        "VIX_vol_regime": 0.0,
        "volume_pressure": 0.0,
    }
    str_defaults = {
        "HTF_Trend": "NEUTRAL",
        "ITF_Trend": "NEUTRAL",
        "LTF_Trend": "NEUTRAL",
        "market_stage": "Neutral/Transition",
        "market_substage": "NEUTRAL_CHOP",
        "rule_recommendation": "HOLD",
    }
    for col, default in bool_defaults.items():
        if col not in df.columns:
            df[col] = default
        df[col] = df[col].fillna(default).astype(bool)
    for col, default in num_defaults.items():
        if col not in df.columns:
            df[col] = default
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(default)
    for col, default in str_defaults.items():
        if col not in df.columns:
            df[col] = default
        df[col] = df[col].fillna(default).astype(str)
    return df


def _is_yes(x) -> bool:
    if x is None:
        return False
    if isinstance(x, bool):
        return x
    if isinstance(x, (int, float)):
        return x != 0
    s = str(x).strip().lower()
    return s in {"y", "yes", "true", "1", "buy", "strong_buy", "strong buy"}


def _is_no(x) -> bool:
    if x is None:
        return False
    if isinstance(x, bool):
        return not x
    if isinstance(x, (int, float)):
        return x == 0
    s = str(x).strip().lower()
    return s in {"n", "no", "false", "0", "sell", "avoid"}


def _weekly_upper_wick_high_volume(df: pd.DataFrame,
                                  wick_ratio_min: float = 0.55,
                                  vol_mult: float = 1.2,
                                  lookback_weeks: int = 12) -> bool:
    """
    True if the most recent WEEK has:
      - large upper wick relative to total candle range
      - and weekly volume is elevated vs trailing weekly average
    """
    try:
        if df is None or df.empty:
            return False
        if not {"date","open","high","low","close","volume"}.issubset(df.columns):
            return False

        d = df.copy()
        d["date"] = pd.to_datetime(d["date"], errors="coerce")
        d = d.dropna(subset=["date"]).sort_values("date")

        # Weekly OHLCV
        w = d.set_index("date").resample("W-FRI").agg({
            "open": "first",
            "high": "max",
            "low": "min",
            "close": "last",
            "volume": "sum"
        }).dropna()

        if len(w) < max(5, lookback_weeks):
            return False

        last = w.iloc[-1]
        rng = float(last["high"] - last["low"])
        if not np.isfinite(rng) or rng <= 0:
            return False

        upper_wick = float(last["high"] - max(last["open"], last["close"]))
        upper_wick_ratio = upper_wick / rng

        vol_base = w["volume"].iloc[-lookback_weeks:-1].mean()
        if not np.isfinite(vol_base) or vol_base <= 0:
            return False

        high_vol = float(last["volume"]) >= vol_mult * float(vol_base)
        big_upper_wick = upper_wick_ratio >= wick_ratio_min

        return bool(big_upper_wick and high_vol)

    except Exception:
        return False


# ------------------------------------------------------------------
# Backward-compat alias
# Some parts of the pipeline call _pct_weekly_high_vs_prev_lower_volume_high(),
# but the implemented function name is _pct_high_from_prev_low_volume_week().
# Keep both so enrichment doesn't fail and Excel columns get populated.
# ------------------------------------------------------------------
def _pct_weekly_high_vs_prev_lower_volume_high(df: pd.DataFrame, lookback_weeks: int = 20) -> float:
    return _pct_high_from_prev_low_volume_week(df, lookback_weeks=lookback_weeks)

def _pct_high_from_prev_low_volume_week(df: pd.DataFrame, lookback_weeks: int = 20) -> float:
    """
    Returns % distance of current weekly high from the most recent 'low-volume week' close.
    """
    try:
        if df is None or df.empty:
            return np.nan
        if not {"date","open","high","low","close","volume"}.issubset(df.columns):
            return np.nan

        d = df.copy()
        d["date"] = pd.to_datetime(d["date"], errors="coerce")
        d = d.dropna(subset=["date"]).sort_values("date")

        w = d.set_index("date").resample("W-FRI").agg({
            "open": "first",
            "high": "max",
            "low": "min",
            "close": "last",
            "volume": "sum"
        }).dropna()

        if len(w) < 5:
            return np.nan

        w_tail = w.tail(lookback_weeks).copy()
        vol_med = w_tail["volume"].median()
        low_vol_weeks = w_tail[w_tail["volume"] <= vol_med * 0.75]

        if low_vol_weeks.empty:
            return np.nan

        ref_close = float(low_vol_weeks.iloc[-1]["close"])
        cur_high = float(w_tail.iloc[-1]["high"])
        if ref_close <= 0:
            return np.nan

        return round((cur_high - ref_close) / ref_close * 100.0, 2)

    except Exception:
        return np.nan


def _load_rank_config() -> RankConfig:
    """Load rank_config.json if present; otherwise use defaults."""
    cfg = RankConfig()
    p = HERE / "rank_config.json"
    if not p.exists():
        return cfg
    try:
        obj = json.loads(p.read_text(encoding="utf-8"))
        th = (obj.get("thresholds") or {})
        lng = (th.get("long") or {})
        sht = (th.get("short") or {})
        sc = (obj.get("scoring") or {})
        w = (sc.get("weights") or {})

        # thresholds
        cfg.vol_surge_breakout = float(lng.get("vol_surge_breakout", cfg.vol_surge_breakout))
        cfg.vol_surge_reversal = float(lng.get("vol_surge_reversal", cfg.vol_surge_reversal))
        cfg.vol_pullback_max = float(lng.get("vol_pullback_max", cfg.vol_pullback_max))
        cfg.rsi_breakout_min = float(lng.get("rsi_breakout_min", cfg.rsi_breakout_min))
        cfg.rsi_pullback_min = float(lng.get("rsi_pullback_min", cfg.rsi_pullback_min))
        cfg.long_score_buy_now = float(lng.get("long_score_buy_now", cfg.long_score_buy_now))
        cfg.long_score_wait_dip = float(lng.get("long_score_wait_dip", cfg.long_score_wait_dip))

        cfg.rsi_short_max = float(sht.get("rsi_short_max", cfg.rsi_short_max))
        cfg.gap_down_pct = float(sht.get("gap_down_pct", cfg.gap_down_pct))
        cfg.overext_dma20_pct = float(sht.get("overext_dma20_pct", cfg.overext_dma20_pct))
        cfg.short_score_short_now = float(sht.get("short_score_short_now", cfg.short_score_short_now))
        cfg.short_score_wait_bounce = float(sht.get("short_score_wait_bounce", cfg.short_score_wait_bounce))
        cfg.borrow_fee_max = float(sht.get("borrow_fee_max", cfg.borrow_fee_max))

        # weights
        cfg.w_trend = float(w.get("trend", cfg.w_trend))
        cfg.w_volume = float(w.get("volume", cfg.w_volume))
        cfg.w_momentum = float(w.get("momentum", cfg.w_momentum))
        cfg.w_volatility = float(w.get("volatility", cfg.w_volatility))
        cfg.w_rel_strength = float(w.get("relative_strength", cfg.w_rel_strength))
        cfg.w_liquidity = float(w.get("liquidity", cfg.w_liquidity))

        return cfg
    except Exception as e:
        print(f"[WARN] rank_config.json parse failed: {e}")
        return cfg


# Global rank config
RANK_CFG = _load_rank_config()


def _load_top_n(default_n: int = 20) -> int:
    p = HERE / "rank_config.json"
    if not p.exists():
        return default_n
    try:
        obj = json.loads(p.read_text(encoding="utf-8"))
        sc = (obj.get("scoring") or {})
        n = int(sc.get("top_n", default_n))
        return max(5, min(100, n))
    except Exception:
        return default_n


TOP_N = _load_top_n()


# Load required model feature list (training-time contract)
REQUIRED_FEATURES: Optional[List[str]] = None
try:
    feats_path = HERE / "model_features.txt"
    if feats_path.exists():
        REQUIRED_FEATURES = [ln.strip() for ln in feats_path.read_text(encoding="utf-8").splitlines() if ln.strip()]
        if REQUIRED_FEATURES:
            print(f"✅ Loaded model feature contract: {len(REQUIRED_FEATURES)} features from model_features.txt")
except Exception as _e:
    REQUIRED_FEATURES = None


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--no-cache", action="store_true", help="Bypass cache entirely")
    p.add_argument("--refresh-cache", choices=["none", "force"], default="none",
                   help="Cache policy: 'force' to re-fetch all symbols")
    p.add_argument("--cache-ttl-mins", type=int, default=360,
                   help="Refresh data if cache older than this many minutes")
    p.add_argument("--allow-stale", action="store_true",
                   help="Allow cache that doesn’t include today’s data")
    p.add_argument("--excel-out", type=str, default="predictions_summary_out.xlsx")
    p.add_argument("--template", type=str, default=TEMPLATE_FILENAME, help="Template xlsx file")
    p.add_argument("--sheet", type=str, default="", help="Optional sheet name (auto-detect if blank)")
    return p.parse_args()


def load_symbol_df_info(sym: str, args) -> tuple[pd.DataFrame, str, str]:
    """Return (df, DATA_SOURCE, error_message)."""
    force = args.no_cache or (args.refresh_cache == "force")
    ttl = 0 if args.no_cache else args.cache_ttl_mins
    df, source, err = fetch_data_daily_with_fallback(
        sym,
        bar_spec="10 Y",
        bar_size="1 day",
        ttl_minutes=ttl,
        require_today=not args.allow_stale,
        force_refresh=force,
    )
    return df, source, err


def load_symbol_df(sym: str, args) -> pd.DataFrame:
    """Backwards-compatible: only returns df."""
    df, _src, _err = load_symbol_df_info(sym, args)
    return df


def load_model() -> Optional[object]:
    p = HERE / "strong_buy_xgb_model_calibrated.pkl"
    if p.exists():
        m = joblib.load(p)
        print(f"✅ Loaded model: {p.name}")
        return m
    print("⚠️ Calibrated model not found; Model Probability will be 0.")
    return None


def load_threshold(default: float = DEFAULT_BUY_PROB) -> float:
    """Load probability threshold for 'strong_buy' classification.

    Priority:
      1) THRESHOLDS_FILE env var (file containing either model_probability_thresholds or best_f1.thr)
      2) strong_buy_thresholds_15pct_45d_proxy.json (preferred, if present)
      3) strong_buy_thresholds.json
      4) fallback default
    """
    import os
    candidates = [
        os.getenv("THRESHOLDS_FILE"),
        str(HERE / "strong_buy_thresholds_15pct_45d_proxy.json"),
        str(HERE / "strong_buy_thresholds.json"),
    ]
    for p in candidates:
        if not p:
            continue
        try:
            pp = Path(p)
            if not pp.exists():
                continue
            obj = json.loads(pp.read_text(encoding="utf-8"))
            # New format: {"model_probability_thresholds": {"BUY": 0.62, ...}}
            thr = None
            if isinstance(obj, dict):
                thr = (obj.get("model_probability_thresholds") or {}).get("BUY")
                if thr is None:
                    # Legacy sweep format: {"best_f1": {"thr": 0.05, ...}}
                    thr = (obj.get("best_f1") or {}).get("thr")
            if thr is not None:
                thrf = float(thr)
                print(f"✅ Using threshold: {thrf:.2f} (source={pp.name})")
                return thrf
        except Exception as e:
            print(f"[WARN] failed to read thresholds from {p}: {e}")
    print(f"[INFO] Using default threshold: {default:.2f}")
    return float(default)

def add_trade_management_columns(df_summary: pd.DataFrame) -> pd.DataFrame:
    """Add actionable columns that convert signals into a trade plan.

    Adds (high-priority):
      - Final_Action
      - Best_Risk_Reward
      - Primary_Entry_Price
      - Add_On_Dip_Price
      - Invalidation_Level
      - Risk_%
      - Reward_%
      - Risk_Reward_Ratio
      - Position_Size_Class
      - Buy_Window_Status

    Notes:
      * Uses existing pipeline columns only.
      * Requires "Current Price" (we populate it in enrichment loop).
    """
    if df_summary is None or df_summary.empty:
        return df_summary

    out = df_summary.copy()

    def _num(x, default=np.nan):
        try:
            v = float(x)
            return v if np.isfinite(v) else default
        except Exception:
            return default

    def _is_up_trend(v) -> bool:
        s = str(v or "").strip().lower()
        return s in {"up", "bull", "bullish", "markup", "mark-up", "✅"}

    def _is_stage_ok(v) -> bool:
        s = str(v or "").strip().lower()
        return ("accum" in s) or ("mark" in s and "down" not in s)

    def _is_buyish_signal(sig) -> bool:
        s = str(sig or "").strip().lower()
        return s.startswith("buy")

    def _choose_primary_entry(r: dict):
        """Return (entry_price, source)"""
        candidates = [
            ("Refined Buy Price", r.get("Refined Buy Price")),
            ("Candle Entry 2w", r.get("Candle Entry 2w")),
            ("VWAP Support", r.get("VWAP Support")),
            ("Candle Entry 4w", r.get("Candle Entry 4w")),
            ("Candle Entry 6w", r.get("Candle Entry 6w")),
            ("Candle Entry 8w", r.get("Candle Entry 8w")),
        ]
        for name, val in candidates:
            pv = _num(val)
            if np.isfinite(pv) and pv > 0:
                return pv, name
        # fallback
        cp = _num(r.get("Current Price"))
        return (cp, "Current Price") if np.isfinite(cp) else (np.nan, "")

    def _choose_add_on(entry_source: str, r: dict):
        # conservative ladder for scaling
        ladder = {
            "Refined Buy Price": "Candle Entry 4w",
            "Candle Entry 2w": "Candle Entry 4w",
            "VWAP Support": "Candle Entry 4w",
            "Candle Entry 4w": "Candle Entry 6w",
            "Candle Entry 6w": "Candle Entry 8w",
            "Candle Entry 8w": "Candle Entry 12w",
        }
        nxt = ladder.get(entry_source, "Candle Entry 4w")
        pv = _num(r.get(nxt))
        return pv if np.isfinite(pv) and pv > 0 else np.nan

    # compute per-row
    primary_entry = []
    add_on = []
    invalidation = []
    risk_pct = []
    reward_pct = []
    rr_ratio = []
    buy_window = []
    final_action = []
    entry_source_list = []
    size_class = []

    # helper metrics for ranking
    rr_for_rank = []
    score_for_rank = []

    for _, row in out.iterrows():
        r = row.to_dict()

        cp = _num(r.get("Current Price"))
        entry, entry_src = _choose_primary_entry(r)
        addp = _choose_add_on(entry_src, r)
        stop = _num(r.get("Atr Trailing Stop"))

        primary_entry.append(entry if np.isfinite(entry) else None)
        add_on.append(addp if np.isfinite(addp) else None)
        invalidation.append(stop if np.isfinite(stop) else None)
        entry_source_list.append(entry_src)

        # Buy window status needs current price
        if np.isfinite(cp) and np.isfinite(entry) and entry > 0:
            dist = (cp - entry) / entry
            if dist <= -0.03:
                bw = "EARLY"
            elif -0.03 < dist <= 0.03:
                bw = "IDEAL"
            elif 0.03 < dist <= 0.10:
                bw = "LATE"
            else:
                bw = "OVEREXTENDED"
        else:
            bw = "UNKNOWN"
        buy_window.append(bw)

        # Risk/Reward
        if np.isfinite(entry) and entry > 0 and np.isfinite(stop) and stop > 0:
            rp = (entry - stop) / entry * 100.0
            rp = max(0.0, rp)
        else:
            rp = np.nan

        # Use 90D Gain (%) as reward proxy (already in your sheet)
        rew = _num(r.get("90D Gain (%)"))
        if not np.isfinite(rew):
            # If backtest gain is unavailable, use a conservative proxy:
            # for BUY/STRONG BUY we assume a default +15% upside (your training label).
            sig = str(r.get("Signal") or r.get("Recommendation") or "").strip().upper()
            if sig in {"BUY", "STRONG_BUY", "STRONG BUY"} and np.isfinite(entry) and entry > 0:
                rew = 15.0
            else:
                rew = np.nan

        risk_pct.append(round(rp, 2) if np.isfinite(rp) else None)
        reward_pct.append(round(rew, 2) if np.isfinite(rew) else None)

        if np.isfinite(rp) and rp > 0 and np.isfinite(rew):
            rr = rew / rp
        else:
            rr = np.nan
        rr_ratio.append(round(rr, 2) if np.isfinite(rr) else None)
        rr_for_rank.append(rr if np.isfinite(rr) else 0.0)

        # Decision gate
        exit_now = bool(r.get("Exit Now", False))
        trend_ok = _is_up_trend(r.get("Trend"))
        stage_ok = _is_stage_ok(r.get("Market Stage"))
        buy_confirm = (
            _is_yes(r.get("Model-Driven Strong Buy"))
            or _is_yes(r.get("Model-Driven Buy"))
            or _is_yes(r.get("Rule-Based Buy"))
            or _is_buyish_signal(r.get("Signal"))
        )

        # Strength filter (2 of 3)
        conf = _num(r.get("Confidence Score"), default=np.nan)
        inst = _num(r.get("Institutional Score"), default=np.nan)
        adx_strength = str(r.get("ADX Strength") or "").strip().lower()
        strength_hits = 0
        if np.isfinite(conf) and conf >= 0.65:
            strength_hits += 1
        if np.isfinite(inst) and inst >= 0.70:
            strength_hits += 1
        if "strong" in adx_strength or _num(r.get("Trend_Strength"), default=0.0) >= 0.60:
            strength_hits += 1
        strength_ok = strength_hits >= 2

        # Momentum/support structure (any one)
        structure_ok = (
            _is_yes(r.get("Breakout"))
            or str(r.get("Darvas Signal") or "").strip().upper() == "BUY"
            or _is_yes(r.get("DipReclaim"))
            or _is_yes(r.get("SMC_Breakout"))
        )

        if exit_now:
            fa = "EXIT"
        elif trend_ok and stage_ok and buy_confirm and strength_ok and structure_ok:
            # If overextended, don't chase
            if bw == "OVEREXTENDED":
                fa = "BUY ON DIP"
            else:
                fa = "BUY (Months)"
        elif buy_confirm and not exit_now:
            fa = "HOLD/WATCH"
        else:
            fa = "AVOID/WAIT"
        final_action.append(fa)

        # Position size class (simple + robust)
        sym_vol = int(_num(r.get("Sym Vol Regime"), default=1))
        vix_vol = int(_num(r.get("VIX Vol Regime"), default=1))
        high_vol = (sym_vol >= 2) or (vix_vol >= 2)
        if fa.startswith("BUY"):
            if (not high_vol) and np.isfinite(conf) and conf >= 0.75 and np.isfinite(inst) and inst >= 0.75:
                sc = "FULL"
            elif high_vol:
                sc = "SMALL"
            else:
                sc = "HALF"
        else:
            sc = "-"
        size_class.append(sc)

        # Score for rank
        rr_safe = rr if np.isfinite(rr) else 0.0
        conf_safe = conf if np.isfinite(conf) else 0.0
        inst_safe = inst if np.isfinite(inst) else 0.0
        score_for_rank.append(rr_safe * 0.55 + conf_safe * 0.25 + inst_safe * 0.20)

    out["Primary_Entry_Price"] = primary_entry
    out["Primary_Entry_Source"] = entry_source_list
    out["Add_On_Dip_Price"] = add_on
    out["Invalidation_Level"] = invalidation
    out["Risk_%"] = risk_pct
    out["Reward_%"] = reward_pct
    out["Risk_Reward_Ratio"] = rr_ratio
    out["Buy_Window_Status"] = buy_window
    out["Final_Action"] = final_action
    out["Position_Size_Class"] = size_class

    # Best_Risk_Reward grade (A+/A/B/C/D) using score distribution
    scores = pd.Series(score_for_rank)
    if len(scores) >= 5:
        q80 = float(scores.quantile(0.80))
        q60 = float(scores.quantile(0.60))
        q40 = float(scores.quantile(0.40))
        q20 = float(scores.quantile(0.20))
        q90 = float(scores.quantile(0.90))
    else:
        q90, q80, q60, q40, q20 = 2.0, 1.5, 1.0, 0.7, 0.4

    grades = []
    for s in scores.tolist():
        if s >= q90:
            g = "A+"
        elif s >= q80:
            g = "A"
        elif s >= q60:
            g = "B"
        elif s >= q40:
            g = "C"
        else:
            g = "D"
        grades.append(g)

    out["Best_Risk_Reward"] = grades
    return out

# -----------------------------
# Excel header matching (fix blanks)
# -----------------------------
import unicodedata, re



def add_pre_breakout_tag(
    df_summary: pd.DataFrame,
    *,
    accum_min: int = 7,
    dist_max: int = 3,
    max_below_vwap_pct: float = 0.02,
) -> pd.DataFrame:
    """Tag PRE_BREAKOUT setups (institutional accumulation just below VWAP).

    Criteria (default):
      - ACCUMULATION_DAYS_20D >= accum_min
      - DISTRIBUTION_DAYS_20D <= dist_max
      - Price proxy is within [0%, max_below_vwap_pct] below VWAP

    Price proxy preference order:
      1) Refined Buy Price
      2) PRICE
      3) close / Close

    VWAP preference order:
      1) VWAP Support
      2) VWAP
      3) vwap_support
    """
    try:
        if df_summary is None or df_summary.empty:
            return df_summary

        out = df_summary.copy()

        def _num_series(colname: str):
            if colname not in out.columns:
                return pd.Series([np.nan] * len(out), index=out.index)
            return pd.to_numeric(out[colname], errors="coerce")

        # Choose columns flexibly
        vwap = _num_series("VWAP Support")
        if vwap.isna().all():
            vwap = _num_series("VWAP")
        if vwap.isna().all():
            vwap = _num_series("vwap_support")

        price = _num_series("Refined Buy Price")
        if price.isna().all():
            price = _num_series("PRICE")
        if price.isna().all():
            # try common variants
            price = _num_series("close")
        if price.isna().all():
            price = _num_series("Close")

        accum = _num_series("ACCUMULATION_DAYS_20D")
        dist = _num_series("DISTRIBUTION_DAYS_20D")

        below_vwap_pct = np.where(
            (vwap.notna()) & (price.notna()) & (vwap != 0),
            (vwap - price) / vwap,
            np.nan
        )

        pre_breakout_mask = (
            (accum >= float(accum_min)) &
            (dist <= float(dist_max)) &
            (below_vwap_pct >= 0.0) &
            (below_vwap_pct <= float(max_below_vwap_pct))
        )

        out["PRE_BREAKOUT"] = np.where(pre_breakout_mask, "YES", "")
        out["PRE_BREAKOUT_REASON"] = np.where(
            pre_breakout_mask,
            f"High accumulation + low distribution; within {max_below_vwap_pct*100:.0f}% below VWAP",
            ""
        )
        out["BELOW_VWAP_PCT"] = np.round(below_vwap_pct * 100, 2)

        return out
    except Exception:
        # Never fail the pipeline due to tagging
        df_summary = df_summary.copy()
        if "PRE_BREAKOUT" not in df_summary.columns:
            df_summary["PRE_BREAKOUT"] = ""
        if "PRE_BREAKOUT_REASON" not in df_summary.columns:
            df_summary["PRE_BREAKOUT_REASON"] = ""
        if "BELOW_VWAP_PCT" not in df_summary.columns:
            df_summary["BELOW_VWAP_PCT"] = np.nan
        return df_summary




def fill_missing_output_columns(df_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure user-facing Excel columns are populated (never left blank), even when upstream data is missing.

    This is *post-processing only* — it does not change the core signal logic.
    """
    if df_summary is None or df_summary.empty:
        return df_summary

    out = df_summary.copy()

    def _to_num(s):
        return pd.to_numeric(s, errors="coerce")

    # -------------------------
    # Confidence Score
    # -------------------------
    if "Confidence Score" not in out.columns:
        out["Confidence Score"] = np.nan
    # If missing, backfill from Model Probability (0..1) -> 0..1 score (keep same scale)
    if "Model Probability" in out.columns:
        mp = _to_num(out["Model Probability"])
        cs = _to_num(out["Confidence Score"])
        out["Confidence Score"] = np.where(cs.isna(), mp, cs)

    # Still missing → backfill from Institutional Score & Volume Weight (normalized-ish)
    cs = _to_num(out["Confidence Score"])
    inst = _to_num(out.get("Institutional Score", np.nan))
    vw = _to_num(out.get("Volume Weight", np.nan))
    fallback = 0.6 * inst + 0.4 * vw
    out["Confidence Score"] = np.where(cs.isna(), fallback, cs)
    out["Confidence Score"] = _to_num(out["Confidence Score"]).round(2)

    # -------------------------
    # Sentiment columns
    # -------------------------
    col = "Sentiment Confidence"

    if col not in out.columns:
        out[col] = None

    out[col] = out[col].astype("object")

    if "News Article Count" in out.columns:
        nac = pd.to_numeric(out["News Article Count"], errors="coerce").fillna(0)
        sc_blank = out[col].isna() | (out[col].astype(str).str.strip() == "")

        out.loc[sc_blank & (nac >= 15), col] = "HIGH"
        out.loc[sc_blank & (nac.between(5, 14)), col] = "MED"
        out.loc[sc_blank & (nac.between(1, 4)), col] = "LOW"
        out.loc[sc_blank & (nac == 0), col] = "N/A"
    else:
        out[col] = out[col].fillna("N/A")


    # If Sentiment Confidence is blank, infer a coarse label from News Article Count
    if "News Article Count" in out.columns:
        nac = _to_num(out["News Article Count"]).fillna(0)
        sc = out["Sentiment Confidence"].astype(object)
        sc_blank = sc.isna() | (sc.astype(str).str.strip() == "")
        out.loc[sc_blank & (nac >= 15), "Sentiment Confidence"] = "HIGH"
        out.loc[sc_blank & (nac.between(5, 14)), "Sentiment Confidence"] = "MED"
        out.loc[sc_blank & (nac.between(1, 4)), "Sentiment Confidence"] = "LOW"
        out.loc[sc_blank & (nac == 0), "Sentiment Confidence"] = "N/A"

    # Add Sentiment Label (BULLISH/NEUTRAL/BEARISH) from News Sentiment Score when available
    if "Sentiment Label" not in out.columns:
        out["Sentiment Label"] = None
    if "News Sentiment Score" in out.columns:
        nss = _to_num(out["News Sentiment Score"])
        # thresholds are intentionally modest (AV scores are ~ -1..+1)
        label = np.where(nss >= 0.15, "BULLISH", np.where(nss <= -0.15, "BEARISH", "NEUTRAL"))
        out["Sentiment Label"] = np.where(out["Sentiment Label"].isna() | (out["Sentiment Label"].astype(str).str.strip() == ""),
                                          label, out["Sentiment Label"])
    out["Sentiment Label"] = out["Sentiment Label"].fillna("N/A")

    # -------------------------
    # Trade-plan columns (Final_Action, Reward_%)
    # -------------------------
    if "Final_Action" not in out.columns:
        out["Final_Action"] = None
    fa_blank = out["Final_Action"].isna() | (out["Final_Action"].astype(str).str.strip() == "")
    # backfill from Recommendation if present
    if "Recommendation" in out.columns:
        reco = out["Recommendation"].astype(str).str.upper()
        out.loc[fa_blank & reco.str.contains("STRONG_BUY", na=False), "Final_Action"] = "BUY"
        out.loc[fa_blank & reco.str.contains("BUY", na=False), "Final_Action"] = "WATCH"
        out.loc[fa_blank & reco.str.contains("SELL", na=False), "Final_Action"] = "EXIT"
    out["Final_Action"] = out["Final_Action"].fillna("WATCH")

    if "Reward_%" not in out.columns:
        out["Reward_%"] = np.nan
    rw = _to_num(out["Reward_%"])
    # If we have Long targets and entry, compute reward to Target1
    pe = _to_num(out.get("Primary_Entry_Price", np.nan))
    t1 = _to_num(out.get("LONG_TARGET_1", np.nan))
    computed = np.where(np.isfinite(pe) & np.isfinite(t1) & (pe > 0), (t1 - pe) / pe * 100.0, np.nan)
    out["Reward_%"] = np.where(rw.isna(), computed, rw)
    out["Reward_%"] = _to_num(out["Reward_%"]).round(2)

    # -------------------------
    # Market microstructure / shortability fields
    # -------------------------
    for col, default in [
        ("BID_ASK_SPREAD_PCT", 0.0),
        ("SHORTABLE_FLAG", "N/A"),
        ("BORROW_FEE_PCT", np.nan),
    ]:
        if col not in out.columns:
            out[col] = default
        # Fill blanks
        if isinstance(default, str):
            out[col] = out[col].fillna(default)
            out[col] = out[col].replace("", default)
        else:
            out[col] = _to_num(out[col]).fillna(default)

    # -------------------------
    # Setup/plan columns (never blank)
    # -------------------------
    for col, default in [
        ("Pattern Detected", "NONE"),
        ("LONG_SETUP_TAG", "NONE"),
        ("SHORT_SETUP_TAG", "NONE"),
        ("PRE_BREAKOUT", "NO"),
        ("PRE_BREAKOUT_REASON", "N/A"),
        ("SPIKE_DRIVER", "N/A"),
        ("DROP_DRIVER", "N/A"),
    ]:
        if col not in out.columns:
            out[col] = default
        out[col] = out[col].fillna(default)
        out[col] = out[col].replace("", default)

    # Numeric plan columns: keep NaN if truly not computable, but avoid Excel blank by using 0 for RR ratios
    for col, default in [
        ("LONG_TARGET_1", np.nan),
        ("LONG_TARGET_2", np.nan),
        ("SHORT_TARGET_1", np.nan),
        ("SHORT_TARGET_2", np.nan),
        ("LONG_RR_RATIO", 0.0),
        ("SHORT_RR_RATIO", 0.0),
        ("BELOW_VWAP_PCT", np.nan),
    ]:
        if col not in out.columns:
            out[col] = default
        out[col] = _to_num(out[col])

    return out

def _canon(s: object) -> str:
    """Canonicalize header strings to match template even with unicode/typos."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    # unify hyphen-like chars to "-"
    s = s.replace("–", "-").replace("—", "-").replace("‑", "-").replace("−", "-")
    s = s.strip().lower()
    # collapse whitespace
    s = re.sub(r"\s+", " ", s)
    # remove non-alphanum
    s2 = re.sub(r"[^a-z0-9]+", "", s)
    return s2

def _find_header_row_in_ws(ws, search_rows: int = 80) -> int:
    max_r = min(search_rows, ws.max_row or 1)
    max_c = ws.max_column or 1
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if _canon(ws.cell(r, c).value) in {"symbol", "ticker", "tickers"}:
                return r
    return 1

def _score_sheet_for_headers(ws) -> int:
    """Pick best sheet by counting header hits in first 80 rows."""
    hr = _find_header_row_in_ws(ws)
    max_c = ws.max_column or 1
    hits = 0
    for c in range(1, max_c + 1):
        k = _canon(ws.cell(hr, c).value)
        if k:
            hits += 1
    # require at least a few headers beyond just Symbol
    return hits

def _pick_sheet(wb, preferred_name: str = ""):
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]
    # pick sheet with most header cells on detected header row
    best_ws = wb.active
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        sc = _score_sheet_for_headers(ws)
        if sc > best_score:
            best_score = sc
            best_ws = ws
    return best_ws

def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(header_row, c).value
        k = _canon(v)
        if k:
            m[k] = c
    return m

def _ensure_header(ws, header_row: int, header_map: Dict[str, int], header: str) -> int:
    """Find an existing matching header; otherwise append."""
    k = _canon(header)
    if not k:
        # shouldn't happen
        new_col = (ws.max_column or 1) + 1
        ws.cell(header_row, new_col).value = header
        return new_col

    if k in header_map:
        return header_map[k]

    # Try fuzzy match: sometimes template uses slightly different wording.
    # Example: "modelprobability" vs "modelprob"
    for existing_k, col_idx in header_map.items():
        if existing_k == k:
            return col_idx
        if existing_k.startswith(k) or k.startswith(existing_k):
            return col_idx

    # append new header
    new_col = (ws.max_column or 1) + 1
    ws.cell(header_row, new_col).value = header
    header_map[k] = new_col
    return new_col

def _excel_safe(v):
    """Convert pandas/numpy types to Excel-safe python scalars; NaN/inf -> None."""
    if v is None:
        return None
    # bool is subclass of int; keep it
    if isinstance(v, (bool, str, int)):
        return v
    try:
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating, float)):
            fv = float(v)
            if not np.isfinite(fv):
                return None
            return fv
    except Exception:
        pass
    # pandas Timestamp
    if hasattr(v, "to_pydatetime"):
        try:
            return v.to_pydatetime()
        except Exception:
            pass
    # fallback
    try:
        # treat empty/nan-like strings
        if isinstance(v, str) and v.strip() == "":
            return None
    except Exception:
        pass
    return v


def write_template_excel(df: pd.DataFrame, template_path: Path, out_path: Path, sheet_name: str = "",
                        extra_sheets: Optional[Dict[str, pd.DataFrame]] = None) -> None:
    """Write df columns into the template WITHOUT wiping existing formulas/values.

    Key rules:
      - Never clear cells (clearing wipes formulas and makes many columns appear blank)
      - Never overwrite a cell with None/NaN/inf
      - Force-write a small set of critical columns by exact header match (case-insensitive)
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = _pick_sheet(wb, preferred_name=sheet_name)

    header_row = _find_header_row_in_ws(ws)
    header_map = _build_header_map(ws, header_row)
    data_start = header_row + 1

    # Ensure every df column exists in template headers (append if missing)
    col_index: Dict[str, int] = {}
    for col in df.columns:
        idx = _ensure_header(ws, header_row, header_map, col)
        col_index[col] = idx

    # Build direct header lookup (case-insensitive exact string)
    direct = {}
    for c in range(1, ws.max_column + 1):
        hv = ws.cell(header_row, c).value
        if hv is None:
            continue
        direct[str(hv).strip().lower()] = c

    force_cols = [
        "Volume Pressure",
        "Whether the current DMA is greater than 50 DMA.",
        "Whether the current DMA is greater than 100 DMA.",
        "Whether the current DMA is greater than 150 DMA.",
        "Whether the current DMA is greater than 200.",
        "Whether Weekly chart has got higher up wicks volume.",
        "How much % high Weekly chart is from previous lower volume.",
    ]

    # Write rows (skip blanks so we don't destroy formulas)
    records = df.to_dict(orient="records")
    for i, rowd in enumerate(records):
        r = data_start + i
        for col, idx in col_index.items():
            val = _excel_safe(rowd.get(col))
            if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
                continue
            ws.cell(r, idx).value = val

    # Force-write critical columns into the exact header if it exists (prevents "looks blank" issues)
    for col in force_cols:
        if col not in df.columns:
            continue
        key = col.strip().lower()
        idx = direct.get(key)
        if idx is None:
            idx = _ensure_header(ws, header_row, header_map, col)
        for i in range(len(df)):
            r = data_start + i
            val = _excel_safe(df.iloc[i][col])
            if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
                continue
            ws.cell(r, idx).value = val

    # Create/update extra sheets (TOP_BUYS, TOP_SHORTS, etc.)
    if extra_sheets:
        for sname, sdf in extra_sheets.items():
            try:
                if sname in wb.sheetnames:
                    del wb[sname]
                nws = wb.create_sheet(title=sname)
                # Write headers
                cols = list(sdf.columns)
                for j, c in enumerate(cols, start=1):
                    nws.cell(1, j).value = str(c)
                # Write rows
                for i in range(len(sdf)):
                    for j, c in enumerate(cols, start=1):
                        v = _excel_safe(sdf.iloc[i][c])
                        if v is None:
                            continue
                        nws.cell(2 + i, j).value = v
                # Freeze header
                nws.freeze_panes = "A2"
            except Exception as e:
                print(f"[WARN] failed to write sheet {sname}: {e}")

    wb.save(out_path)



def add_excel_derived_columns(df):
    """Defensive Excel-facing column normalizer.
    Ensures commonly expected display columns exist and are backfilled.
    Never raises; returns df unchanged on any error.
    """
    try:
        import pandas as pd
        if df is None or len(df) == 0:
            return df

        def ensure(col, default=None):
            if col not in df.columns:
                df[col] = default

        ensure("Recommendation", "")
        ensure("Decision Reason", "")
        ensure("Refined Buy Price", None)
        ensure("VWAP Support", None)
        ensure("Trend", "")
        ensure("Institutional Score", 0.0)
        ensure("Confidence Score", 0.0)

        if "FINAL_ACTION" in df.columns:
            df["Recommendation"] = df["Recommendation"].where(
                df["Recommendation"].astype(str).str.len() > 0,
                df["FINAL_ACTION"].astype(str),
            )

        if "FINAL_REASON" in df.columns:
            df["Decision Reason"] = df["Decision Reason"].where(
                df["Decision Reason"].astype(str).str.len() > 0,
                df["FINAL_REASON"].astype(str),
            )

        for src in ["REFINED_BUY_PRICE", "LONG_ENTRY_ZONE_LOW", "BUY_PRICE", "buy_price", "refined_buy_price"]:
            if src in df.columns:
                df["Refined Buy Price"] = df["Refined Buy Price"].fillna(df[src])
                break

        for src in ["VWAP", "vwap"]:
            if src in df.columns:
                df["VWAP Support"] = df["VWAP Support"].fillna(df[src])
                break

        for src in ["CONFIDENCE_SCORE", "confidence_score", "FINAL_CONFIDENCE"]:
            if src in df.columns:
                df["Confidence Score"] = pd.to_numeric(df["Confidence Score"], errors="coerce") \
                    .fillna(pd.to_numeric(df[src], errors="coerce")) \
                    .fillna(0.0)
                break

        for src in ["INSTITUTIONAL_SCORE", "institutional_score"]:
            if src in df.columns:
                df["Institutional Score"] = pd.to_numeric(df["Institutional Score"], errors="coerce") \
                    .fillna(pd.to_numeric(df[src], errors="coerce")) \
                    .fillna(0.0)
                break

        for src in ["TREND_LABEL", "HTF_Trend", "trend"]:
            if src in df.columns:
                df["Trend"] = df["Trend"].where(
                    df["Trend"].astype(str).str.len() > 0,
                    df[src].astype(str),
                )
                break

        return df
    except Exception:
        return df

def main():
    args = parse_args()

    if not CONFIG_SYMBOLS:
        raise SystemExit("config.symbols is empty")

    template_path = HERE / args.template
    out_path = Path(args.excel_out)

    # First pass: base analysis via analyze_symbol_all (same as ddddd_buy.py)
    summary: List[dict] = []
    _RUN_RAW_DF = {}  # symbol -> raw daily df (reuse across passes)
    _RUN_SRC = {}     # symbol -> data source string
    _RUN_ERR = {}     # symbol -> error string
    for symbol in CONFIG_SYMBOLS:
        df_raw, _src, _err = load_symbol_df_info(symbol, args)
        _RUN_RAW_DF[symbol] = df_raw
        _RUN_SRC[symbol] = _src
        _RUN_ERR[symbol] = _err
        if df_raw is None or df_raw.empty:
            print(f"⚠️ {symbol}: no data (fetch/cache failure)")
            continue
        df = ensure_enrichment_schema(compute_indicators(df_raw.copy(), symbol=symbol))
        if df is None or df.empty:
            print(f"⚠️ {symbol}: indicators failed/empty")
            continue

        # res = analyze_symbol_all(symbol)
        res = analyze_symbol(symbol, df_raw=df_raw)
        if not res:
            continue

        # normalize recommendation key (required for legacy logic)
        if "Recommendation" in res and "recommendation" not in res:
            res["recommendation"] = res["Recommendation"]

        # guarantee explainability fields BEFORE append
        res.setdefault("Decision Reason", None)
        res.setdefault("Confidence Grade", None)
        res.setdefault("Expected Holding Period", None)

        # NOW append (this line must be last)
        summary.append(res)

    if not summary:
        raise SystemExit("⚠️ No valid predictions could be generated.")

    # Load QQQ once for relative strength (optional)
    try:
        qqq_df_raw, qqq_src, _qqq_err = load_symbol_df_info("QQQ", args)
        qqq_df = qqq_df_raw if (qqq_df_raw is not None and not qqq_df_raw.empty) else None
        if qqq_df is None:
            print("[INFO] QQQ data unavailable; REL_STRENGTH_20D_VS_QQQ will be NA")
        else:
            print(f"[INFO] QQQ data loaded for RS calc (source={qqq_src})")
    except Exception:
        qqq_df = None

    # Second pass: enrichment
    hit_list, gain_list, days_list = [], [], []
    for res in summary:
        symbol = res.get("Symbol", "")
        try:
            # Reuse the already-fetched daily df from pass-1 to avoid duplicate fetch/parquet reads
            df_raw = _RUN_RAW_DF.get(symbol)
            data_source = _RUN_SRC.get(symbol)
            data_err = _RUN_ERR.get(symbol)
            if df_raw is None or getattr(df_raw, 'empty', True):
                df_raw, data_source, data_err = load_symbol_df_info(symbol, args)
                _RUN_RAW_DF[symbol] = df_raw
                _RUN_SRC[symbol] = data_source
                _RUN_ERR[symbol] = data_err
            res["DATA_SOURCE"] = data_source
            res["DATA_ERROR"] = data_err if data_err else None
            try:
                res["ASOF_DATE"] = pd.to_datetime(df_raw["date"].iloc[-1]) if (df_raw is not None and not df_raw.empty and "date" in df_raw.columns) else None
            except Exception:
                res["ASOF_DATE"] = None
            df = ensure_enrichment_schema(compute_indicators(df_raw.copy(), symbol=symbol))

            # Latest price (needed for Buy_Window_Status / chase checks)
            try:
                res["Current Price"] = round(float(df["close"].iloc[-1]), 2)
            except Exception:
                res["Current Price"] = None

            # -----------------------------------------------------------------
            # NEW SRS FIELDS: DMA / Volume / Momentum / Volatility / Trend health
            # -----------------------------------------------------------------
            common = compute_common_indicators_rls(df, qqq_df=qqq_df)
            # Map into exact requested schema
            for k in [
                "DMA20", "DMA50", "DMA100", "DMA150", "DMA200",
                "PCT_FROM_DMA20", "PCT_FROM_DMA50", "PCT_FROM_DMA200",
                "DMA_STACK", "ABOVE_DMA20", "ABOVE_DMA50", "ABOVE_DMA200",
                "VOL_TODAY", "AVG_VOL_20D", "VOL_SURGE_RATIO", "VOL_TREND_5D",
                "RSI14", "MACD", "MACD_SIGNAL", "MACD_HIST",
                "ATR14", "ATR14_PCT", "GAP_PCT", "REL_STRENGTH_20D_VS_QQQ",
            ]:
                res[k] = common.get(k, None)

            # Execution / shortability (unknown unless you wire IBKR shortability endpoints)
            res.setdefault("BID_ASK_SPREAD_PCT", None)
            res.setdefault("LIQUIDITY_SCORE", None)
            res.setdefault("SHORTABLE_FLAG", None)
            res.setdefault("BORROW_FEE_PCT", None)

            # Liquidity score (uses AVG_VOL_20D + spread if available)
            try:
                spread_pct = float(res.get("BID_ASK_SPREAD_PCT")) if res.get("BID_ASK_SPREAD_PCT") is not None else float("nan")
                res["LIQUIDITY_SCORE"] = float(liquidity_score(float(res.get("AVG_VOL_20D")) if res.get("AVG_VOL_20D") is not None else float("nan"), spread_pct))
            except Exception:
                res["LIQUIDITY_SCORE"] = None

            # -----------------------------------------------------------------
            # LONG / SHORT setups + scores + plans
            # -----------------------------------------------------------------
            long_setup = detect_long_setup(common, df, RANK_CFG)
            short_setup = detect_short_setup(common, df, RANK_CFG)

            res["LONG_SETUP_TAG"] = long_setup
            res["SHORT_SETUP_TAG"] = short_setup

            # Scores
            spread_pct = float(res.get("BID_ASK_SPREAD_PCT")) if res.get("BID_ASK_SPREAD_PCT") is not None else float("nan")
            long_score = score_long(common, RANK_CFG, bid_ask_spread_pct=spread_pct)
            short_score = score_short(
                common,
                RANK_CFG,
                bid_ask_spread_pct=spread_pct,
                shortable_flag=None if res.get("SHORTABLE_FLAG") is None else bool(res.get("SHORTABLE_FLAG")),
                borrow_fee_pct=float(res.get("BORROW_FEE_PCT")) if res.get("BORROW_FEE_PCT") is not None else float("nan"),
            )
            res["LONG_SCORE"] = float(long_score) if np.isfinite(long_score) else None
            res["SHORT_SCORE"] = float(short_score) if np.isfinite(short_score) else None

            # Verdicts
            if long_setup and np.isfinite(long_score) and long_score >= RANK_CFG.long_score_buy_now:
                res["LONG_VERDICT"] = "BUY_NOW"
            elif long_setup and np.isfinite(long_score) and long_score >= RANK_CFG.long_score_wait_dip:
                res["LONG_VERDICT"] = "WAIT_FOR_DIP"
            elif long_setup:
                res["LONG_VERDICT"] = "WATCH"
            else:
                res["LONG_VERDICT"] = "AVOID"

            if short_setup and np.isfinite(short_score) and short_score >= RANK_CFG.short_score_short_now:
                res["SHORT_VERDICT"] = "SHORT_NOW"
            elif short_setup and np.isfinite(short_score) and short_score >= RANK_CFG.short_score_wait_bounce:
                res["SHORT_VERDICT"] = "WAIT_FOR_BOUNCE"
            elif short_setup:
                res["SHORT_VERDICT"] = "WATCH"
            else:
                res["SHORT_VERDICT"] = "AVOID"

            res["SHORT_FEASIBILITY"] = feasibility_label(
                None if res.get("SHORTABLE_FLAG") is None else bool(res.get("SHORTABLE_FLAG")),
                float(res.get("BORROW_FEE_PCT")) if res.get("BORROW_FEE_PCT") is not None else float("nan"),
                RANK_CFG,
            )

            # Plans
            lp = build_long_plan(common, df, long_setup, RANK_CFG)
            sp = build_short_plan(common, df, short_setup, RANK_CFG)
            for k2, v2 in lp.items():
                res[k2] = v2
            for k2, v2 in sp.items():
                res[k2] = v2

            # Explainers
            res["SPIKE_DRIVER"] = lp.get("SPIKE_DRIVER")
            res["DROP_DRIVER"] = sp.get("DROP_DRIVER")

            # --- New DMA alignment columns (interpreting "current DMA" as 20-DMA) ---
            try:
                close_num = pd.to_numeric(df["close"], errors="coerce")
                sma20 = float(close_num.rolling(20).mean().iloc[-1])
                sma50 = float(close_num.rolling(50).mean().iloc[-1])
                sma100 = float(close_num.rolling(100).mean().iloc[-1])
                sma150 = float(close_num.rolling(150).mean().iloc[-1])
                sma200 = float(close_num.rolling(200).mean().iloc[-1])
            except Exception:
                sma20 = sma50 = sma100 = sma150 = sma200 = float("nan")

            res["Whether the current DMA is greater than 50 DMA."] = "✅" if (np.isfinite(sma20) and np.isfinite(sma50) and sma20 > sma50) else "❌"
            res["Whether the current DMA is greater than 100 DMA."] = "✅" if (np.isfinite(sma20) and np.isfinite(sma100) and sma20 > sma100) else "❌"
            res["Whether the current DMA is greater than 150 DMA."] = "✅" if (np.isfinite(sma20) and np.isfinite(sma150) and sma20 > sma150) else "❌"
            res["Whether the current DMA is greater than 200."] = "✅" if (np.isfinite(sma20) and np.isfinite(sma200) and sma20 > sma200) else "❌"

            # --- Weekly chart checks (upper wick + volume, and % distance vs prev lower-volume week high) ---
            try:
                wdf = fetch_data_cached(
                    symbol,
                    bar_spec="5 Y",
                    bar_size="1 week",
                    ttl_minutes=args.cache_ttl_mins,
                    require_today=False,
                    force_refresh=False,
                )
            except Exception:
                wdf = None

            # If weekly fetch fails, fall back to resampling daily into weekly
            if (wdf is None or getattr(wdf, "empty", True)) and df is not None and not df.empty:
                try:
                    wd = df.copy()
                    if "date" in wd.columns:
                        wd["date"] = pd.to_datetime(wd["date"], errors="coerce")
                        wd = wd.dropna(subset=["date"]).set_index("date")
                    else:
                        wd.index = pd.to_datetime(wd.index, errors="coerce")
                    wd = wd.sort_index()
                    wdf = pd.DataFrame({
                        "open": wd["open"].resample("W-FRI").first(),
                        "high": wd["high"].resample("W-FRI").max(),
                        "low": wd["low"].resample("W-FRI").min(),
                        "close": wd["close"].resample("W-FRI").last(),
                        "volume": wd["volume"].resample("W-FRI").sum() if "volume" in wd.columns else np.nan,
                    }).dropna(subset=["open", "high", "low", "close"])
                except Exception:
                    wdf = None

            weekly_hi_wick = _weekly_upper_wick_high_volume(wdf) if wdf is not None and not getattr(wdf, "empty", True) else False
            weekly_pct = _pct_weekly_high_vs_prev_lower_volume_high(wdf) if wdf is not None and not getattr(wdf, "empty", True) else None

            res["Whether Weekly chart has got higher up wicks volume."] = "✅" if weekly_hi_wick else "❌"
            res["How much % high Weekly chart is from previous lower volume."] = weekly_pct

            # EPS flags
            eps_df = fetch_quarterly_eps(symbol)
            eps_flags = eps_growth_flags(eps_df)
            res["EPS Increase 2Q"] = eps_flags["EPS Increase 2Q"]
            res["EPS Increase 3Q"] = eps_flags["EPS Increase 3Q"]
            res["EPS Increase 4Q"] = eps_flags["EPS Increase 4Q"]

            # News sentiment
            sent = fetch_market_sentiment(symbol)
            res["News Sentiment Score"] = round(float(sent.get("news_sentiment_score") or 0.0), 4)
            res["News Positive Ratio"] = round(float(sent.get("news_positive_ratio") or 0.0), 4)
            res["News Article Count"] = int(sent.get("news_article_count") or 0)
            conf = sent.get("sentiment_confidence")
            if conf is None or conf == "":
                n_art = int(sent.get("news_article_count") or 0)
                if n_art >= 15:
                    conf = "HIGH"
                elif n_art >= 5:
                    conf = "MED"
                elif n_art > 0:
                    conf = "LOW"
                else:
                    conf = None
            res["Sentiment Confidence"] = conf

            # Price-action signals
            df = compute_upward_trend(df)
            signal_score, signal_count, pa_reco = compute_signal_score(df)
            res["Signal Score"] = round(float(signal_score), 2)
            res["Signal Count"] = int(signal_count)
            res["Signal"] = pa_reco

            # Pattern booleans
            smc = bool(detect_smc_accumulation_breakout(df))
            mean_rev = bool(detect_mean_reversion_buy(df))
            res["SMC_Breakout"] = smc
            res["Mean_Reversion"] = mean_rev
            res["Bullish_Engulfing"] = bool(detect_bullish_engulfing(df))
            res["Hammer"] = bool(detect_hammer(df))
            res["Trend_Strength"] = int(df.iloc[-1].get("trend_strength", 0))

            # Rich pattern list (includes multi-bottom if detected) + Wyckoff-style stage/substage
            try:
                pats = compute_patterns(df, res)
                res["Pattern_List"] = ", ".join(pats) if pats else "NONE"
                res["Primary_Pattern"] = pats[0] if pats else "NONE"
            except Exception:
                res["Pattern_List"] = "NONE"
                res["Primary_Pattern"] = "NONE"

            try:
                stg, sub = compute_market_stage_substage(df)
                res["Market Stage"] = stg
                res["Market Sub-Stage"] = sub
            except Exception:
                res["Market Stage"] = "Neutral/Transition"
                res["Market Sub-Stage"] = "NEUTRAL_CHOP"

            # Regimes / volume pressure
            res["Sym Vol Regime"] = int(df.iloc[-1].get("sym_vol_regime", 0))
            res["VIX Vol Regime"] = int(df.iloc[-1].get("VIX_vol_regime", 0))
            res["Volume Pressure"] = float(df.iloc[-1].get("volume_pressure", 0.0))

            # Snapshot
            snap = _compute_indicator_snap(df)
            # _compute_indicator_snap already computes EMA_Uptrend / EMA21_Slope / ADX_Strength safely.
            res["MACD Cross"] = "✅" if snap.get("MACD_Crossover", False) else "❌"
            rsi14 = snap.get("RSI_14")
            res["RSI"] = round(rsi14, 1) if rsi14 is not None and np.isfinite(rsi14) else None
            res["RSI State"] = snap.get("RSI_State", "N/A")
            res["OBV Trend"] = snap.get("OBV_Trend", "N/A")
            res["At BB Lower"] = "Y" if snap.get("At_BB_Lower", False) else ""

            # Rule-based SMA cross + confidence threshold
            df["SMA_20"] = df["close"].rolling(window=20).mean()
            df["SMA_50"] = df["close"].rolling(window=50).mean()
            df["SMA_cross_signal"] = (df["SMA_20"] > df["SMA_50"]) & (df["SMA_20"].shift(1) <= df["SMA_50"].shift(1))
            conf = float(res.get("Confidence Score", 0.0) or 0.0)
            rule_based_buy = bool(df["SMA_cross_signal"].iloc[-1]) and conf > BUY_THRESH
            res["Rule-Based Buy"] = "✅" if rule_based_buy else "❌"

            # Model proba + band
            prob = _predict_proba_for_last_row(symbol, df)
            res["Model Probability"] = round(float(prob), 2)
            res["Model-Driven Buy"] = "✅" if prob >= BUY_THRESH else "❌"
            res["Model-Driven Strong Buy"] = "✅" if prob >= STRONG_BUY_THRESH else "❌"
            res["Confidence Band"] = get_confidence_band(prob)

            # Exit signals (entry = Refined Buy Price)
            entry = res.get("Refined Buy Price", None)
            try:
                entry_price = float(entry) if entry is not None and not (isinstance(entry, float) and np.isnan(entry)) else float(df["close"].iloc[-1])
            except Exception:
                entry_price = float(df["close"].iloc[-1])

            exit_info = compute_exit_signals(df, entry_price=entry_price)
            res["Exit Now"] = bool(exit_info.get("ExitNow", False))
            res["Atr Trailing Stop"] = exit_info.get("AtrTrailingStop", None)
            res["Exit Reasons"] = exit_info.get("ExitReasons", "")

            # Safety
            if bool(res.get("Exit Now", False)):
                res["Signal"] = "HOLD (Exit risk)"
                res["Model-Driven Buy"] = "❌"
                res["Model-Driven Strong Buy"] = "❌"
                res["Confidence Band"] = "NO TRADE"

            # Tech fallback
            if prob < BUY_THRESH:
                tech_score = _tech_fallback_score(snap, df, smc, mean_rev)
                res["Tech Fallback Score"] = tech_score
                # ❌ DO NOT upgrade to BUY
            else:
                res["Tech Fallback Score"] = 0.0

            # Backtest (90D)
            hit, gain, days_to_peak = evaluate_backtest_accuracy(symbol, df, entry_price, gain_thresh=0.04, use_close=True)

        except Exception as e:
            print(f"⚠️ {symbol}: enrichment/backtest failed ({e})")
            hit, gain, days_to_peak = False, 0.0, -1
            res.setdefault("Model Probability", 0.0)
            res.setdefault("Model-Driven Buy", "❌")
            res.setdefault("Model-Driven Strong Buy", "❌")
            res.setdefault("Confidence Band", "WATCH")

        hit_list.append("✅" if hit else "❌")
        gain_list.append(round(float(gain), 2) if np.isfinite(float(gain)) else None)
        days_list.append(int(days_to_peak) if isinstance(days_to_peak, int) and days_to_peak >= 0 else "N/A")

    df_summary = pd.DataFrame(summary)

    # Add derived user-friendly columns (Breakout/Undervalued/Reversal/Patterns/DipReclaim)
    df_summary = add_excel_derived_columns(df_summary)

    # Add actionable trade-management columns
    df_summary = add_trade_management_columns(df_summary)

    # Tag PRE_BREAKOUT setups (coiled spring just below VWAP)
    df_summary = add_pre_breakout_tag(df_summary)
    # Final pass: backfill any user-facing output columns so Excel never has blanks
    df_summary = fill_missing_output_columns(df_summary)


    # Ensure Candle Entry columns exist
    for col in [
        "Candle Entry 2w", "Candle Entry 4w", "Candle Entry 6w", "Candle Entry 8w",
        "Candle Entry 12w", "Candle Entry 18w", "Candle Entry 30w"
    ]:
        if col not in df_summary.columns:
            df_summary[col] = None

    # Price Reversal aggregate
    rev_cols = ["Mean_Reversion", "Bullish_Engulfing", "Hammer"]
    for c in rev_cols:
        if c not in df_summary.columns:
            df_summary[c] = False
    has_rev = df_summary[rev_cols].fillna(False).astype(bool).any(axis=1)
    df_summary["Price Reversal"] = np.where(has_rev, "✅", "❌")

    # 90D results
    df_summary["90D Hit"] = hit_list
    df_summary["90D Gain (%)"] = gain_list
    df_summary["Days to Peak"] = days_list

    # Sort
    if "Model Probability" in df_summary.columns:
        df_summary.sort_values(by="Model Probability", ascending=False, inplace=True)

    columns_to_display = [
        # Action & execution
        "Final_Action", "Best_Risk_Reward", "Buy_Window_Status", "Position_Size_Class",
        "Primary_Entry_Price", "Primary_Entry_Source", "Add_On_Dip_Price", "Invalidation_Level",
        "Risk_%", "Reward_%", "Risk_Reward_Ratio",

        # Price context
        "Current Price",
        "DATA_SOURCE", "ASOF_DATE",
        "DMA20", "DMA50", "DMA100", "DMA150", "DMA200",
        "PCT_FROM_DMA20", "PCT_FROM_DMA50", "PCT_FROM_DMA200",
        "DMA_STACK", "ABOVE_DMA20", "ABOVE_DMA50", "ABOVE_DMA200",
        "VOL_TODAY", "AVG_VOL_20D", "VOL_SURGE_RATIO", "VOL_TREND_5D",
        "RSI14", "MACD", "MACD_SIGNAL", "MACD_HIST", "ATR14", "ATR14_PCT", "GAP_PCT",
        "REL_STRENGTH_20D_VS_QQQ",
        "BID_ASK_SPREAD_PCT", "LIQUIDITY_SCORE", "SHORTABLE_FLAG", "BORROW_FEE_PCT",
        "LONG_SCORE", "LONG_SETUP_TAG", "LONG_VERDICT", "LONG_ENTRY_ZONE", "LONG_INVALIDATION", "LONG_TARGET_1", "LONG_TARGET_2",
        "SHORT_SCORE", "SHORT_SETUP_TAG", "SHORT_VERDICT", "SHORT_ENTRY_ZONE", "SHORT_INVALIDATION", "SHORT_TARGET_1", "SHORT_TARGET_2",
        "SHORT_FEASIBILITY", "SPIKE_DRIVER", "DROP_DRIVER",
        "Whether the current DMA is greater than 50 DMA.",
        "Whether the current DMA is greater than 100 DMA.",
        "Whether the current DMA is greater than 150 DMA.",
        "Whether the current DMA is greater than 200.",
        "Whether Weekly chart has got higher up wicks volume.",
        "How much % high Weekly chart is from previous lower volume.",

        "Symbol", "Refined Buy Price", "Candle Entry 2w", "Candle Entry 4w",
        "Candle Entry 6w", "Candle Entry 8w", "Candle Entry 12w", "Candle Entry 18w", "Candle Entry 30w",
        "VWAP Support", "ADX", "Institutional Score",
        "Volume Weight", "Confidence Score", "Sector Correlation",
        "Trend", "Recommendation", "Decision Reason", "Confidence Grade", "Expected Holding Period",
        "Momentum Recommendation", "Momentum Decision Reason", "Momentum Confidence Grade", "Momentum Expected Holding Period", "Darvas Breakout %", "Darvas Signal",
        "Rule-Based Buy", "Model-Driven Buy", "Model-Driven Strong Buy", "Model Probability",
        "ML Entry Target", "ML Entry Mode", "ML Entry Bias ATR",
        "Confidence Band", "Tech Fallback Score", "Signal",
                "Breakout", "Undervalued", "Trend Reversal", "Pattern Detected", "DipReclaim",
        "90D Hit", "90D Gain (%)", "Days to Peak",
        "EMA Uptrend", "EMA21 Slope", "ADX Strength", "MACD Cross", "RSI", "RSI State",
        "OBV Trend", "At BB Lower",
        "Volume Surge", "Near Support", "Signal Score",
        "Price Reversal",
        "SMC_Breakout", "Mean_Reversion", "Bullish_Engulfing", "Hammer", "Trend_Strength",
        "Market Stage", "Market Sub-Stage",
        "Sym Vol Regime", "VIX Vol Regime", "Volume Pressure",
        "Exit Now", "Atr Trailing Stop", "Exit Reasons",
        "EPS Increase 2Q", "EPS Increase 3Q", "EPS Increase 4Q",
        "News Sentiment Score", "News Positive Ratio", "News Article Count", "Sentiment Confidence",
    ]

    # Ensure optional cols exist
    for col in ["Volume Surge", "Near Support", "Signal Score"]:
        if col not in df_summary.columns:
            df_summary[col] = None

    # Build output frame in that order; create any missing display cols as None so Excel gets filled
    for col in columns_to_display:
        if col not in df_summary.columns:
            df_summary[col] = None
    out_df = df_summary[columns_to_display].copy()

    # Replace NaN/inf with None before writing
    out_df = out_df.replace([np.inf, -np.inf], np.nan)
    out_df = out_df.where(pd.notna(out_df), None)

    # Build TOP_BUYS / TOP_SHORTS sheets
    def _top_sheet(df_all: pd.DataFrame, score_col: str, setup_col: str, verdict_col: str, top_n: int = 20) -> pd.DataFrame:
        cols = [
            "Symbol", "Sector", "Current Price",
            score_col, setup_col, verdict_col,
            "LONG_ENTRY_ZONE" if score_col == "LONG_SCORE" else "SHORT_ENTRY_ZONE",
            "LONG_INVALIDATION" if score_col == "LONG_SCORE" else "SHORT_INVALIDATION",
            "LONG_TARGET_1" if score_col == "LONG_SCORE" else "SHORT_TARGET_1",
            "LONG_TARGET_2" if score_col == "LONG_SCORE" else "SHORT_TARGET_2",
            "DMA20", "DMA50", "DMA200", "DMA_STACK",
            "VOL_SURGE_RATIO", "RSI14", "MACD_HIST", "ATR14_PCT",
            "REL_STRENGTH_20D_VS_QQQ", "LIQUIDITY_SCORE", "DATA_SOURCE",
            "SPIKE_DRIVER" if score_col == "LONG_SCORE" else "DROP_DRIVER",
        ]
        if score_col == "SHORT_SCORE":
            cols.insert(cols.index("LIQUIDITY_SCORE") + 1, "SHORT_FEASIBILITY")
            cols.insert(cols.index("SHORT_FEASIBILITY") + 1, "BORROW_FEE_PCT")

        work = df_all.copy()
        work = work[work[setup_col].notna()]
        work = work[work[verdict_col].isin(["BUY_NOW", "WAIT_FOR_DIP", "WATCH"]) if score_col == "LONG_SCORE" else work[verdict_col].isin(["SHORT_NOW", "WAIT_FOR_BOUNCE", "WATCH"]) ]
        work = work.sort_values(by=score_col, ascending=False, na_position="last")
        for c in cols:
            if c not in work.columns:
                work[c] = None
        return work[cols].head(top_n)

    top_buys = _top_sheet(df_summary, "LONG_SCORE", "LONG_SETUP_TAG", "LONG_VERDICT", top_n=TOP_N)
    top_shorts = _top_sheet(df_summary, "SHORT_SCORE", "SHORT_SETUP_TAG", "SHORT_VERDICT", top_n=TOP_N)

    write_template_excel(
        out_df,
        template_path=template_path,
        out_path=out_path,
        sheet_name=args.sheet,
        extra_sheets={"TOP_BUYS": top_buys, "TOP_SHORTS": top_shorts},
    )
    print(f"\n📊 Excel saved → {out_path}")


def _num(x, default=np.nan):
    try:
        v = float(x)
        return v if np.isfinite(v) else default
    except Exception:
        return default

def _yes_no(flag: bool) -> str:
    return 'Y' if bool(flag) else 'N'

if __name__ == "__main__":
    main()
