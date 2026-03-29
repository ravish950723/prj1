"""buy_srs_update_excel.py

Purpose:
  Update your existing Excel output by adding + populating the requested SRS columns
  (DMA/Volume/VWAP/ATR% + LONG/SHORT scores and trade plans).

Key design:
  - NEVER clears cells.
  - Updates rows by matching the Symbol value (does not assume row order).
  - If data cannot be fetched/computed, writes "N/A" (or 0 for count/score fields)
    so the column is never blank.

Run:
  python buy_srs_update_excel.py --in predictions_summary.xlsx --out predictions_summary_out.xlsx

Notes:
  - Uses fetching.fetch_data_daily_with_fallback() which returns (df, source, err).
  - Uses rank_long_short.py functions + rank_config.json.
"""

import argparse
import logging
import math
import os
import sys
import time
from dataclasses import asdict
from datetime import datetime
from typing import Dict, Any, Optional, Tuple
from compute import  compute_patterns_full , patterns_to_strings

def _is_yes(v) -> bool:
    try:
        return str(v).strip().upper() in {'Y','YES','TRUE','1','✅'}
    except Exception:
        return False



from upward import detect_smc_accumulation_breakout, detect_mean_reversion_buy, detect_bullish_engulfing
from upward import detect_hammer, compute_upward_trend
from backtest import _to_float
try:
    from darvas import darvas_box_signal
except Exception:
    darvas_box_signal = None
from institutional_investor import score_institutional_investor
from rank_long_short import   feasibility_label


import numpy as np
import pandas as pd
import openpyxl

import math

from config import symbols as CONFIG_SYMBOLS  # expected in your project

from fetching import fetch_data_daily_with_fallback
from rank_long_short import (
    load_rank_config,
    RankConfig,
    compute_common_indicators_rls,
    detect_long_setup,
    detect_short_setup,
    score_long,
    score_short,
    build_long_plan,
    build_short_plan,
    explain_spike_drop,
    verdicts,
)


REQUIRED_COLS = []  # Will be replaced at runtime with the actual Excel headers + required computed fields.

DEFAULTS: Dict[str, Any] = {
    "DMA20": "N/A", "DMA50": "N/A", "DMA200": "N/A",
    "PCT_FROM_DMA50": "N/A", "PCT_FROM_DMA200": "N/A",
    "VOL_TODAY": "N/A", "AVG_VOL_20D": "N/A", "VOL_SURGE_RATIO": "N/A",
    "VWAP": "N/A", "VWAP_DISTANCE_PCT": "N/A",
    "ATR14_PCT": "N/A",
    "DISTRIBUTION_DAYS_20D": 0, "ACCUMULATION_DAYS_20D": 0,
    "LONG_SETUP_TAG": "N/A", "LONG_SCORE": 0.0, "LONG_VERDICT": "AVOID",
    "SHORT_SETUP_TAG": "N/A", "SHORT_SCORE": 0.0, "SHORT_VERDICT": "AVOID",
    "FINAL_ACTION": "AVOID", "CONFIDENCE_SCORE": 0.0,
    "LONG_ENTRY_ZONE_LOW": "N/A", "LONG_ENTRY_ZONE_HIGH": "N/A", "LONG_INVALIDATION": "N/A",
    "LONG_TARGET_1": "N/A", "LONG_TARGET_2": "N/A", "LONG_RR_RATIO": "N/A",
    "SHORT_ENTRY_ZONE_LOW": "N/A", "SHORT_ENTRY_ZONE_HIGH": "N/A", "SHORT_INVALIDATION": "N/A",
    "SHORT_TARGET_1": "N/A", "SHORT_TARGET_2": "N/A", "SHORT_RR_RATIO": "N/A",
    "DATA_SOURCE": "ERROR", "SPIKE_DRIVER": "N/A", "DROP_DRIVER": "N/A",
    "Pattern_List": "NONE", "Primary_Pattern": "NONE",
    "Market Stage": "Neutral/Transition", "Market Sub-Stage": "NEUTRAL_CHOP",
}


# ---- Column + feature completion helpers ------------------------------------

# These are the columns that exist in your Excel template but were not being written
# because they were not included in REQUIRED_COLS. We always ensure they are present
# and filled (at least with "N/A") so there are no blank cells.
TEMPLATE_COLUMNS_MINIMUM = [
    # Core technical/state fields
    "RSI", "RSI State", "ADX", "ADX Strength", "EMA Uptrend", "EMA21 Slope",
    "OBV Trend", "At BB Lower", "Breakout", "Undervalued", "Rule-Based Buy",
    "Atr Trailing Stop", "Near Support", "Volume Surge",
    # Trade math
    "Risk_%", "Reward_%", "Days to Peak", "Exit Reasons",
    # Sentiment placeholders
    "Sentiment Label", "Sentiment Confidence",
    # Explanations / drivers
    "SPIKE_DRIVER", "DROP_DRIVER", "PRE_BREAKOUT", "PRE_BREAKOUT_REASON",
]

def _rsi_state(rsi: float) -> str:
    if not np.isfinite(rsi):
        return "N/A"
    if rsi >= 70:
        return "OVERBOUGHT"
    if rsi <= 30:
        return "OVERSOLD"
    if rsi >= 55:
        return "BULLISH"
    if rsi <= 45:
        return "BEARISH"
    return "NEUTRAL"

def _adx_strength(adx: float) -> str:
    if not np.isfinite(adx):
        return "N/A"
    if adx >= 30:
        return "STRONG"
    if adx >= 20:
        return "MODERATE"
    return "WEAK"

def _safe_float(x):
    try:
        v = float(x)
        return v if np.isfinite(v) else np.nan
    except Exception:
        return np.nan

def _compute_basic_ta_fallback(df: pd.DataFrame) -> Dict[str, Any]:
    """Compute a light set of indicators using only pandas/numpy.
    This is used when upstream functions didn't populate fields.
    Expects normalized OHLCV with columns: open/high/low/close/volume.
    """
    out: Dict[str, Any] = {}
    if df is None or df.empty:
        return out
    c = df["close"].astype(float)
    h = df["high"].astype(float)
    l = df["low"].astype(float)
    v = df.get("volume", pd.Series(index=df.index, data=np.nan)).astype(float)

    # RSI(14)
    delta = c.diff()
    gain = delta.clip(lower=0).rolling(14).mean()
    loss = (-delta.clip(upper=0)).rolling(14).mean()
    rs = gain / loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    out["RSI"] = float(rsi.iloc[-1])

    # ATR(14)
    prev_close = c.shift(1)
    tr = pd.concat([(h - l), (h - prev_close).abs(), (l - prev_close).abs()], axis=1).max(axis=1)
    atr14 = tr.rolling(14).mean()
    out["ATR14"] = float(atr14.iloc[-1])
    out["ATR14_PCT"] = float((atr14.iloc[-1] / c.iloc[-1]) * 100.0) if c.iloc[-1] else "N/A"

    # EMA21 slope (approx)
    ema21 = c.ewm(span=21, adjust=False).mean()
    if len(ema21) >= 6:
        out["EMA21 Slope"] = float((ema21.iloc[-1] - ema21.iloc[-6]) / 5.0)
    else:
        out["EMA21 Slope"] = "N/A"

    # EMA uptrend heuristic: EMA20 > EMA50
    ema20 = c.ewm(span=20, adjust=False).mean()
    ema50 = c.ewm(span=50, adjust=False).mean()
    out["EMA Uptrend"] = "Y" if float(ema20.iloc[-1]) > float(ema50.iloc[-1]) else ""

    # Bollinger lower (20,2)
    mid = c.rolling(20).mean()
    sd = c.rolling(20).std()
    lower = mid - 2 * sd
    out["At BB Lower"] = "Y" if float(c.iloc[-1]) <= float(lower.iloc[-1]) else ""

    # OBV trend: slope of rolling OBV
    obv = (np.sign(c.diff()).fillna(0) * v.fillna(0)).cumsum()
    if len(obv) >= 6:
        out["OBV Trend"] = "UP" if float(obv.iloc[-1] - obv.iloc[-6]) > 0 else "DOWN"
    else:
        out["OBV Trend"] = "N/A"

    # ADX (very lightweight approximation using directional movement)
    # Not as accurate as TA-lib, but consistent and fills the column.
    up_move = h.diff()
    down_move = -l.diff()
    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
    tr14 = tr.rolling(14).sum()
    plus_di = 100 * pd.Series(plus_dm, index=df.index).rolling(14).sum() / tr14.replace(0, np.nan)
    minus_di = 100 * pd.Series(minus_dm, index=df.index).rolling(14).sum() / tr14.replace(0, np.nan)
    dx = (100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan))
    adx = dx.rolling(14).mean()
    out["ADX"] = float(adx.iloc[-1]) if np.isfinite(adx.iloc[-1]) else "N/A"

    # Volume surge ratio if missing
    if v.notna().any():
        avg20 = v.rolling(20).mean()
        out["AVG_VOL_20D"] = float(avg20.iloc[-1]) if np.isfinite(avg20.iloc[-1]) else "N/A"
        out["VOL_TODAY"] = float(v.iloc[-1]) if np.isfinite(v.iloc[-1]) else "N/A"
        if np.isfinite(avg20.iloc[-1]) and avg20.iloc[-1] != 0:
            out["VOL_SURGE_RATIO"] = float(v.iloc[-1] / avg20.iloc[-1])
        else:
            out["VOL_SURGE_RATIO"] = "N/A"

    return out

def fill_template_fields_best_effort(out: Dict[str, Any], df: pd.DataFrame) -> Dict[str, Any]:
    """Ensure commonly-empty template fields get filled for every symbol."""
    # 1) If upstream didn't produce RSI/ATR/ADX/etc, compute fallback
    needs_any = any(out.get(k) in (None, "", "N/A", "NA") for k in ["RSI", "ADX", "ATR14"])
    if needs_any:
        out.update(_compute_basic_ta_fallback(df))

    # 2) RSI State + ADX Strength
    rsi = _safe_float(out.get("RSI"))
    adx = _safe_float(out.get("ADX"))
    out["RSI State"] = out.get("RSI State") or _rsi_state(rsi)
    out["ADX Strength"] = out.get("ADX Strength") or _adx_strength(adx)

    # 3) Breakout / Undervalued / Rule-Based Buy heuristics
    # Breakout: price above 20d high and volume surge >= 1.5
    try:
        c = df["close"].astype(float)
        high20 = c.rolling(20).max().iloc[-2]  # prior 20d high
        price = float(c.iloc[-1])
        vs = _safe_float(out.get("VOL_SURGE_RATIO"))
        if np.isfinite(high20) and np.isfinite(price) and price > high20 and (not np.isfinite(vs) or vs >= 1.5):
            out["Breakout"] = out.get("Breakout") or "Y"
        else:
            out["Breakout"] = out.get("Breakout") or ""
    except Exception:
        out["Breakout"] = out.get("Breakout") or "N/A"

    # Undervalued: trading >=10% below DMA200 (if present)
    pct_dma200 = _safe_float(out.get("PCT_FROM_DMA200"))
    if np.isfinite(pct_dma200):
        out["Undervalued"] = out.get("Undervalued") or ("Y" if pct_dma200 <= -10 else "")
    else:
        out["Undervalued"] = out.get("Undervalued") or "N/A"

    # Rule-Based Buy: (RSI bullish/oversold) AND (EMA uptrend or at BB lower)
    rs_state = str(out.get("RSI State") or "")
    ema_up = str(out.get("EMA Uptrend") or "")
    at_bb = str(out.get("At BB Lower") or "")
    if out.get("Rule-Based Buy") in (None, "", "N/A", "NA"):
        cond = (rs_state in {"BULLISH", "OVERSOLD"}) and (ema_up == "Y" or at_bb == "Y")
        out["Rule-Based Buy"] = "Y" if cond else ""

    # 4) ATR trailing stop (simple)
    if out.get("Atr Trailing Stop") in (None, "", "N/A", "NA"):
        price = _safe_float(out.get("Current Price"))
        atr = _safe_float(out.get("ATR14"))
        if np.isfinite(price) and np.isfinite(atr):
            out["Atr Trailing Stop"] = round(price - 3.0 * atr, 4)
        else:
            out["Atr Trailing Stop"] = "N/A"

    # 5) Risk_% and Reward_% if missing and we have plan levels
    if out.get("Risk_%") in (None, "", "N/A", "NA"):
        entry = _safe_float(out.get("Primary_Entry_Price") or out.get("LONG_ENTRY_ZONE_LOW") or out.get("Current Price"))
        inv = _safe_float(out.get("Invalidation_Level") or out.get("LONG_INVALIDATION"))
        if np.isfinite(entry) and np.isfinite(inv) and entry != 0:
            out["Risk_%"] = round(abs(entry - inv) / entry * 100.0, 2)
        else:
            out["Risk_%"] = "N/A"

    if out.get("Reward_%") in (None, "", "N/A", "NA"):
        entry = _safe_float(out.get("Primary_Entry_Price") or out.get("LONG_ENTRY_ZONE_LOW") or out.get("Current Price"))
        tgt = _safe_float(out.get("LONG_TARGET_1") or out.get("SHORT_TARGET_1"))
        if np.isfinite(entry) and np.isfinite(tgt) and entry != 0:
            out["Reward_%"] = round(abs(tgt - entry) / entry * 100.0, 2)
        else:
            out["Reward_%"] = "N/A"

    # 6) Sentiment + exit reasons defaults if empty
    out["Sentiment Label"] = out.get("Sentiment Label") or "N/A"
    out["Sentiment Confidence"] = out.get("Sentiment Confidence") or "N/A"
    out["Exit Reasons"] = out.get("Exit Reasons") or "N/A"
    out["Days to Peak"] = out.get("Days to Peak") or "N/A"

    # 7) Pre-breakout placeholders
    out["PRE_BREAKOUT"] = out.get("PRE_BREAKOUT") or ""
    out["PRE_BREAKOUT_REASON"] = out.get("PRE_BREAKOUT_REASON") or ""

    return out

def get_excel_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        headers.append(str(cell.value).strip())
    return [h for h in headers if h]

def _rsi_state(rsi_value) -> str:
    """
    Convert RSI into a human readable state.
    """
    try:
        r = float(rsi_value)
        if math.isnan(r):
            return "N/A"
    except Exception:
        return "N/A"

    if r >= 70:
        return "OVERBOUGHT"
    if r <= 30:
        return "OVERSOLD"
    return "NEUTRAL"


def _to_float_safe(x, d: float = 0.0) -> float:
    """Robust float conversion for Excel/text placeholders like 'N/A', including non-breaking spaces."""
    try:
        if isinstance(x, str):
            x = x.replace("\u00A0", " ").strip()
        x = _nz(x, d)
        return float(x)
    except Exception:
        return float(d)


def _bb_width_pct(close: pd.Series, window: int = 20, n_std: float = 2.0) -> float:
    """Bollinger Band width as % of mid band (latest)."""
    try:
        mid = close.rolling(window).mean()
        sd = close.rolling(window).std(ddof=0)
        upper = mid + n_std * sd
        lower = mid - n_std * sd
        w = (upper - lower) / mid * 100.0
        val = float(w.iloc[-1])
        return val
    except Exception:
        return float("nan")


def _find_local_extrema(series: pd.Series, kind: str = "min", lookback: int = 120):
    """Return indices of simple local minima/maxima within lookback window."""
    s = series.dropna()
    if len(s) < 5:
        return []
    s = s.iloc[-lookback:]
    idxs = []
    vals = s.values
    for i in range(1, len(vals) - 1):
        if kind == "min":
            if vals[i] <= vals[i - 1] and vals[i] <= vals[i + 1]:
                idxs.append(s.index[i])
        else:
            if vals[i] >= vals[i - 1] and vals[i] >= vals[i + 1]:
                idxs.append(s.index[i])
    return idxs


def _cluster_levels(df: pd.DataFrame, idxs, col: str, tol_pct: float = 2.0, min_sep_days: int = 8):
    """Cluster extrema levels within tolerance; return best cluster (count, level)."""
    if not idxs:
        return (0, None, [])
    levels = []
    for ix in idxs:
        try:
            levels.append((ix, float(df.loc[ix, col])))
        except Exception:
            continue
    if not levels:
        return (0, None, [])
    # sort by date
    levels.sort(key=lambda x: x[0])
    best = (0, None, [])
    for i in range(len(levels)):
        base_ix, base_lvl = levels[i]
        cluster = [(base_ix, base_lvl)]
        last_ix = base_ix
        for j in range(i + 1, len(levels)):
            ix, lvl = levels[j]
            # spacing constraint (avoid counting same swing)
            try:
                if (ix - last_ix).days < min_sep_days:
                    continue
            except Exception:
                pass
            if base_lvl == 0:
                continue
            if abs(lvl - base_lvl) / abs(base_lvl) * 100.0 <= tol_pct:
                cluster.append((ix, lvl))
                last_ix = ix
        if len(cluster) > best[0]:
            best = (len(cluster), base_lvl, cluster)
    return best


def detect_price_patterns(df: pd.DataFrame, common: Dict[str, Any], cfg) -> Dict[str, Any]:
    """Lightweight pattern detector for daily bars. Returns Pattern_List + Primary_Pattern."""
    out = {
        "Pattern_List": "N/A",
        "Primary_Pattern": "N/A",
        "Pattern Detected": "N/A",
    }
    try:
        if df is None or df.empty or len(df) < 80:
            return out

        close = df["Close"].astype(float)
        high = df["High"].astype(float)
        low = df["Low"].astype(float)

        patterns = []

        # --- Triple/Double Bottom & Top (swing clustering)
        lows_idx = _find_local_extrema(low, "min", lookback=180)
        highs_idx = _find_local_extrema(high, "max", lookback=180)

        c_lo, lvl_lo, _ = _cluster_levels(df, lows_idx, "Low", tol_pct=2.0, min_sep_days=10)
        c_hi, lvl_hi, _ = _cluster_levels(df, highs_idx, "High", tol_pct=2.0, min_sep_days=10)

        if c_lo >= 3:
            patterns.append("Triple Bottom")
        elif c_lo == 2:
            patterns.append("Double Bottom")

        if c_hi >= 3:
            patterns.append("Triple Top")
        elif c_hi == 2:
            patterns.append("Double Top")

        # --- Breakout (Darvas / Range breakout proxy)
        price = float(common.get("PRICE", close.iloc[-1]))
        vol_surge = _to_float_safe(common.get("VOL_SURGE_RATIO", common.get("VOL_SURGE_RATIO", "N/A")), 0.0)
        hi20 = float(high.rolling(20).max().iloc[-2]) if len(high) >= 21 else float("nan")
        if np.isfinite(hi20) and price > hi20 and vol_surge >= getattr(cfg, "vol_surge_breakout", 1.25):
            patterns.append("Range Breakout")
            patterns.append("Darvas Box Breakout")

        # --- VWAP reclaim / pullback proxy
        vwap = _to_float_safe(common.get("VWAP", "N/A"), float("nan"))
        if np.isfinite(vwap) and abs((price - vwap) / vwap * 100.0) <= 1.0:
            patterns.append("VWAP Pullback")

        # --- Bollinger squeeze proxy
        bbw = _bb_width_pct(close, 20, 2.0)
        if np.isfinite(bbw):
            # compare to its own 6m distribution
            bbw_hist = ((close.rolling(20).mean() + 2*close.rolling(20).std(ddof=0)) - (close.rolling(20).mean() - 2*close.rolling(20).std(ddof=0))) / close.rolling(20).mean() * 100.0
            bbw_hist = bbw_hist.dropna().iloc[-160:]
            if len(bbw_hist) >= 40:
                q20 = float(bbw_hist.quantile(0.2))
                if bbw <= q20:
                    patterns.append("Bollinger Squeeze")

        # --- VCP proxy: shrinking ATR% over last month + squeeze-ish
        atr = _to_float_safe(common.get("ATR14_PCT", "N/A"), float("nan"))
        if "ATR14_PCT_SERIES" in common:
            # not used now
            pass
        # simple: compare last 10 vs prior 20 median true range %
        tr_pct = (df["High"] - df["Low"]) / df["Close"] * 100.0
        if len(tr_pct) >= 40:
            last10 = float(tr_pct.iloc[-10:].median())
            prev20 = float(tr_pct.iloc[-30:-10].median())
            if prev20 > 0 and last10 / prev20 <= 0.8:
                patterns.append("Volatility Contraction (VCP proxy)")

        if not patterns:
            out["Pattern_List"] = "NONE"
            out["Primary_Pattern"] = "NONE"
            out["Pattern Detected"] = "NONE"
            return out

        # pick primary by priority
        priority = [
            "Range Breakout",
            "Darvas Box Breakout",
            "Triple Bottom",
            "Double Bottom",
            "Bollinger Squeeze",
            "Volatility Contraction (VCP proxy)",
            "VWAP Pullback",
            "Triple Top",
            "Double Top",
        ]
        primary = next((p for p in priority if p in patterns), patterns[0])
        out["Pattern_List"] = " | ".join(patterns)
        out["Primary_Pattern"] = primary
        out["Pattern Detected"] = primary
        return out
    except Exception:
        return out


def classify_market_stage(common: Dict[str, Any], cfg) -> Dict[str, str]:
    """Assign Market Stage + Market Sub-Stage using DMA alignment, distance, and volume surge."""
    stage = "Neutral/Transition"
    sub = "NEUTRAL_CHOP"
    try:
        price = _to_float_safe(common.get("PRICE", "N/A"), float("nan"))
        dma20 = _to_float_safe(common.get("DMA20", "N/A"), float("nan"))
        dma50 = _to_float_safe(common.get("DMA50", "N/A"), float("nan"))
        dma200 = _to_float_safe(common.get("DMA200", "N/A"), float("nan"))
        pct50 = _to_float_safe(common.get("PCT_FROM_DMA50", "N/A"), float("nan"))
        pct200 = _to_float_safe(common.get("PCT_FROM_DMA200", "N/A"), float("nan"))
        vol_surge = _to_float_safe(common.get("VOL_SURGE_RATIO", "N/A"), 0.0)
        rsi = _to_float_safe(common.get("RSI14", common.get("RSI", "N/A")), float("nan"))

        bull_stack = np.isfinite(dma20) and np.isfinite(dma50) and np.isfinite(dma200) and (dma20 > dma50 > dma200)
        bear_stack = np.isfinite(dma20) and np.isfinite(dma50) and np.isfinite(dma200) and (dma20 < dma50 < dma200)

        if bear_stack or (np.isfinite(price) and np.isfinite(dma200) and price < dma200 and np.isfinite(dma50) and dma50 < dma200):
            stage = "Mark-Down"
            sub = "MARKDOWN_TREND" if (np.isfinite(pct200) and pct200 <= -5) or bear_stack else "MARKDOWN_EARLY"
        elif bull_stack or (np.isfinite(price) and np.isfinite(dma200) and price > dma200 and np.isfinite(dma50) and dma50 > dma200):
            stage = "Mark-Up"
            # pullback within markup
            if np.isfinite(dma20) and np.isfinite(dma50) and dma20 < dma50:
                sub = "MARKUP_DOWN"
            else:
                sub = "MARKUP_TREND"
        else:
            # sideways regime
            stage = "Neutral/Transition"
            sub = "NEUTRAL_RANGE" if (np.isfinite(pct50) and abs(pct50) <= 2.0) else "NEUTRAL_CHOP"

        # accumulation breakout override: breakout-like conditions
        if np.isfinite(pct50) and pct50 >= -2 and pct50 <= 6 and vol_surge >= getattr(cfg, "vol_surge_breakout", 1.25) and (not np.isfinite(rsi) or rsi >= getattr(cfg, "rsi_breakout_min", 48.0)):
            stage = "Accumulation"
            sub = "ACCUMULATION_BREAKOUT"

        return {"Market Stage": stage, "Market Sub-Stage": sub}
    except Exception:
        return {"Market Stage": stage, "Market Sub-Stage": sub}

def _safe_pct(new_val, old_val) -> float:
    """
    Safe percent change: (new-old)/old * 100
    Returns NaN if inputs invalid or old_val is 0.
    """
    try:
        n = float(new_val)
        o = float(old_val)
        if math.isnan(n) or math.isnan(o) or o == 0.0:
            return float("nan")
        return (n - o) / o * 100.0
    except Exception:
        return float("nan")


def setup_logger(log_path: str) -> logging.Logger:
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    logger = logging.getLogger("SRS_UPDATE")
    logger.setLevel(logging.INFO)
    logger.handlers = []

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(log_path, mode="a", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    return logger


def excel_safe(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (list, tuple, dict)):
        return str(v)
    if isinstance(v, (np.generic,)):
        return v.item()
    if isinstance(v, float):
        if math.isnan(v) or math.isinf(v):
            return None
    return v




def normalize_ohlcv(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize OHLCV dataframe coming from different sources (IBKR cache, AlphaVantage, CSV, etc.)
    into canonical lowercase columns: date, open, high, low, close, volume.

    Fixes a common pandas pitfall: after lowercasing columns, duplicates may appear (e.g. 'Close' and 'close').
    When duplicates exist, df['close'] becomes a DataFrame, and pd.to_numeric raises:
      TypeError: arg must be a list, tuple, 1-d array, or Series
    """
    if df is None:
        return pd.DataFrame()

    # Some callers may pass dict/records/Series by mistake.
    if not isinstance(df, pd.DataFrame):
        try:
            df = pd.DataFrame(df)
        except Exception:
            return pd.DataFrame()

    df = df.copy()

    # Flatten MultiIndex columns if present
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ["_".join([str(x) for x in col if x is not None]).strip() for col in df.columns]

    # helper: find columns by case-insensitive match
    def _find_cols(candidates):
        cand_lower = [str(x).lower() for x in candidates]
        found = []
        for c in df.columns:
            if str(c).lower() in cand_lower:
                found.append(c)
        return found

    def _coalesce(cols):
        cols = [c for c in cols if c in df.columns]
        if not cols:
            return pd.Series([np.nan] * len(df), index=df.index)
        # If duplicates or multi-columns, coalesce left->right
        parts = []
        for c in cols:
            s = df[c]
            if isinstance(s, pd.DataFrame):
                s = s.iloc[:, 0]
            parts.append(pd.Series(s.values, index=df.index))
        tmp = pd.concat(parts, axis=1)
        return tmp.bfill(axis=1).iloc[:, 0]

    # Build canonical columns (support common variants)
    date_cols = _find_cols(["date", "datetime", "time", "timestamp"])
    open_cols = _find_cols(["open", "o"])
    high_cols = _find_cols(["high", "h"])
    low_cols  = _find_cols(["low", "l"])
    close_cols= _find_cols(["close", "adj close", "adj_close", "adjclose", "last", "price", "c"])
    vol_cols  = _find_cols(["volume", "vol", "v"])

    out = pd.DataFrame(index=df.index)
    out["date"] = _coalesce(date_cols) if date_cols else np.nan
    out["open"] = _coalesce(open_cols)
    out["high"] = _coalesce(high_cols)
    out["low"]  = _coalesce(low_cols)
    out["close"] = _coalesce(close_cols)
    out["volume"] = _coalesce(vol_cols)

    # If date wasn't a column, try index
    if out["date"].isna().all():
        if isinstance(df.index, pd.DatetimeIndex):
            out["date"] = df.index
        elif "Date" in df.columns:
            out["date"] = df["Date"]

    # Coerce types
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    for c in ["open", "high", "low", "close", "volume"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")

    # Drop rows with no close
    out = out.dropna(subset=["close"])
    out = out.sort_values("date")
    out = out.reset_index(drop=True)
    return out


def normalize_header(s: Any) -> str:
    if s is None:
        return ""
    return " ".join(str(s).strip().split())


def ensure_header(ws, header: str) -> int:
    header = normalize_header(header)
    headers = [normalize_header(c.value) for c in ws[1]]
    if header in headers:
        return headers.index(header) + 1
    col = len(headers) + 1
    ws.cell(row=1, column=col, value=header)
    return col


def get_symbol_col(ws) -> int:
    headers = [normalize_header(c.value).lower() for c in ws[1]]
    for i, h in enumerate(headers):
        if h == "symbol":
            return i + 1
    return ensure_header(ws, "Symbol")


def build_symbol_row_map(ws) -> Tuple[Dict[str, int], int]:
    sym_col = get_symbol_col(ws)
    m: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=sym_col).value
        if v is None:
            continue
        s = str(v).strip().upper()
        if s:
            m[s] = r
    return m, sym_col


def vwap_proxy(df: pd.DataFrame, lookback: int = 20) -> float:
    d = df.tail(lookback).copy()
    if d.empty:
        return float("nan")
    high = pd.to_numeric(d.get("high"), errors="coerce")
    low = pd.to_numeric(d.get("low"), errors="coerce")
    close = pd.to_numeric(d.get("close"), errors="coerce")
    vol = pd.to_numeric(d.get("volume"), errors="coerce").fillna(0.0)
    tp = (high + low + close) / 3.0
    denom = float(np.nansum(vol))
    if denom <= 0:
        return float("nan")
    return float(np.nansum(tp * vol) / denom)


def dist_acc_days(df: pd.DataFrame, lookback: int = 20) -> Tuple[int, int]:
    d = df.tail(lookback + 1).copy()
    if len(d) < lookback + 1:
        return 0, 0
    d["close"] = pd.to_numeric(d.get("close"), errors="coerce")
    d["volume"] = pd.to_numeric(d.get("volume"), errors="coerce")
    d["prev_close"] = d["close"].shift(1)
    avg = float(d["volume"].tail(lookback).mean())
    red = (d["close"] < d["prev_close"]) & (d["volume"] > avg)
    green = (d["close"] > d["prev_close"]) & (d["volume"] > avg)
    return int(red.tail(lookback).sum()), int(green.tail(lookback).sum())





def _infer_trend_label(df: pd.DataFrame) -> str:
    """
    Simple trend label based on 50/200 DMA and last close position.
    Returns: UP, DOWN, SIDEWAYS
    """
    try:
        close = pd.to_numeric(df.get("close"), errors="coerce")
        if close is None or close.isna().all() or len(close) < 220:
            return "NA"
        ma50 = close.rolling(50).mean()
        ma200 = close.rolling(200).mean()
        c = float(close.iloc[-1])
        m50 = float(ma50.iloc[-1])
        m200 = float(ma200.iloc[-1])
        if not (np.isfinite(c) and np.isfinite(m50) and np.isfinite(m200)):
            return "NA"
        if m50 > m200 and c >= m50:
            return "UP"
        if m50 < m200 and c <= m50:
            return "DOWN"
        return "SIDEWAYS"
    except Exception:
        return "NA"


def _macd_cross_flag(df: pd.DataFrame) -> str:
    """
    Compute a basic MACD signal cross flag for the latest bar.
    Returns 'Y' if MACD crossed above signal on the last bar, else ''.
    """
    try:
        close = pd.to_numeric(df.get("close"), errors="coerce")
        if close is None or close.isna().all() or len(close) < 60:
            return "N/A"
        ema12 = close.ewm(span=12, adjust=False).mean()
        ema26 = close.ewm(span=26, adjust=False).mean()
        macd = ema12 - ema26
        signal = macd.ewm(span=9, adjust=False).mean()
        if len(macd) < 2 or len(signal) < 2:
            return "N/A"
        prev = float(macd.iloc[-2] - signal.iloc[-2])
        curr = float(macd.iloc[-1] - signal.iloc[-1])
        if np.isfinite(prev) and np.isfinite(curr) and prev <= 0 and curr > 0:
            return "Y"
        return ""
    except Exception:
        return "N/A"


def compute_for_symbol(symbol: str, cfg: RankConfig, qqq_df: Optional[pd.DataFrame], logger: logging.Logger,
                       ttl_minutes: int, force_refresh: bool) -> Dict[str, Any]:
    out = {"Symbol": symbol, "_OK": False}
    out.update(DEFAULTS)

    df, source, err = fetch_data_daily_with_fallback(
        symbol,
        bar_spec="10 Y",
        bar_size="1 day",
        ttl_minutes=ttl_minutes,
        require_today=False,
        force_refresh=force_refresh,
    )
    out["DATA_SOURCE"] = source

    try:
        common = compute_common_indicators_rls(df, qqq_df)

        if not common:
            logger.warning(f"{symbol}: common indicators EMPTY")
        else:
            logger.info(f"{symbol}: indicators computed")

            # 🔥 FORCE WRITE INTO OUTPUT
            for k, v in common.items():
                if v is not None and str(v) != "nan":
                    out[k] = v

    except Exception as e:
        logger.error(f"{symbol}: compute_common_indicators_rls failed: {e}")

    print(f"DEBUG {symbol} COMMON:", common)

    if df is None or df.empty or len(df) < 60:
        logger.warning(f"{symbol} | source={source} | rows={0 if df is None else len(df)} | fetch_err={err or 'n/a'}")
        return out

    try:
        df = normalize_ohlcv(df)
        # Use the richer business-row computation (fills Trend/RSI state/decision reason/etc)
        # ==========================================================
        # CRITICAL FIX: apply common indicators into output
        # ==========================================================
        try:
            common = compute_common_indicators_rls(df, qqq_df)

            if not common:
                logger.warning(f"{symbol}: common indicators EMPTY")
            else:
                logger.info(f"{symbol}: indicators computed")
                print(f"DEBUG {symbol} COMMON:", common)

                for k, v in common.items():
                    if v is not None and str(v) != "nan":
                        out[k] = v

        except Exception as e:
            logger.error(f"{symbol}: compute_common_indicators_rls failed: {e}")

        try:
            core = compute_business_row(symbol, df, qqq_df, cfg, logger) or {}
            out.update(core)
        except Exception:
            # keep whatever we have; downstream fills best-effort
            pass

        # Ensure Trend present
        if str(out.get("Trend") or "").strip() in {"", "NA", "N/A"}:
            out["Trend"] = _infer_trend_label(df)

        # Ensure Decision Reason non-empty
        if not str(out.get("Decision Reason") or "").strip():
            out["Decision Reason"] = str(out.get("FINAL_REASON") or out.get("decision_reason") or "NO_SIGNAL")

        # MACD Cross
        out["MACD Cross"] = _macd_cross_flag(df)

        # Market Stage / Market Sub-Stage
        try:
            stg, sub = compute_market_stage_substage(df)
            out["Market Stage"] = stg
            out["Market Sub-Stage"] = sub
        except Exception:
            out["Market Stage"] = "Neutral/Transition"
            out["Market Sub-Stage"] = "NEUTRAL_CHOP"

        # Pattern_List / Primary_Pattern (richer pattern view)
        try:
            # prefer darvas-based signal if available
            if darvas_box_signal is not None:
                try:
                    tmp = darvas_box_signal(df.copy())
                    out['Darvas Signal'] = 'Y' if int(tmp.get('darvas_signal', pd.Series([0])).iloc[-1]) == 1 else ''
                    out['Darvas Breakout %'] = float(tmp.get('darvas_breakout_pct', pd.Series([0.0])).iloc[-1])
                except Exception:
                    pass

            pats = compute_patterns(df, out)
            out['Pattern_List'] = ', '.join(pats) if pats else 'NONE'
            out['Primary_Pattern'] = (pats[0] if pats and pats[0] != 'NONE' else 'NONE')
        except Exception:
            try:
                pats = compute_patterns(df, out)(df, out)
                out["Pattern_List"] = ", ".join(pats) if pats else "NONE"
                out["Primary_Pattern"] = (pats[0] if pats and pats[0] != "NONE" else "NONE")
            except Exception:
                out["Pattern_List"] = out.get("Pattern_List") or "NONE"
                out["Primary_Pattern"] = out.get("Primary_Pattern") or "NONE"

        # Sentiment placeholders (filled in buy.py when news is available)
        out["Sentiment Label"] = out.get("Sentiment Label") or "N/A"
        out["Sentiment Confidence"] = out.get("Sentiment Confidence") or "N/A"

        # Exit reasons (best-effort)
        try:
            from exit_signals import compute_exit_signals
            entry_proxy = float(out.get("Current Price") or out.get("PRICE") or np.nan)
            if np.isfinite(entry_proxy):
                ex = compute_exit_signals(df, entry_proxy) or {}
                out["Exit Reasons"] = ex.get("reasons") or ex.get("reason") or "N/A"
            else:
                out["Exit Reasons"] = "N/A"
        except Exception:
            out["Exit Reasons"] = "N/A"

        # IBKR shortability / borrow fee are not always available via historical bars.
        out["SHORTABLE_FLAG"] = out.get("SHORTABLE_FLAG") or "UNKNOWN"
        out["BORROW_FEE_PCT"] = out.get("BORROW_FEE_PCT") or "N/A"

        # Mirror underscore columns to the legacy space-named columns (the Excel has both)
        out["Confidence Score"] = out.get("CONFIDENCE_SCORE", out.get("Confidence Score", "N/A"))
        out["Final_Action"] = out.get("FINAL_ACTION", out.get("Final_Action", "N/A"))

        # Reward_%: prefer Reward_% else Best_Risk_Reward if numeric
        br = out.get("Reward_%")
        if br in (None, "", "N/A", "NA"):
            br2 = out.get("Best_Risk_Reward")
            if br2 not in (None, "", "N/A", "NA") and isinstance(br2, (int, float, np.floating)):
                out["Reward_%"] = round(float(br2), 2)

        # Pattern Detected: best-effort tag (use setup tags)
        if out.get("LONG_SETUP_TAG") and out.get("LONG_SETUP_TAG") != "N/A":
            out["Pattern Detected"] = f"LONG_{out.get('LONG_SETUP_TAG')}"
        elif out.get("SHORT_SETUP_TAG") and out.get("SHORT_SETUP_TAG") != "N/A":
            out["Pattern Detected"] = f"SHORT_{out.get('SHORT_SETUP_TAG')}"
        else:
            out["Pattern Detected"] = out.get("Pattern Detected") or "NONE"

        # Log + return (ALWAYS)
        logger.info(
            f"{symbol} | {source} | rows={len(df)} | VOLsurge={out.get('VOL_SURGE_RATIO')} | ATR%={out.get('ATR14_PCT')} | "
            f"L={out.get('LONG_SETUP_TAG')} {out.get('LONG_SCORE')} {out.get('LONG_VERDICT')} | "
            f"S={out.get('SHORT_SETUP_TAG')} {out.get('SHORT_SCORE')} {out.get('SHORT_VERDICT')} | "
            f"FINAL={out.get('FINAL_ACTION')} conf={out.get('CONFIDENCE_SCORE')}"
        )

        out = fill_template_fields_best_effort(out, df)
        out["_OK"] = True
        return out

    except Exception as e:
        out["_OK"] = False
        out["ERROR"] = str(e)
        logger.exception(f"{symbol} | compute failed: {e}")
        return out
def write_top_sheets(wb: openpyxl.Workbook, rows: list[Dict[str, Any]], top_n: int = 20) -> None:
    # remove existing if present
    for name in ("TOP_BUYS", "TOP_SHORTS"):
        if name in wb.sheetnames:
            ws_old = wb[name]
            wb.remove(ws_old)

    df = pd.DataFrame([r for r in rows if r.get("_OK")])

    # TOP_BUYS
    buys = df.sort_values(["LONG_SCORE", "CONFIDENCE_SCORE"], ascending=False).head(top_n)
    ws_b = wb.create_sheet("TOP_BUYS")
    cols_b = [
        "Symbol", "DATA_SOURCE", "LONG_SCORE", "LONG_SETUP_TAG", "LONG_VERDICT",
        "LONG_ENTRY_ZONE_LOW", "LONG_ENTRY_ZONE_HIGH", "LONG_INVALIDATION",
        "LONG_TARGET_1", "LONG_TARGET_2", "LONG_RR_RATIO",
        "DMA20", "DMA50", "DMA200", "PCT_FROM_DMA50", "PCT_FROM_DMA200",
        "VOL_TODAY", "AVG_VOL_20D", "VOL_SURGE_RATIO", "VWAP", "VWAP_DISTANCE_PCT", "ATR14_PCT",
        "DISTRIBUTION_DAYS_20D", "ACCUMULATION_DAYS_20D",
        "SPIKE_DRIVER", "DROP_DRIVER",
    ]
    ws_b.append(cols_b)
    for _, r in buys.iterrows():
        ws_b.append([excel_safe(r.get(c)) for c in cols_b])

    # TOP_SHORTS
    shorts = df.sort_values(["SHORT_SCORE", "CONFIDENCE_SCORE"], ascending=False).head(top_n)
    ws_s = wb.create_sheet("TOP_SHORTS")
    cols_s = [
        "Symbol", "DATA_SOURCE", "SHORT_SCORE", "SHORT_SETUP_TAG", "SHORT_VERDICT",
        "SHORT_ENTRY_ZONE_LOW", "SHORT_ENTRY_ZONE_HIGH", "SHORT_INVALIDATION",
        "SHORT_TARGET_1", "SHORT_TARGET_2", "SHORT_RR_RATIO",
        "DMA20", "DMA50", "DMA200", "PCT_FROM_DMA50", "PCT_FROM_DMA200",
        "VOL_TODAY", "AVG_VOL_20D", "VOL_SURGE_RATIO", "VWAP", "VWAP_DISTANCE_PCT", "ATR14_PCT",
        "DISTRIBUTION_DAYS_20D", "ACCUMULATION_DAYS_20D",
        "SPIKE_DRIVER", "DROP_DRIVER",
    ]
    ws_s.append(cols_s)
    for _, r in shorts.iterrows():
        ws_s.append([excel_safe(r.get(c)) for c in cols_s])



def _nz(x, d=0.0):
    try:
        if x is None:
            return d
        # NaN check
        if isinstance(x, float) and x != x:
            return d
        return x
    except Exception:
        return d


def _to_float_safe(x, default=np.nan) -> float:
    """Best-effort numeric conversion. Returns default on N/A / blanks / parse errors."""
    try:
        if x is None:
            return float(default)
        if isinstance(x, (int, float, np.integer, np.floating)):
            # NaN stays NaN
            return float(x)
        if isinstance(x, str):
            t = x.strip()
            if t == "" or t.upper() in {"N/A", "NA", "NONE", "NULL", "NAN", "-"}:
                return float(default)
            # remove percent sign / commas
            t = t.replace("%", "").replace(",", "")
            return float(t)
        return float(x)
    except Exception:
        return float(default)

def _rolling_slope(series: pd.Series, window: int = 20) -> float:
    """Slope of last `window` points (simple linear regression)."""
    try:
        s = pd.to_numeric(series.tail(window), errors="coerce").dropna()
        if len(s) < max(5, window // 2):
            return float("nan")
        y = s.values
        x = np.arange(len(y), dtype=float)
        # slope of best-fit line
        xm = x.mean()
        ym = y.mean()
        denom = ((x - xm) ** 2).sum()
        if denom == 0:
            return float("nan")
        slope = (((x - xm) * (y - ym)).sum()) / denom
        return float(slope)
    except Exception:
        return float("nan")

def detect_triple_bottom(df: pd.DataFrame, lookback: int = 140, tol: float = 0.03, min_sep: int = 8) -> bool:
    """Heuristic: 3 local lows within `tol` band, separated by `min_sep` bars, followed by bounce."""
    try:
        if df is None or len(df) < 60:
            return False
        d = df.tail(lookback).copy()
        close = pd.to_numeric(d.get("close") or d.get("Close"), errors="coerce")
        if close is None:
            return False
        close = close.dropna()
        if len(close) < 60:
            return False

        # local minima indices
        vals = close.values
        idxs = []
        for i in range(2, len(vals) - 2):
            if vals[i] <= vals[i-1] and vals[i] <= vals[i+1] and vals[i] <= vals[i-2] and vals[i] <= vals[i+2]:
                idxs.append(i)
        if len(idxs) < 3:
            return False

        # pick 3 minima separated by min_sep
        lows = []
        for i in idxs:
            if not lows or (i - lows[-1][0]) >= min_sep:
                lows.append((i, vals[i]))
            if len(lows) >= 6:
                break
        if len(lows) < 3:
            return False

        # try combinations of 3 lows
        from itertools import combinations
        for comb in combinations(lows, 3):
            lows_vals = np.array([v for _, v in comb], dtype=float)
            med = np.median(lows_vals)
            if med <= 0:
                continue
            if np.max(np.abs(lows_vals - med) / med) <= tol:
                # bounce confirmation: latest close above median low by 5%
                if vals[-1] >= med * 1.05:
                    return True
        return False
    except Exception:
        return False

def compute_market_stage_substage(df: pd.DataFrame) -> tuple[str, str]:
    """Coarse Wyckoff-style stage + substage using 50/200 DMAs and slope heuristics."""
    stage = "Neutral/Transition"
    sub = "NEUTRAL_CHOP"
    try:
        if df is None or len(df) < 220:
            return stage, sub

        d = df.copy()
        if "close" in d.columns:
            close = pd.to_numeric(d["close"], errors="coerce")
        elif "Close" in d.columns:
            close = pd.to_numeric(d["Close"], errors="coerce")
        else:
            return stage, sub

        # Prefer precomputed DMAs if present, else compute from close
        if "DMA50" in d.columns:
            dma50 = pd.to_numeric(d["DMA50"], errors="coerce")
        elif "dma50" in d.columns:
            dma50 = pd.to_numeric(d["dma50"], errors="coerce")
        elif "SMA_50" in d.columns:
            dma50 = pd.to_numeric(d["SMA_50"], errors="coerce")
        else:
            dma50 = close.rolling(50).mean()

        if "DMA200" in d.columns:
            dma200 = pd.to_numeric(d["DMA200"], errors="coerce")
        elif "dma200" in d.columns:
            dma200 = pd.to_numeric(d["dma200"], errors="coerce")
        elif "SMA_200" in d.columns:
            dma200 = pd.to_numeric(d["SMA_200"], errors="coerce")
        else:
            dma200 = close.rolling(200).mean()

        c = float(close.iloc[-1]) if len(close) else float("nan")
        d50 = float(dma50.iloc[-1]) if len(dma50) else float("nan")
        d200 = float(dma200.iloc[-1]) if len(dma200) else float("nan")
        if not (np.isfinite(c) and np.isfinite(d50) and np.isfinite(d200)):
            return stage, sub

        s200 = _rolling_slope(dma200, 30)
        s50 = _rolling_slope(dma50, 20)

        above200 = c >= d200
        above50 = c >= d50
        below200 = c < d200
        below50 = c < d50

        # Mark-Down
        if below200 and below50 and (np.isfinite(s200) and s200 < 0):
            return "Mark-Down", "MARKDOWN_TREND"

        # Early markdown
        if below200 and (above50 or (np.isfinite(s200) and s200 <= 0)):
            return "Mark-Down", "MARKDOWN_EARLY"

        # Mark-Up
        if above200 and (np.isfinite(s200) and s200 >= 0):
            if above50 and (np.isfinite(s50) and s50 > 0):
                return "Mark-Up", "MARKUP_TREND"
            if below50:
                return "Mark-Up", "MARKUP_PULLBACK"
            return "Mark-Up", "MARKUP_TREND"

        # Neutral range near 200
        if abs((c - d200) / d200) <= 0.03:
            return "Neutral/Transition", "NEUTRAL_RANGE"

        return stage, sub
    except Exception:
        return stage, sub

def compute_patterns(df: pd.DataFrame, row: Dict[str, Any]) -> list[str]:
    """Derive a list of pattern strings from booleans/fields already in the row and from price series."""
    pats: list[str] = []
    # Existing flags in row (Excel-friendly)
    if str(row.get("Darvas Signal", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("DARVAS_BREAKOUT")
    if str(row.get("SMC_Breakout", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("SMC_BREAKOUT")
    if str(row.get("Mean_Reversion", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("MEAN_REVERSION")
    if str(row.get("Bullish_Engulfing", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("BULLISH_ENGULFING")
    if str(row.get("Hammer", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("HAMMER")
    if str(row.get("Price Reversal", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("PRICE_REVERSAL")
    if str(row.get("MACD Cross", "")).strip().upper() in {"Y", "YES", "TRUE", "1", "✅"}:
        pats.append("MACD_CROSSOVER")

    # Series-driven: Triple Bottom
    if detect_triple_bottom(df):
        pats.append("TRIPLE_BOTTOM")

    # If nothing, NONE
    if not pats:
        pats = ["NONE"]
    # de-dup while keeping order
    seen=set()
    out=[]
    for p in pats:
        if p not in seen:
            out.append(p); seen.add(p)
    return out

def backfill_legacy_columns(row: Dict[str, Any]) -> Dict[str, Any]:
    """Backfill legacy/UI columns so the Excel sheet doesn't show blanks."""
    # Refined Buy Price
    row["Refined Buy Price"] = _nz(
        row.get("REFINED_BUY_PRICE") or
        row.get("LONG_ENTRY_ZONE_LOW") or
        row.get("BUY_PRICE"),
        "N/A"
    )

    # Recommendation & reason
    row["Recommendation"] = row.get("FINAL_ACTION", row.get("Final_Action", "AVOID"))
    row["Decision Reason"] = row.get("FINAL_REASON", "")

    # Trend / institutional
    row["Trend"] = row.get("TREND_LABEL") or row.get("HTF_Trend") or row.get("Trend") or "NA"
    row["Institutional Score"] = float(_nz(row.get("INSTITUTIONAL_SCORE"), _nz(row.get("Institutional Score"), 0.0)))

    # VWAP support (display)
    if "VWAP" in row:
        row["VWAP Support"] = row.get("VWAP")

    # Ensure RR fields are numeric, not blank
    if "SHORT_RR_RATIO" in row:
        row["SHORT_RR_RATIO"] = _to_float_safe(row.get("SHORT_RR_RATIO"), default=0.0)
    if "LONG_RR_RATIO" in row:
        row["LONG_RR_RATIO"] = _to_float_safe(row.get("LONG_RR_RATIO"), default=0.0)

    return row

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_xlsx", default="predictions_summary_out.xlsx")
    ap.add_argument("--out", dest="out_xlsx", default="predictions_summary_out.xlsx")
    ap.add_argument("--rank-config", dest="rank_cfg", default="rank_config.json")
    ap.add_argument("--ttl", dest="ttl", type=int, default=360)
    ap.add_argument("--refresh", action="store_true")
    ap.add_argument("--top", dest="top_n", type=int, default=20)
    args = ap.parse_args()

    logger = setup_logger(os.path.join("logs", "buy_srs_update.log"))
    cfg = load_rank_config(args.rank_cfg)
    logger.info(f"Loaded rank config: {asdict(cfg)}")

    wb = openpyxl.load_workbook(args.in_xlsx)
    ws = wb[wb.sheetnames[0]]

    # Ensure headers exist (dynamic: use whatever is already in the Excel + ensure template columns)
    existing_headers = get_excel_headers(ws)
    # Guarantee minimum template fields exist
    desired_headers = list(dict.fromkeys(existing_headers + TEMPLATE_COLUMNS_MINIMUM + list(DEFAULTS.keys())))
    col_map = {c: ensure_header(ws, c) for c in desired_headers}

    sym_map, sym_col = build_symbol_row_map(ws)

    # QQQ for rel strength (best effort)
    qqq_df = None
    try:
        qqq_df, qqq_src, qqq_err = fetch_data_daily_with_fallback("QQQ", require_today=False, ttl_minutes=args.ttl, force_refresh=args.refresh)
        qqq_df = normalize_ohlcv(qqq_df)
        if qqq_df is not None and not qqq_df.empty:
            logger.info(f"QQQ loaded for rel-strength: rows={len(qqq_df)} src={qqq_src}")
        else:
            qqq_df = None
    except Exception:
        qqq_df = None

    symbols_cfg = [str(s).strip().upper() for s in (CONFIG_SYMBOLS or []) if str(s).strip()]
    symbols_xl = list(sym_map.keys())
    # union (config first)
    symbols = list(dict.fromkeys(symbols_cfg + symbols_xl))
    logger.info(f"Total symbols: config={len(symbols_cfg)} excel={len(symbols_xl)} union={len(symbols)}")

    computed_rows: list[Dict[str, Any]] = []

    for idx, sym in enumerate(symbols, start=1):
        t0 = time.time()
        row = compute_for_symbol(sym, cfg, qqq_df, logger, ttl_minutes=args.ttl, force_refresh=args.refresh)
        row = backfill_legacy_columns(row)
        computed_rows.append(row)

        # write to Summary by symbol row
        excel_row = sym_map.get(sym)
        if excel_row is None:
            excel_row = ws.max_row + 1
            ws.cell(row=excel_row, column=sym_col, value=sym)
            sym_map[sym] = excel_row

        if row.get("_OK"):
            for col_name, col_idx in col_map.items():
                v = row.get(col_name, DEFAULTS.get(col_name, "N/A"))
                ws.cell(row=excel_row, column=col_idx, value=excel_safe(v))
        else:
            # IMPORTANT: Do NOT overwrite existing Excel cells with DEFAULTS when compute failed.
            # This preserves the values already produced by buy.py (stage 1).
            pass

        dt = time.time() - t0
        if idx % 10 == 0:
            logger.info(f"Progress: {idx}/{len(symbols)} processed. last={sym} dt={dt:.2f}s")

    # Top sheets
    try:
        write_top_sheets(wb, computed_rows, top_n=args.top_n)
        logger.info("Created TOP_BUYS and TOP_SHORTS sheets")
    except Exception as e:
        logger.warning(f"Top sheets failed: {e}")

    # Save (if locked, timestamp)
    try:
        wb.save(args.out_xlsx)
        logger.info(f"Saved: {args.out_xlsx}")
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = args.out_xlsx.replace(".xlsx", f"_{ts}.xlsx")
        wb.save(alt)
        logger.warning(f"Output locked; saved to: {alt}")


if __name__ == "__main__":
    main()
def compute_business_row(symbol: str, df: pd.DataFrame, qqq_df: pd.DataFrame | None, rank_cfg, logger: logging.Logger) -> dict[str, Any]:
    """
    FULL business-row computation.
    Goal: avoid blanks/defaults and always output deterministic values from OHLCV + common indicators.
    """
    out: dict[str, Any] = {"Symbol": symbol}

    # Defaults first (never leave missing keys)
    out["Patterns"] = "NONE"
    out["Primary Pattern"] = "NONE"
    out["Trend"] = "N/A"
    out["Decision Reason"] = "NO_SIGNAL"
    out["Institutional Score"] = "N/A"
    out["Confidence Score"] = 0.0
    out["Tech Fallback Score"] = 0.0
    out["Rule-Based Buy"] = "NO"
    out["Darvas Breakout %"] = "N/A"
    out["EMA21 Slope"] = "N/A"
    out["ADX Strength"] = "N/A"
    out["RSI State"] = "N/A"
    out["OBV Trend"] = "N/A"
    out["At BB Lower"] = False
    out["BELOW_VWAP_PCT"] = "N/A"
    out["Current Price"] = "N/A"
    out["Invalidation_Level"] = "N/A"
    out["Atr Trailing Stop"] = "N/A"
    out["Risk_%"] = "N/A"
    out["Reward_%"] = "N/A"
    out["90D Gain (%)"] = "N/A"
    out["Days to Peak"] = "N/A"
    out["Sentiment Label"] = "N/A"
    out["Sentiment Confidence"] = "N/A"
    out["Exit Reasons"] = "N/A"
    out["Position_Size_Class"] = "N/A"
    out["BORROW_FEE_PCT"] = "N/A"
    out["SHORT_SETUP_TAG"] = "N/A"
    out["SHORT_INVALIDATION"] = "N/A"
    out["SHORT_TARGET_1"] = "N/A"
    out["SHORT_TARGET_2"] = "N/A"
    out["SHORT_RR_RATIO"] = "N/A"
    out["SHORT_FEASIBILITY"] = "N/A"
    out["LONG_SETUP_TAG"] = "N/A"
    out["LONG_SCORE"] = "N/A"
    out["SHORT_SCORE"] = "N/A"
    out["Breakout"] = False
    out["Undervalued"] = False

    if df is None or df.empty:
        out["Decision Reason"] = "NO_DATA"
        return out

    d = df.copy()
    # Normalize numeric types
    for c in ("open","high","low","close","volume"):
        if c in d.columns:
            d[c] = pd.to_numeric(d[c], errors="coerce")
    d = d.dropna(subset=["close"]).reset_index(drop=True)
    if d.empty:
        out["Decision Reason"] = "NO_DATA"
        return out

    # Compute common indicators (may include RS vs QQQ, VWAP dist, vol surge, ATR%, etc.)
    try:
        common = compute_common_indicators_rls(d, qqq_df=qqq_df)
    except Exception:
        common = {}

    close = _to_float(common.get("CLOSE", common.get("close", d["close"].iloc[-1])))
    out["Current Price"] = close if np.isfinite(close) else "N/A"

    # Candle/signal detectors
    try:
        smc = bool(detect_smc_accumulation_breakout(d))
    except Exception:
        smc = False
    try:
        mean_rev = bool(detect_mean_reversion_buy(d))
    except Exception:
        mean_rev = False
    try:
        engulf = bool(detect_bullish_engulfing(d))
    except Exception:
        engulf = False
    try:
        hammer = bool(detect_hammer(d))
    except Exception:
        hammer = False

    # Darvas
    try:
        darvas_sig, darvas_pct = darvas_box_signal(d)
    except Exception:
        darvas_sig, darvas_pct = False, float("nan")
    out["Darvas Breakout %"] = float(darvas_pct) if np.isfinite(darvas_pct) else "N/A"

    # Pattern framework (structural + tactical)
    try:
        flags = compute_patterns_full(d, common=common)
    except Exception:
        flags = {}
    flags.update({
        "SMC_BREAKOUT": smc,
        "MEAN_REVERSION": mean_rev,
        "BULLISH_ENGULFING": engulf,
        "HAMMER": hammer,
    })
    flags["DARVAS_BREAKOUT"] = bool(flags.get("DARVAS_BREAKOUT", False) or darvas_sig)

    try:
        patterns_str, primary = patterns_to_strings(flags)
    except Exception:
        patterns_str, primary = "NONE", "NONE"
    out["Patterns"] = patterns_str or "NONE"
    out["Primary Pattern"] = primary or "NONE"

    # Trend / indicator states
    trend = str(common.get("TREND", common.get("trend", "")) or "")
    if trend.strip().upper() in ("NA","N/A","NONE"):
        trend = ""
    if not trend:
        try:
            trend = compute_upward_trend(d)
        except Exception:
            trend = ""
    out["Trend"] = trend if trend else "N/A"

    out["EMA21 Slope"] = common.get("EMA21_SLOPE", common.get("ema21_slope", "N/A"))
    out["ADX Strength"] = common.get("ADX_STRENGTH", common.get("adx_strength", "N/A"))
    rsi = _to_float(common.get("RSI14", common.get("rsi14", np.nan)))
    out["RSI State"] = _rsi_state(rsi)
    out["OBV Trend"] = common.get("OBV_TREND", common.get("obv_trend", "N/A"))
    out["At BB Lower"] = bool(common.get("AT_BB_LOWER", common.get("at_bb_lower", False)))
    out["BELOW_VWAP_PCT"] = common.get("BELOW_VWAP_PCT", common.get("below_vwap_pct", "N/A"))

    # 90D gain + days to peak (pure price-based)
    try:
        if len(d) >= 91:
            c0 = float(d["close"].iloc[-1])
            c90 = float(d["close"].iloc[-91])
            out["90D Gain (%)"] = ((c0 / c90) - 1.0) * 100.0 if c90 else "N/A"
            window = d["close"].iloc[-91:].to_numpy()
            out["Days to Peak"] = int(np.argmax(window))
    except Exception:
        pass

    # Long/Short setup and scoring
    # Normalize common keys for rank_long_short expectations
    common["RSI"] = common.get("RSI", common.get("RSI14", 0))
    common["MACD_hist"] = common.get("MACD_hist", common.get("MACD_HIST", 0))
    common["MACD_signal"] = common.get("MACD_signal", common.get("MACD_SIGNAL", 0))
    common["ABOVE_DMA20"] = bool(common.get("ABOVE_DMA20", False))
    common["ABOVE_DMA50"] = bool(common.get("ABOVE_DMA50", False))
    common["ABOVE_DMA200"] = bool(common.get("ABOVE_DMA200", False))
    try:
        common["VOL_SURGE_RATIO"] = float(common.get("VOL_SURGE_RATIO", 0) or 0)
    except Exception:
        common["VOL_SURGE_RATIO"] = 0.0
    try:
        common["VWAP_DISTANCE_PCT"] = float(common.get("VWAP_DISTANCE_PCT", 0) or 0)
    except Exception:
        common["VWAP_DISTANCE_PCT"] = 0.0
    try:
        common["ATR14_PCT"] = float(common.get("ATR14_PCT", 0) or 0)
    except Exception:
        common["ATR14_PCT"] = 0.0

    try:
        long_setup = detect_long_setup(common, d, rank_cfg)
    except Exception:
        long_setup = None

    try:
        short_setup = detect_short_setup(common, d, rank_cfg)
    except Exception:
        short_setup = None

    out["LONG_SETUP_TAG"] = long_setup if long_setup else "N/A"
    out["SHORT_SETUP_TAG"] = short_setup if short_setup else "N/A"

    try:
        long_score = float(score_long(common, rank_cfg))
    except Exception:
        long_score = 0.0
    try:
        short_score = float(score_short(common, rank_cfg))
    except Exception:
        short_score = 0.0

    out["LONG_SCORE"] = long_score
    out["SHORT_SCORE"] = short_score



    # Plans (targets, invalidation, RR)
    try:
        long_plan = build_long_plan(common, d, long_setup, rank_cfg) or {}
    except Exception:
        long_plan = {}
    try:
        short_plan = build_short_plan(common, d, short_setup, rank_cfg) or {}
    except Exception:
        short_plan = {}


    inv = long_plan.get("invalidation_level", np.nan)
    atr_stop = long_plan.get("atr_trailing_stop", np.nan)
    target = long_plan.get("target_1", np.nan)

    out["Invalidation_Level"] = inv if np.isfinite(_to_float(inv)) else "N/A"
    out["Atr Trailing Stop"] = atr_stop if np.isfinite(_to_float(atr_stop)) else "N/A"

    risk_pct = float(_safe_pct(close, _to_float(inv))) if np.isfinite(close) and np.isfinite(_to_float(inv)) else float("nan")
    rew_pct = float(_safe_pct(_to_float(target), close)) if np.isfinite(close) and np.isfinite(_to_float(target)) else float("nan")
    out["Risk_%"] = abs(risk_pct) if np.isfinite(risk_pct) else "N/A"
    out["Reward_%"] = rew_pct if np.isfinite(rew_pct) else "N/A"

    out["SHORT_INVALIDATION"] = short_plan.get("invalidation_level", "N/A")
    out["SHORT_TARGET_1"] = short_plan.get("target_1", "N/A")
    out["SHORT_TARGET_2"] = short_plan.get("target_2", "N/A")
    rr = short_plan.get("rr_ratio", np.nan)
    out["SHORT_RR_RATIO"] = rr if np.isfinite(_to_float(rr)) else "N/A"
    out["SHORT_FEASIBILITY"] = feasibility_label(short_plan) if short_plan else "N/A"

    # Institutional score
    try:
        inst = score_institutional_investor(d)
        out["Institutional Score"] = inst if np.isfinite(_to_float(inst)) else "N/A"
    except Exception:
        out["Institutional Score"] = "N/A"

    # Confidence score: use best of long/short
    conf = float(np.nanmax([long_score if np.isfinite(long_score) else np.nan, short_score if np.isfinite(short_score) else np.nan]))
    if not np.isfinite(conf):
        conf = 0.0
    out["Confidence Score"] = conf

    # Tech fallback score: long_score if finite else 0
    out["Tech Fallback Score"] = long_score if np.isfinite(long_score) else 0.0

    # Rule-based buy (simple institutional gating)
    rule_buy = (
        (np.isfinite(long_score) and long_score >= float(rank_cfg.long_score_buy_now)) and
        ((str(out["Trend"]).upper().startswith("UP")) or (out["Primary Pattern"] in ("DARVAS_BREAKOUT","SMC_BREAKOUT","CUP_HANDLE","TRIPLE_BOTTOM","DOUBLE_BOTTOM")))
    )
    out["Rule-Based Buy"] = "YES" if rule_buy else "NO"

    # Sentiment proxy (only price-based; true sentiment requires news)
    try:
        if out["RSI State"] == "BULLISH" and str(out["Trend"]).upper().startswith("UP"):
            out["Sentiment Label"] = "BULLISH"
            out["Sentiment Confidence"] = "MEDIUM"
        elif out["RSI State"] in ("OVERSOLD",) and not str(out["Trend"]).upper().startswith("DOWN"):
            out["Sentiment Label"] = "RECOVERY"
            out["Sentiment Confidence"] = "LOW"
        elif str(out["Trend"]).upper().startswith("DOWN"):
            out["Sentiment Label"] = "BEARISH"
            out["Sentiment Confidence"] = "MEDIUM"
        else:
            out["Sentiment Label"] = "NEUTRAL"
            out["Sentiment Confidence"] = "LOW"
    except Exception:
        pass

    # Exit reasons (rule-based)
    exit_reasons = []
    try:
        if str(out["Trend"]).upper().startswith("DOWN"):
            exit_reasons.append("DOWN_TREND")
        if out["RSI State"] == "WEAK":
            exit_reasons.append("RSI_WEAK")
        bv = _to_float(out.get("BELOW_VWAP_PCT", np.nan))
        if np.isfinite(bv) and bv > 2.0:
            exit_reasons.append("BELOW_VWAP")
        if out.get("Reward_%") == "N/A" and out.get("Risk_%") != "N/A":
            exit_reasons.append("NO_TARGETS")
    except Exception:
        pass
    out["Exit Reasons"] = ", ".join(exit_reasons) if exit_reasons else "N/A"

    # Position sizing class (based on ATR% if available)
    atrp = _to_float(common.get("ATR14_PCT", common.get("atr14_pct", np.nan)))
    if np.isfinite(atrp):
        if atrp <= 3: out["Position_Size_Class"] = "LARGE"
        elif atrp <= 6: out["Position_Size_Class"] = "MEDIUM"
        else: out["Position_Size_Class"] = "SMALL"
    else:
        out["Position_Size_Class"] = "N/A"

    # Breakout / Undervalued proxies
    out["Breakout"] = bool(common.get("BREAKOUT", darvas_sig or smc))
    out["Undervalued"] = bool(common.get("UNDERVALUED", False))

    # Expose individual pattern flags to columns if those columns exist
    for k, v in (flags or {}).items():
        out[k] = bool(v)

    # Decision Reason (guaranteed non-empty string)
    reasons = []
    if out["Primary Pattern"] and out["Primary Pattern"] != "NONE":
        reasons.append(f"PRIMARY={out['Primary Pattern']}")
    if smc:
        reasons.append("SMC_BREAKOUT")
    if mean_rev:
        reasons.append("MEAN_REVERSION")
    if darvas_sig:
        reasons.append("DARVAS")
    if np.isfinite(rsi):
        reasons.append(f"RSI={rsi:.1f}({out['RSI State']})")
    vws = common.get("VWAP_DISTANCE_PCT", np.nan)
    if np.isfinite(_to_float(vws)):
        reasons.append(f"VWAP_DIST={float(vws):.2f}%")
    vs = common.get("VOL_SURGE_RATIO", np.nan)
    if np.isfinite(_to_float(vs)):
        reasons.append(f"VOL_SURGE={float(vs):.2f}")
    out["Decision Reason"] = "; ".join(reasons) if reasons else "NO_SIGNAL"

    return out




def _rsi_state(rsi_value) -> str:
    """
    Convert RSI into a human readable state.
    """
    try:
        r = float(rsi_value)
        if math.isnan(r):
            return "N/A"
    except Exception:
        return "N/A"

    if r >= 70:
        return "OVERBOUGHT"
    if r <= 30:
        return "OVERSOLD"
    return "NEUTRAL"


def _safe_pct(new_val, old_val) -> float:
    """
    Safe percent change: (new-old)/old * 100
    Returns NaN if inputs invalid or old_val is 0.
    """
    try:
        n = float(new_val)
        o = float(old_val)
        if math.isnan(n) or math.isnan(o) or o == 0.0:
            return float("nan")
        return (n - o) / o * 100.0
    except Exception:
        return float("nan")