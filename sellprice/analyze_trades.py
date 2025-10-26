
from utils import print_signal_summary

import re
import pandas as pd
from tabulate import tabulate
from utils import signal_storage, export_signals_to_excel  # raw memory dump helper



def run_post_analysis():
    """
    Post-run summary + exports:
      - CSV:   signal_summary.csv   (filtered/sorted consolidated table incl. ATRpct if present)
      - XLSX:  signal_summary.xlsx  (Summary, TopByType, TopByStage, RawLog, SellTargets, SellTargetsPct, RiskMetrics)
      - XLSX:  signals_raw.xlsx     (raw, unfiltered log dump from utils.signal_storage)
    """
    import re
    import pandas as pd
    from tabulate import tabulate
    from utils import signal_storage, export_signals_to_excel  # raw memory dump helper

    try:
        # 1) Consolidated "signals" table from your existing printer
        signals = print_signal_summary()

        # If nothing to summarize, still emit predictable (empty) Excel files
        if not signals:
            print("[INFO] No signals to summarize.")
            export_signals_to_excel("signals_raw.xlsx")
            with pd.ExcelWriter("signal_summary.xlsx") as xw:
                pd.DataFrame(columns=["Symbol","Date","Signal","Confidence","Type","Condition","ATRpct","SizePct"]).to_excel(
                    xw, sheet_name="Summary", index=False
                )
                for name in ["TopByType","TopByStage","RawLog","SellTargets","SellTargetsPct","RiskMetrics"]:
                    pd.DataFrame().to_excel(xw, sheet_name=name, index=False)
            print("💾 Saved: signal_summary.xlsx")
            return

        # 2) Build DataFrame
        df = pd.DataFrame(signals)

        # Normalize dtypes
        for col in ("Confidence","ATRpct","SizePct"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        for col in ("Type","Signal","Condition"):
            if col in df.columns:
                df[col] = df[col].astype(str)

        # 3) Apply simple floors (tune to taste)
        floors = {'SHORT': 55, 'SELL': 50, 'BUY': 60}
        if 'Type' in df.columns and 'Confidence' in df.columns:
            df = df[df.apply(lambda r: r['Confidence'] >= floors.get(str(r['Type']).upper(), 50), axis=1)].copy()

        # 4) Sort and persist CSV (keeps your original behavior)
        sort_cols, sort_asc = [], []
        if 'Type' in df.columns:
            sort_cols.append('Type'); sort_asc.append(True)
        if 'Confidence' in df.columns:
            sort_cols.append('Confidence'); sort_asc.append(False)
        if sort_cols:
            df.sort_values(sort_cols, ascending=sort_asc, inplace=True)

        # Guarantee stable Summary columns (adds ATRpct if missing)
        summary_cols = ["Symbol","Date","Signal","Confidence","Type","Condition","ATRpct","SizePct","Pattern detected","Breakout"]
        for c in summary_cols:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[summary_cols]

        df.to_csv("signal_summary.csv", index=False)
        print("💾 Saved: signal_summary.csv")

        # ---------- Build extra Excel views ----------

        # A) Top-by-Stage (bucket the 'Stage:' WATCH rows)
        def _extract_stage(label: str) -> str:
            if not isinstance(label, str):
                return ""
            return label.replace('Stage:', '').strip()

        def _stage_bucket(stage: str) -> str:
            stage = (stage or "").strip()
            if stage.startswith('Mark-Up'): return 'Mark-Up'
            if stage.startswith('Accumulation'): return 'Accumulation'
            if stage.startswith('Distribution'): return 'Distribution'
            if stage.startswith('Mark-Down'): return 'Mark-Down'
            return 'Neutral/Transition'

        df_stage_top = pd.DataFrame(columns=['Symbol','Date','Stage','Confidence','Bucket'])
        if {'Signal','Type','Confidence'}.issubset(df.columns):
            stage_rows = df[
                df['Signal'].str.startswith('Stage:', na=False)
                & (df['Type'].str.upper() == 'WATCH')
            ].copy()
            if not stage_rows.empty:
                stage_rows['Stage'] = stage_rows['Signal'].apply(_extract_stage)
                stage_rows['Bucket'] = stage_rows['Stage'].apply(_stage_bucket)
                top_by_bucket = []
                for bucket in ['Mark-Up','Accumulation','Distribution','Mark-Down','Neutral/Transition']:
                    sub = stage_rows[stage_rows['Bucket'] == bucket].nlargest(5, 'Confidence')
                    if not sub.empty:
                        top_by_bucket.append(sub[['Symbol','Date','Stage','Confidence']].assign(Bucket=bucket))
                if top_by_bucket:
                    df_stage_top = pd.concat(top_by_bucket, ignore_index=True)

        # B) Top-N per Type
        TOP_N = 10
        df_top_types = pd.DataFrame(columns=df.columns.tolist() + ['_Type'])
        if 'Type' in df.columns and 'Confidence' in df.columns:
            stacks = []
            for t in ['SHORT','SELL','BUY','WATCH']:
                sub = df[df['Type'].str.upper() == t].nlargest(TOP_N, 'Confidence')
                if not sub.empty:
                    stacks.append(sub.assign(_Type=t))
            if stacks:
                df_top_types = pd.concat(stacks, ignore_index=True)

        # C) Raw (unfiltered) log for auditing
        df_raw = pd.DataFrame(signal_storage) if signal_storage else pd.DataFrame()

        # D) SELL TARGETS wide tables (absolute + pct)
        df_targets = pd.DataFrame()
        df_targets_pct = pd.DataFrame()
        if not df_raw.empty and 'Type' in df_raw.columns and 'Signal' in df_raw.columns:
            df_t = df_raw[df_raw['Type'].astype(str).str.upper() == 'TARGET'].copy()

            def _parse_abs(row):
                m = re.match(r"SELL_TGT_(\d+w):\s*([0-9.]+)", str(row.get("Signal", "")))
                if not m: return None
                return (row.get("Symbol"), row.get("Date"), m.group(1), float(m.group(2)))

            def _parse_pct(row):
                m = re.match(r"SELL_TGT_(\d+w):\s*([0-9.]+)", str(row.get("Signal", "")))
                if not m: return None
                pct = row.get("Confidence", None)
                try: pct = float(pct) if pct is not None else None
                except Exception: pct = None
                return (row.get("Symbol"), row.get("Date"), m.group(1), pct)

            parsed_abs = [x for x in df_t.apply(_parse_abs, axis=1).tolist() if x is not None]
            parsed_pct = [x for x in df_t.apply(_parse_pct, axis=1).tolist() if x is not None]

            if parsed_abs:
                df_long = pd.DataFrame(parsed_abs, columns=["Symbol","Date","Horizon","Target"])
                df_targets = (
                    df_long.pivot_table(index=["Symbol","Date"], columns="Horizon", values="Target", aggfunc="last")
                    .reset_index()
                )

            if parsed_pct:
                df_longp = pd.DataFrame(parsed_pct, columns=["Symbol","Date","Horizon","Pct"])
                df_targets_pct = (
                    df_longp.pivot_table(index=["Symbol","Date"], columns="Horizon", values="Pct", aggfunc="last")
                    .reset_index()
                )

            desired = ["Symbol","Date","6w","8w","12w","18w","30w"]
            if not df_targets.empty:
                cols = [c for c in desired if c in df_targets.columns] + [c for c in df_targets.columns if c not in desired]
                df_targets = df_targets[cols]
            if not df_targets_pct.empty:
                cols = [c for c in desired if c in df_targets_pct.columns] + [c for c in df_targets_pct.columns if c not in desired]
                df_targets_pct = df_targets_pct[cols]
                # Presentation rounding
                for c in ["6w","8w","12w","18w","30w"]:
                    if c in df_targets_pct.columns:
                        df_targets_pct[c] = pd.to_numeric(df_targets_pct[c], errors='coerce').round(2)

        # E) RiskMetrics (ATRpct per symbol/date from Summary)
        df_risk = pd.DataFrame()
        if 'ATRpct' in df.columns:
            # last available per symbol (sorted by Date if Date is sortable)
            df_tmp = df[['Symbol','Date','ATRpct']].copy()
            # try to get most recent per symbol
            try:
                df_tmp['Date'] = pd.to_datetime(df_tmp['Date'])
                df_tmp.sort_values(['Symbol','Date'], inplace=True)
                df_risk = df_tmp.groupby('Symbol', as_index=False).last()
            except Exception:
                # fallback: just dedupe, keep last row ordering
                df_risk = df_tmp.drop_duplicates(subset=['Symbol'], keep='last')
            # ensure numeric + round
            df_risk['ATRpct'] = pd.to_numeric(df_risk['ATRpct'], errors='coerce').round(2)

        # 5) Write Excel (multi-sheet), then add heatmaps on SellTargetsPct and RiskMetrics.ATRpct
        try:
            with pd.ExcelWriter("signal_summary.xlsx") as xw:
                df.to_excel(xw, sheet_name="Summary", index=False)
                df_top_types.to_excel(xw, sheet_name="TopByType", index=False)
                df_stage_top.to_excel(xw, sheet_name="TopByStage", index=False)
                df_raw.to_excel(xw, sheet_name="RawLog", index=False)
                df_targets.to_excel(xw, sheet_name="SellTargets", index=False)
                df_targets_pct.to_excel(xw, sheet_name="SellTargetsPct", index=False)
                df_risk.to_excel(xw, sheet_name="RiskMetrics", index=False)
            print("💾 Saved: signal_summary.xlsx")

            # Apply conditional formatting
            try:
                from openpyxl import load_workbook
                from openpyxl.formatting.rule import ColorScaleRule

                wb = load_workbook("signal_summary.xlsx")

                # Heatmap on SellTargetsPct sheet (columns for 6w..30w)
                if "SellTargetsPct" in wb.sheetnames:
                    ws = wb["SellTargetsPct"]
                    header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column+1)]
                    horizon_cols = [j for j, name in enumerate(header, start=1) if name in ("6w","8w","12w","18w","30w")]
                    if horizon_cols and ws.max_row >= 2:
                        col_min, col_max = min(horizon_cols), max(horizon_cols)
                        start_col_letter = ws.cell(row=1, column=col_min).column_letter
                        end_col_letter = ws.cell(row=1, column=col_max).column_letter
                        rng = f"{start_col_letter}2:{end_col_letter}{ws.max_row}"
                        rule = ColorScaleRule(
                            start_type='min', start_color='F8696B',
                            mid_type='percentile', mid_value=50, mid_color='FFEB84',
                            end_type='max', end_color='63BE7B'
                        )
                        ws.conditional_formatting.add(rng, rule)
                        # number format as percent with 2 decimals
                        for col in range(col_min, col_max+1):
                            for row in range(2, ws.max_row+1):
                                ws.cell(row=row, column=col).number_format = '0.00"%"'

                # Heatmap on RiskMetrics!ATRpct
                if "RiskMetrics" in wb.sheetnames:
                    ws = wb["RiskMetrics"]
                    # find ATRpct column
                    header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column+1)]
                    try:
                        atr_col = header.index("ATRpct") + 1
                        if ws.max_row >= 2:
                            col_letter = ws.cell(row=1, column=atr_col).column_letter
                            rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                            rule = ColorScaleRule(
                                start_type='min', start_color='63BE7B',  # low ATR% = green
                                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                end_type='max', end_color='F8696B'       # high ATR% = red
                            )
                            ws.conditional_formatting.add(rng, rule)
                            for row in range(2, ws.max_row+1):
                                ws.cell(row=row, column=atr_col).number_format = '0.00"%"'
                    except ValueError:
                        pass

                wb.save("signal_summary.xlsx")
            except Exception as e:
                print(f"⚠️ Conditional formatting skipped: {e}")

        except Exception as e:
            print(f"⚠️ Excel export failed: {e}")

        # 6) Console prints (keep your original feel)
        if not df_stage_top.empty:
            for bucket in ['Mark-Up','Accumulation','Distribution','Mark-Down','Neutral/Transition']:
                sub = df_stage_top[df_stage_top['Bucket'] == bucket][['Symbol','Date','Stage','Confidence']]
                if not sub.empty:
                    print(f"\nTop {min(5, len(sub))} {bucket}:")
                    print(tabulate(sub, headers='keys', tablefmt='fancy_grid', showindex=False))

        print("\n📊 Trade Signal Summary Table:\n")
        print(tabulate(df, headers='keys', tablefmt='fancy_grid', showindex=False))

        # Also emit a separate raw Excel for quick access (redundant but convenient)
        export_signals_to_excel("signals_raw.xlsx")

    except Exception as e:
        print(f"⚠️ Error during post-analysis summary: {e}")




def _norm_type(t: str) -> str:
    t = (t or "").upper()
    return {"BUY","SELL","SHORT","WATCH"}.intersection({t}).pop() if t in {"BUY","SELL","SHORT","WATCH"} else "WATCH"



def _print_top_by_stage(df_all, top_n=5):
    # Keep only the Stage rows we logged from sell.py
    stage_rows = df_all[(df_all['Signal'].str.startswith('Stage:')) & (df_all['Type'].str.upper() == 'WATCH')].copy()
    if stage_rows.empty:
        print("\n📭 No stage rows found.")
        return

    # Extract stage label
    stage_rows['Stage'] = stage_rows['Signal'].str.replace('Stage:', '').str.strip()

    def _bucket(label):
        if label.startswith('Mark-Up'): return 'Mark-Up'
        if label.startswith('Accumulation'): return 'Accumulation'
        if label.startswith('Distribution'): return 'Distribution'
        if label.startswith('Mark-Down'): return 'Mark-Down'
        return 'Neutral/Transition'

    stage_rows['Bucket'] = stage_rows['Stage'].apply(_bucket)

    for bucket in ['Mark-Up', 'Accumulation', 'Distribution', 'Mark-Down', 'Neutral/Transition']:
        sub = stage_rows[stage_rows['Bucket'] == bucket].nlargest(top_n, 'Confidence')
        if sub.empty:
            continue
        print(f"\nTop {min(top_n, len(sub))} {bucket}:")
        print(tabulate(sub[['Symbol','Date','Stage','Confidence']], headers='keys', tablefmt='fancy_grid', showindex=False))

# Call this at the end of run_post_analysis(), after the existing prints:
#   _print_top_by_stage(df)
