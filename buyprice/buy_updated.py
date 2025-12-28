#!/usr/bin/env python3
"""
buy_clean_synced_to_ddddd_v2.py

Fixes for "most Excel values are blank":
1) Template header matching is now *canonicalized* (unicode normalized, hyphen variants unified,
   non-alphanumerics stripped) so we correctly match headers like "Rule‑Based Buy" vs "Rule-Based Buy".
   If we fail to match, we used to append new columns on the right, leaving the original template
   columns blank — which looks like "most values are blank".
2) We now coerce NaN/inf to None before writing, so Excel cells are truly filled/blank as intended.
3) We can auto-detect the correct sheet by searching for a header row containing "Symbol"/"Ticker",
   instead of assuming "Summary".
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
import joblib

from config import symbols as CONFIG_SYMBOLS
from fetching import fetch_data_cached
from compute import compute_indicators, analyze_symbol_all
from backtest import evaluate_backtest_accuracy
from eps_features import fetch_quarterly_eps, eps_growth_flags, fetch_market_sentiment
from exit_signals import compute_exit_signals
from symbol_analysis import analyze_symbol
from upward import (
    detect_smc_accumulation_breakout,
    detect_mean_reversion_buy,
    detect_bullish_engulfing,
    detect_hammer,
    compute_upward_trend,
    compute_signal_score
)

# -----------------------------
# Thresholds / Bands (sync with ddddd_buy.py)
# -----------------------------
WATCH_PROB = 0.30
DEFAULT_BUY_PROB = 0.50
STRONG_BUY_PROB = 0.65
TECH_BUY_FALLBACK = 0.70

HERE = Path(__file__).resolve().parent
TEMPLATE_FILENAME = "predictions_summary.xlsx"


# Load required model feature list (training-time contract)
REQUIRED_FEATURES: Optional[List[str]] = None
try:
    feats_path = HERE / "model_features.txt"
    if feats_path.exists():
        REQUIRED_FEATURES = [ln.strip() for ln in feats_path.read_text(encoding="utf-8").splitlines() if ln.strip()]
        if REQUIRED_FEATURES:
            print(f"✅ Loaded model feature contract: {len(REQUIRED_FEATURES)} features from model_features.txt")
except Exception as _e:
    REQUIRED_FEATURES = None


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--no-cache", action="store_true", help="Bypass cache entirely")
    p.add_argument("--refresh-cache", choices=["none", "force"], default="none",
                   help="Cache policy: 'force' to re-fetch all symbols")
    p.add_argument("--cache-ttl-mins", type=int, default=360,
                   help="Refresh data if cache older than this many minutes")
    p.add_argument("--allow-stale", action="store_true",
                   help="Allow cache that doesn’t include today’s data")
    p.add_argument("--excel-out", type=str, default="predictions_summary_out.xlsx")
    p.add_argument("--template", type=str, default=TEMPLATE_FILENAME, help="Template xlsx file")
    p.add_argument("--sheet", type=str, default="", help="Optional sheet name (auto-detect if blank)")
    return p.parse_args()


def load_symbol_df(sym: str, args) -> pd.DataFrame:
    if args.no_cache:
        return fetch_data_cached(sym, ttl_minutes=0, force_refresh=True, require_today=not args.allow_stale)
    if args.refresh_cache == "force":
        return fetch_data_cached(sym, ttl_minutes=args.cache_ttl_mins, force_refresh=True, require_today=not args.allow_stale)
    return fetch_data_cached(sym, ttl_minutes=args.cache_ttl_mins, force_refresh=False, require_today=not args.allow_stale)


def load_model() -> Optional[object]:
    p = HERE / "strong_buy_xgb_model_calibrated.pkl"
    if p.exists():
        m = joblib.load(p)
        print(f"✅ Loaded model: {p.name}")
        return m
    print("⚠️ Calibrated model not found; Model Probability will be 0.")
    return None


def load_threshold(default: float = DEFAULT_BUY_PROB) -> float:
    p = HERE / "strong_buy_thresholds.json"
    if p.exists():
        try:
            obj = json.loads(p.read_text())
            thr = float(obj.get("best_f1", {}).get("thr", default))
            print(f"✅ Using threshold (best_f1): {thr:.2f}")
            return thr
        except Exception as e:
            print(f"[WARN] failed to read thresholds: {e}")
    print(f"[INFO] Using default threshold: {default:.2f}")
    return float(default)


MODEL = load_model()
BUY_THRESH = max(0.40, float(load_threshold()))
STRONG_BUY_THRESH = float(STRONG_BUY_PROB)


def get_confidence_band(prob: float) -> str:
    if prob >= STRONG_BUY_THRESH:
        return "STRONG BUY"
    if prob >= BUY_THRESH:
        return "BUY"
    if prob >= WATCH_PROB:
        return "WATCH"
    return "NO TRADE"


def _load_fallback_feature_list(path: str = "model_features.txt") -> Optional[List[str]]:
    try:
        p = HERE / path
        feats = [ln.strip() for ln in p.read_text(encoding="utf-8").splitlines() if ln.strip()]
        return feats or None
    except Exception:
        return None


def _expected_features_for_model(model, df_sample: pd.DataFrame) -> List[str]:
    """
    Determine the exact feature list (and order) to feed to the trained model.

    Priority:
      1) model_features.txt (training-time contract)
      2) model.feature_names_in_ (if present)
      3) numeric columns present in the current df (last resort)
    """
    if REQUIRED_FEATURES:
        return list(REQUIRED_FEATURES)

    if model is not None and hasattr(model, "feature_names_in_"):
        try:
            names = [str(c).strip() for c in model.feature_names_in_]
            if names:
                return names
        except Exception:
            pass

    file_feats = _load_fallback_feature_list("model_features.txt")
    if file_feats:
        return file_feats

    return df_sample.select_dtypes(include="number").columns.tolist()


def _predict_proba_for_last_row(symbol: str, df: pd.DataFrame) -> float:
    """
    Predict calibrated probability for the latest row, with strict feature alignment.

    IMPORTANT: We must align live features to the *training* feature contract in model_features.txt,
    otherwise XGBoost will error or silently mis-score and many Excel cells appear blank.
    """
    if MODEL is None:
        return 0.0

    # Use latest bar only
    latest = df.iloc[-1:].copy()
    latest.columns = latest.columns.astype(str).map(str.strip)

    # === Feature alignment contract ===
    # Load the training-time feature contract if present; fall back safely otherwise.
    try:
        feats_path = HERE / "model_features.txt"
        if feats_path.exists():
            with open(feats_path, "r", encoding="utf-8") as f:
                required = [line.strip() for line in f if line.strip()]
        else:
            required = _expected_features_for_model(MODEL, latest)
    except Exception:
        required = _expected_features_for_model(MODEL, latest)

    # Ensure all required columns exist
    for col in required:
        if col not in latest.columns:
            latest[col] = 0.0

    # This is the exact alignment step you asked about:
    X = latest.reindex(columns=required, fill_value=0.0)

    # Safety: numeric only
    X = X.apply(pd.to_numeric, errors="coerce").replace([np.inf, -np.inf], np.nan).fillna(0.0)

    try:
        proba = float(MODEL.predict_proba(X)[0][1])
        return max(0.0, min(1.0, proba))
    except Exception as e:
        print(f"⚠️ Model prediction failed for {symbol}: {e}")
        return 0.0


def _compute_indicator_snap(df: pd.DataFrame) -> dict:
    last = df.iloc[-1]
    snap = {}
    snap["EMA_Uptrend"] = bool(last.get("EMA_uptrend", False))
    snap["EMA21_Slope"] = float(last.get("EMA_21_slope", np.nan))

    adx = float(last.get("ADX_14", np.nan))
    snap["ADX"] = adx
    snap["ADX_Strength"] = "Strong" if np.isfinite(adx) and adx >= 25 else "Weak"

    macd_cross = bool(last.get("MACD_crossover", False))
    snap["MACD_Crossover"] = "✅" if macd_cross else "❌"

    rsi = float(last.get("RSI_14", np.nan))
    snap["RSI_14"] = rsi
    if np.isfinite(rsi) and rsi >= 70:
        snap["RSI_State"] = "Overbought"
    elif np.isfinite(rsi) and rsi <= 30:
        snap["RSI_State"] = "Oversold"
    else:
        snap["RSI_State"] = "Neutral"

    try:
        obv_slope = float(df["OBV"].diff().rolling(5).mean().iloc[-1])
    except Exception:
        obv_slope = float("nan")
    snap["OBV_Slope"] = obv_slope
    snap["OBV_Trend"] = "Up" if np.isfinite(obv_slope) and obv_slope > 0 else ("Down" if np.isfinite(obv_slope) and obv_slope < 0 else "Flat")

    close = float(last.get("close", np.nan))
    bb_lower = float(last.get("BB_lower", np.nan))
    bb_upper = float(last.get("BB_upper", np.nan))
    snap["At_BB_Lower"] = "✅" if np.isfinite(close) and np.isfinite(bb_lower) and close <= bb_lower * 1.01 else "❌"
    snap["At_BB_Upper"] = "✅" if np.isfinite(close) and np.isfinite(bb_upper) and close >= bb_upper * 0.99 else "❌"
    return snap


def _tech_fallback_score(snap: dict, df: pd.DataFrame, smc_breakout: bool, mean_rev: bool) -> float:
    score = 0.0
    if snap.get("EMA_Uptrend"):
        score += 0.25
    if float(snap.get("ADX", 0) or 0) >= 25:
        score += 0.20
    if snap.get("MACD_Crossover") == "✅":
        score += 0.20
    rsi = float(snap.get("RSI_14") or 50)
    if 40 <= rsi <= 60:
        score += 0.10
    if float(snap.get("OBV_Slope", 0) or 0) > 0:
        score += 0.10
    if smc_breakout:
        score += 0.10
    if mean_rev:
        score += 0.05
    near_support = bool(df.iloc[-1].get("near_support", False))
    if near_support:
        score += 0.05
    return round(score, 2)



# -----------------------------
# Derived columns for Excel (Breakout / Undervalued / Trend Reversal / Pattern Detected / DipReclaim)
# -----------------------------

def _is_yes(v: object) -> bool:
    """Interpret various truthy encodings used across the pipeline (✅/❌, True/False, 1/0, yes/no)."""
    if v is None:
        return False
    if isinstance(v, (bool, np.bool_)):
        return bool(v)
    # numeric
    if isinstance(v, (int, float, np.integer, np.floating)):
        try:
            return float(v) != 0.0 and np.isfinite(float(v))
        except Exception:
            return False
    s = str(v).strip().lower()
    return s in {"1", "true", "t", "yes", "y", "✅", "check", "checked"}

def _yes_no(flag: bool) -> str:
    return "✅" if bool(flag) else "❌"

def _pattern_list(row: dict) -> str:
    patterns = []
    if _is_yes(row.get("Darvas Signal")):
        patterns.append("Darvas")
    if _is_yes(row.get("SMC_Breakout")):
        patterns.append("SMC Breakout")
    if _is_yes(row.get("Mean_Reversion")):
        patterns.append("Mean Reversion")
    if _is_yes(row.get("Bullish_Engulfing")):
        patterns.append("Bullish Engulfing")
    if _is_yes(row.get("Hammer")):
        patterns.append("Hammer")
    if _is_yes(row.get("Price Reversal")):
        patterns.append("Price Reversal")
    return ", ".join(patterns) if patterns else ""

def add_excel_derived_columns(df_summary: pd.DataFrame) -> pd.DataFrame:
    """
    Add user-friendly boolean columns based on existing signal columns already computed in the pipeline.

    Columns added:
      - Breakout: Darvas or SMC breakout
      - Undervalued: tech fallback / mean-reversion / support confluence heuristic
      - Trend Reversal: reversal/candle hints
      - Pattern Detected: comma-separated list of detected patterns
      - DipReclaim: 'dip + reclaim' style condition (near support + reversal + trend resume)
    """
    if df_summary is None or df_summary.empty:
        return df_summary

    out = df_summary.copy()

    # Helper that safely gets numeric
    def _num(x, default=np.nan):
        try:
            v = float(x)
            return v if np.isfinite(v) else default
        except Exception:
            return default

    breakout = []
    underval = []
    reversal = []
    patterns = []
    dipreclaim = []

    for _, r in out.iterrows():
        rd = r.to_dict()

        # Breakout
        b = _is_yes(rd.get("Darvas Signal")) or _is_yes(rd.get("SMC_Breakout"))
        breakout.append(_yes_no(b))

        # Undervalued heuristic (since we may not have a 'Current Price' column in summary)
        tech = _num(rd.get("Tech Fallback Score"), default=np.nan)
        at_bb = _is_yes(rd.get("At BB Lower"))
        near = _is_yes(rd.get("Near Support"))
        mean_rev = _is_yes(rd.get("Mean_Reversion"))
        u = (np.isfinite(tech) and tech >= 0.70) or (mean_rev and near) or (at_bb and near)
        underval.append(_yes_no(u))

        # Trend reversal heuristic
        macd_cross = _is_yes(rd.get("MACD Cross"))
        rsi_state = str(rd.get("RSI State") or "").lower()
        rev = (
            _is_yes(rd.get("Price Reversal"))
            or _is_yes(rd.get("Bullish_Engulfing"))
            or _is_yes(rd.get("Hammer"))
            or (macd_cross and ("oversold" in rsi_state))
        )
        reversal.append(_yes_no(rev))

        # Pattern list
        patterns.append(_pattern_list(rd))

        # DipReclaim (near support + reversal + (trend up or strong trend strength) + volume/obv confirmation)
        ema_up = _is_yes(rd.get("EMA Uptrend"))
        vol_surge = _is_yes(rd.get("Volume Surge"))
        obv_trend = _is_yes(rd.get("OBV Trend"))
        ts = _num(rd.get("Trend_Strength"), default=0.0)
        dip = near and (ema_up or ts >= 0.60) and rev and (vol_surge or obv_trend)
        dipreclaim.append(_yes_no(dip))

    out["Breakout"] = breakout
    out["Undervalued"] = underval
    out["Trend Reversal"] = reversal
    out["Pattern Detected"] = patterns
    out["DipReclaim"] = dipreclaim
    return out


# -----------------------------
# Trade management / decision columns
# -----------------------------

def add_trade_management_columns(df_summary: pd.DataFrame) -> pd.DataFrame:
    """Add actionable columns that convert signals into a trade plan.

    Adds (high-priority):
      - Final_Action
      - Best_Risk_Reward
      - Primary_Entry_Price
      - Add_On_Dip_Price
      - Invalidation_Level
      - Risk_%
      - Reward_%
      - Risk_Reward_Ratio
      - Position_Size_Class
      - Buy_Window_Status

    Notes:
      * Uses existing pipeline columns only.
      * Requires "Current Price" (we populate it in enrichment loop).
    """
    if df_summary is None or df_summary.empty:
        return df_summary

    out = df_summary.copy()

    def _num(x, default=np.nan):
        try:
            v = float(x)
            return v if np.isfinite(v) else default
        except Exception:
            return default

    def _is_up_trend(v) -> bool:
        s = str(v or "").strip().lower()
        return s in {"up", "bull", "bullish", "markup", "mark-up", "✅"}

    def _is_stage_ok(v) -> bool:
        s = str(v or "").strip().lower()
        return ("accum" in s) or ("mark" in s and "down" not in s)

    def _is_buyish_signal(sig) -> bool:
        s = str(sig or "").strip().lower()
        return s.startswith("buy")

    def _choose_primary_entry(r: dict):
        """Return (entry_price, source)"""
        candidates = [
            ("Refined Buy Price", r.get("Refined Buy Price")),
            ("Candle Entry 2w", r.get("Candle Entry 2w")),
            ("VWAP Support", r.get("VWAP Support")),
            ("Candle Entry 4w", r.get("Candle Entry 4w")),
            ("Candle Entry 6w", r.get("Candle Entry 6w")),
            ("Candle Entry 8w", r.get("Candle Entry 8w")),
        ]
        for name, val in candidates:
            pv = _num(val)
            if np.isfinite(pv) and pv > 0:
                return pv, name
        # fallback
        cp = _num(r.get("Current Price"))
        return (cp, "Current Price") if np.isfinite(cp) else (np.nan, "")

    def _choose_add_on(entry_source: str, r: dict):
        # conservative ladder for scaling
        ladder = {
            "Refined Buy Price": "Candle Entry 4w",
            "Candle Entry 2w": "Candle Entry 4w",
            "VWAP Support": "Candle Entry 4w",
            "Candle Entry 4w": "Candle Entry 6w",
            "Candle Entry 6w": "Candle Entry 8w",
            "Candle Entry 8w": "Candle Entry 12w",
        }
        nxt = ladder.get(entry_source, "Candle Entry 4w")
        pv = _num(r.get(nxt))
        return pv if np.isfinite(pv) and pv > 0 else np.nan

    # compute per-row
    primary_entry = []
    add_on = []
    invalidation = []
    risk_pct = []
    reward_pct = []
    rr_ratio = []
    buy_window = []
    final_action = []
    entry_source_list = []
    size_class = []

    # helper metrics for ranking
    rr_for_rank = []
    score_for_rank = []

    for _, row in out.iterrows():
        r = row.to_dict()

        cp = _num(r.get("Current Price"))
        entry, entry_src = _choose_primary_entry(r)
        addp = _choose_add_on(entry_src, r)
        stop = _num(r.get("Atr Trailing Stop"))

        primary_entry.append(entry if np.isfinite(entry) else None)
        add_on.append(addp if np.isfinite(addp) else None)
        invalidation.append(stop if np.isfinite(stop) else None)
        entry_source_list.append(entry_src)

        # Buy window status needs current price
        if np.isfinite(cp) and np.isfinite(entry) and entry > 0:
            dist = (cp - entry) / entry
            if dist <= -0.03:
                bw = "EARLY"
            elif -0.03 < dist <= 0.03:
                bw = "IDEAL"
            elif 0.03 < dist <= 0.10:
                bw = "LATE"
            else:
                bw = "OVEREXTENDED"
        else:
            bw = "UNKNOWN"
        buy_window.append(bw)

        # Risk/Reward
        if np.isfinite(entry) and entry > 0 and np.isfinite(stop) and stop > 0:
            rp = (entry - stop) / entry * 100.0
            rp = max(0.0, rp)
        else:
            rp = np.nan

        # Use 90D Gain (%) as reward proxy (already in your sheet)
        rew = _num(r.get("90D Gain (%)"))
        if not np.isfinite(rew):
            rew = np.nan

        risk_pct.append(round(rp, 2) if np.isfinite(rp) else None)
        reward_pct.append(round(rew, 2) if np.isfinite(rew) else None)

        if np.isfinite(rp) and rp > 0 and np.isfinite(rew):
            rr = rew / rp
        else:
            rr = np.nan
        rr_ratio.append(round(rr, 2) if np.isfinite(rr) else None)
        rr_for_rank.append(rr if np.isfinite(rr) else 0.0)

        # Decision gate
        exit_now = bool(r.get("Exit Now", False))
        trend_ok = _is_up_trend(r.get("Trend"))
        stage_ok = _is_stage_ok(r.get("Market Stage"))
        buy_confirm = (
            _is_yes(r.get("Model-Driven Strong Buy"))
            or _is_yes(r.get("Model-Driven Buy"))
            or _is_yes(r.get("Rule-Based Buy"))
            or _is_buyish_signal(r.get("Signal"))
        )

        # Strength filter (2 of 3)
        conf = _num(r.get("Confidence Score"), default=np.nan)
        inst = _num(r.get("Institutional Score"), default=np.nan)
        adx_strength = str(r.get("ADX Strength") or "").strip().lower()
        strength_hits = 0
        if np.isfinite(conf) and conf >= 0.65:
            strength_hits += 1
        if np.isfinite(inst) and inst >= 0.70:
            strength_hits += 1
        if "strong" in adx_strength or _num(r.get("Trend_Strength"), default=0.0) >= 0.60:
            strength_hits += 1
        strength_ok = strength_hits >= 2

        # Momentum/support structure (any one)
        structure_ok = (
            _is_yes(r.get("Breakout"))
            or str(r.get("Darvas Signal") or "").strip().upper() == "BUY"
            or _is_yes(r.get("DipReclaim"))
            or _is_yes(r.get("SMC_Breakout"))
        )

        if exit_now:
            fa = "EXIT"
        elif trend_ok and stage_ok and buy_confirm and strength_ok and structure_ok:
            # If overextended, don't chase
            if bw == "OVEREXTENDED":
                fa = "BUY ON DIP"
            else:
                fa = "BUY (Months)"
        elif buy_confirm and not exit_now:
            fa = "HOLD/WATCH"
        else:
            fa = "AVOID/WAIT"
        final_action.append(fa)

        # Position size class (simple + robust)
        sym_vol = int(_num(r.get("Sym Vol Regime"), default=1))
        vix_vol = int(_num(r.get("VIX Vol Regime"), default=1))
        high_vol = (sym_vol >= 2) or (vix_vol >= 2)
        if fa.startswith("BUY"):
            if (not high_vol) and np.isfinite(conf) and conf >= 0.75 and np.isfinite(inst) and inst >= 0.75:
                sc = "FULL"
            elif high_vol:
                sc = "SMALL"
            else:
                sc = "HALF"
        else:
            sc = "-"
        size_class.append(sc)

        # Score for rank
        rr_safe = rr if np.isfinite(rr) else 0.0
        conf_safe = conf if np.isfinite(conf) else 0.0
        inst_safe = inst if np.isfinite(inst) else 0.0
        score_for_rank.append(rr_safe * 0.55 + conf_safe * 0.25 + inst_safe * 0.20)

    out["Primary_Entry_Price"] = primary_entry
    out["Primary_Entry_Source"] = entry_source_list
    out["Add_On_Dip_Price"] = add_on
    out["Invalidation_Level"] = invalidation
    out["Risk_%"] = risk_pct
    out["Reward_%"] = reward_pct
    out["Risk_Reward_Ratio"] = rr_ratio
    out["Buy_Window_Status"] = buy_window
    out["Final_Action"] = final_action
    out["Position_Size_Class"] = size_class

    # Best_Risk_Reward grade (A+/A/B/C/D) using score distribution
    scores = pd.Series(score_for_rank)
    if len(scores) >= 5:
        q80 = float(scores.quantile(0.80))
        q60 = float(scores.quantile(0.60))
        q40 = float(scores.quantile(0.40))
        q20 = float(scores.quantile(0.20))
        q90 = float(scores.quantile(0.90))
    else:
        q90, q80, q60, q40, q20 = 2.0, 1.5, 1.0, 0.7, 0.4

    grades = []
    for s in scores.tolist():
        if s >= q90:
            g = "A+"
        elif s >= q80:
            g = "A"
        elif s >= q60:
            g = "B"
        elif s >= q40:
            g = "C"
        else:
            g = "D"
        grades.append(g)

    out["Best_Risk_Reward"] = grades
    return out


# -----------------------------
# Momentum (breakout vs pullback) columns
# -----------------------------
def add_momentum_columns(df_summary: pd.DataFrame) -> pd.DataFrame:
    """Populate Momentum* columns using a lightweight institutional momentum rubric.

    This is intentionally rule-based (no ML dependency) so the Excel columns
    are always populated.

    Required inputs (best-effort; missing inputs fall back gracefully):
      EMA Uptrend, EMA21 Slope, ADX, MACD Cross, RSI, OBV Trend, Volume Surge, SMC_Breakout
    """
    out = df_summary.copy()

    for col in [
        "Momentum Recommendation",
        "Momentum Decision Reason",
        "Momentum Confidence Grade",
        "Momentum Expected Holding Period",
    ]:
        if col not in out.columns:
            out[col] = None

    def _boolish(v) -> bool:
        if isinstance(v, (bool, np.bool_)):
            return bool(v)
        s = str(v).strip().upper()
        return s in ("✅", "TRUE", "1", "YES", "Y")

    def _momentum_for_row(r: dict):
        ema_up = _boolish(r.get("EMA Uptrend"))
        macd = _boolish(r.get("MACD Cross"))
        smc = _boolish(r.get("SMC_Breakout"))
        vol_surge = _boolish(r.get("Volume Surge"))

        obv = str(r.get("OBV Trend", "")).upper()
        obv_bull = ("UP" in obv) or ("BULL" in obv) or (obv.strip() == "✅")

        rsi = pd.to_numeric(pd.Series([r.get("RSI")]), errors="coerce").iloc[0]
        adx = pd.to_numeric(pd.Series([r.get("ADX")]), errors="coerce").iloc[0]
        slope = pd.to_numeric(pd.Series([r.get("EMA21 Slope")]), errors="coerce").iloc[0]

        score = 0
        reasons = []
        if ema_up:
            score += 2; reasons.append("EMA uptrend")
        if np.isfinite(slope) and slope > 0:
            score += 1; reasons.append("EMA21 rising")
        if np.isfinite(adx) and adx >= 20:
            score += 1; reasons.append("ADX strong")
        if macd:
            score += 1; reasons.append("MACD bullish")
        if np.isfinite(rsi) and 50 <= rsi <= 75:
            score += 1; reasons.append("RSI healthy")
        if obv_bull:
            score += 1; reasons.append("OBV up")
        if vol_surge:
            score += 1; reasons.append("Volume surge")
        if smc:
            score += 1; reasons.append("SMC breakout")

        if score >= 7 and (smc or vol_surge):
            rec = "MOMO_BREAKOUT"
            hold = "4-12 weeks"
        elif score >= 5:
            rec = "MOMO_PULLBACK"
            hold = "2-8 weeks"
        else:
            rec = "NO_MOMO"
            hold = "N/A"

        if score >= 8:
            grade = "A"
        elif score >= 6:
            grade = "B"
        elif score >= 4:
            grade = "C"
        else:
            grade = "D"

        reason = " + ".join(reasons) if reasons else "Insufficient momentum confirmation"
        return rec, reason, grade, hold

    for idx, row in out.iterrows():
        # Only fill if blank/NaN
        if pd.isna(out.at[idx, "Momentum Recommendation"]):
            rec, reason, grade, hold = _momentum_for_row(row.to_dict())
            out.at[idx, "Momentum Recommendation"] = rec
            out.at[idx, "Momentum Decision Reason"] = reason
            out.at[idx, "Momentum Confidence Grade"] = grade
            out.at[idx, "Momentum Expected Holding Period"] = hold

    return out


# -----------------------------
# Excel header matching (fix blanks)
# -----------------------------
import unicodedata, re

def _canon(s: object) -> str:
    """Canonicalize header strings to match template even with unicode/typos."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    # unify hyphen-like chars to "-"
    s = s.replace("–", "-").replace("—", "-").replace("‑", "-").replace("−", "-")
    s = s.strip().lower()
    # collapse whitespace
    s = re.sub(r"\s+", " ", s)
    # remove non-alphanum
    s2 = re.sub(r"[^a-z0-9]+", "", s)
    return s2

def _find_header_row_in_ws(ws, search_rows: int = 80) -> int:
    max_r = min(search_rows, ws.max_row or 1)
    max_c = ws.max_column or 1
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if _canon(ws.cell(r, c).value) in {"symbol", "ticker", "tickers"}:
                return r
    return 1

def _score_sheet_for_headers(ws) -> int:
    """Pick best sheet by counting header hits in first 80 rows."""
    hr = _find_header_row_in_ws(ws)
    max_c = ws.max_column or 1
    hits = 0
    for c in range(1, max_c + 1):
        k = _canon(ws.cell(hr, c).value)
        if k:
            hits += 1
    # require at least a few headers beyond just Symbol
    return hits

def _pick_sheet(wb, preferred_name: str = ""):
    if preferred_name and preferred_name in wb.sheetnames:
        return wb[preferred_name]
    # pick sheet with most header cells on detected header row
    best_ws = wb.active
    best_score = -1
    for name in wb.sheetnames:
        ws = wb[name]
        sc = _score_sheet_for_headers(ws)
        if sc > best_score:
            best_score = sc
            best_ws = ws
    return best_ws

def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, (ws.max_column or 1) + 1):
        v = ws.cell(header_row, c).value
        k = _canon(v)
        if k:
            m[k] = c
    return m

def _ensure_header(ws, header_row: int, header_map: Dict[str, int], header: str) -> int:
    """Find an existing matching header; otherwise append."""
    k = _canon(header)
    if not k:
        # shouldn't happen
        new_col = (ws.max_column or 1) + 1
        ws.cell(header_row, new_col).value = header
        return new_col

    if k in header_map:
        return header_map[k]

    # Try fuzzy match: sometimes template uses slightly different wording.
    # Example: "modelprobability" vs "modelprob"
    for existing_k, col_idx in header_map.items():
        if existing_k == k:
            return col_idx
        if existing_k.startswith(k) or k.startswith(existing_k):
            return col_idx

    # append new header
    new_col = (ws.max_column or 1) + 1
    ws.cell(header_row, new_col).value = header
    header_map[k] = new_col
    return new_col

def _excel_safe(v):
    """Convert pandas/numpy types to Excel-safe python scalars; NaN/inf -> None."""
    if v is None:
        return None
    # bool is subclass of int; keep it
    if isinstance(v, (bool, str, int)):
        return v
    try:
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating, float)):
            fv = float(v)
            if not np.isfinite(fv):
                return None
            return fv
    except Exception:
        pass
    # pandas Timestamp
    if hasattr(v, "to_pydatetime"):
        try:
            return v.to_pydatetime()
        except Exception:
            pass
    # fallback
    try:
        # treat empty/nan-like strings
        if isinstance(v, str) and v.strip() == "":
            return None
    except Exception:
        pass
    return v


def write_template_excel(df: pd.DataFrame, template_path: Path, out_path: Path, sheet_name: str = "") -> None:
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    ws = _pick_sheet(wb, preferred_name=sheet_name)

    header_row = _find_header_row_in_ws(ws)
    header_map = _build_header_map(ws, header_row)
    data_start = header_row + 1

    # Ensure every df column exists in template headers (append if missing)
    col_index: Dict[str, int] = {}
    for col in df.columns:
        idx = _ensure_header(ws, header_row, header_map, col)
        col_index[col] = idx

    # Clear existing values under the columns we will write (also clears formulas)
    max_existing = ws.max_row or data_start
    for r in range(data_start, max_existing + 1):
        for idx in col_index.values():
            ws.cell(r, idx).value = None

    # Write rows
    records = df.to_dict(orient="records")
    for i, rowd in enumerate(records):
        r = data_start + i
        for col, idx in col_index.items():
            ws.cell(r, idx).value = _excel_safe(rowd.get(col))

    wb.save(out_path)


def main():
    args = parse_args()

    if not CONFIG_SYMBOLS:
        raise SystemExit("config.symbols is empty")

    template_path = HERE / args.template
    out_path = Path(args.excel_out)

    # First pass: base analysis via analyze_symbol_all (same as ddddd_buy.py)
    summary: List[dict] = []
    for symbol in CONFIG_SYMBOLS:
        df_raw = load_symbol_df(symbol, args)
        if df_raw is None or df_raw.empty:
            print(f"⚠️ {symbol}: no data (fetch/cache failure)")
            continue
        df = compute_indicators(df_raw.copy(), symbol=symbol)
        if df is None or df.empty:
            print(f"⚠️ {symbol}: indicators failed/empty")
            continue

        # res = analyze_symbol_all(symbol)
        res = analyze_symbol(symbol, df_raw=df_raw)
        if not res:
            continue

        # normalize recommendation key (required for legacy logic)
        if "Recommendation" in res and "recommendation" not in res:
            res["recommendation"] = res["Recommendation"]

        # guarantee explainability fields BEFORE append
        res.setdefault("Decision Reason", None)
        res.setdefault("Confidence Grade", None)
        res.setdefault("Expected Holding Period", None)

        # NOW append (this line must be last)
        summary.append(res)

    if not summary:
        raise SystemExit("⚠️ No valid predictions could be generated.")

    # Second pass: enrichment
    hit_list, gain_list, days_list = [], [], []
    for res in summary:
        symbol = res.get("Symbol", "")
        try:
            df_raw = load_symbol_df(symbol, args)
            df = compute_indicators(df_raw.copy(), symbol=symbol)

            # Latest price (needed for Buy_Window_Status / chase checks)
            try:
                res["Current Price"] = round(float(df["close"].iloc[-1]), 2)
            except Exception:
                res["Current Price"] = None

            # EPS flags
            eps_df = fetch_quarterly_eps(symbol)
            eps_flags = eps_growth_flags(eps_df)
            res["EPS Increase 2Q"] = eps_flags["EPS Increase 2Q"]
            res["EPS Increase 3Q"] = eps_flags["EPS Increase 3Q"]
            res["EPS Increase 4Q"] = eps_flags["EPS Increase 4Q"]

            # News sentiment
            sent = fetch_market_sentiment(symbol)
            res["News Sentiment Score"] = round(float(sent.get("news_sentiment_score") or 0.0), 4)
            res["News Positive Ratio"] = round(float(sent.get("news_positive_ratio") or 0.0), 4)
            res["News Article Count"] = int(sent.get("news_article_count") or 0)
            conf = sent.get("sentiment_confidence")
            if conf is None or conf == "":
                n_art = int(sent.get("news_article_count") or 0)
                if n_art >= 15:
                    conf = "HIGH"
                elif n_art >= 5:
                    conf = "MED"
                elif n_art > 0:
                    conf = "LOW"
                else:
                    conf = None
            res["Sentiment Confidence"] = conf

            # Price-action signals
            df = compute_upward_trend(df)
            signal_score, signal_count, pa_reco = compute_signal_score(df)
            res["Signal Score"] = round(float(signal_score), 2)
            res["Signal Count"] = int(signal_count)
            res["Signal"] = pa_reco

            # Pattern booleans
            smc = bool(detect_smc_accumulation_breakout(df))
            mean_rev = bool(detect_mean_reversion_buy(df))
            res["SMC_Breakout"] = smc
            res["Mean_Reversion"] = mean_rev
            res["Bullish_Engulfing"] = bool(detect_bullish_engulfing(df))
            res["Hammer"] = bool(detect_hammer(df))
            res["Trend_Strength"] = int(df.iloc[-1].get("trend_strength", 0))

            # Regimes / volume pressure
            res["Sym Vol Regime"] = int(df.iloc[-1].get("sym_vol_regime", 0))
            res["VIX Vol Regime"] = int(df.iloc[-1].get("VIX_vol_regime", 0))
            res["Volume Pressure"] = float(df.iloc[-1].get("volume_pressure", 0.0))

            # Snapshot
            snap = _compute_indicator_snap(df)
            res["EMA Uptrend"] = "✅" if snap["EMA_Uptrend"] else "❌"
            res["EMA21 Slope"] = round(snap["EMA21_Slope"], 4) if np.isfinite(snap["EMA21_Slope"]) else None
            res["ADX Strength"] = snap["ADX_Strength"]
            res["MACD Cross"] = snap["MACD_Crossover"]
            res["RSI"] = round(snap["RSI_14"], 1) if np.isfinite(snap["RSI_14"]) else None
            res["RSI State"] = snap["RSI_State"]
            res["OBV Trend"] = snap["OBV_Trend"]
            res["At BB Lower"] = snap["At_BB_Lower"]

            # Rule-based SMA cross + confidence threshold
            df["SMA_20"] = df["close"].rolling(window=20).mean()
            df["SMA_50"] = df["close"].rolling(window=50).mean()
            df["SMA_cross_signal"] = (df["SMA_20"] > df["SMA_50"]) & (df["SMA_20"].shift(1) <= df["SMA_50"].shift(1))
            conf = float(res.get("Confidence Score", 0.0) or 0.0)
            rule_based_buy = bool(df["SMA_cross_signal"].iloc[-1]) and conf > BUY_THRESH
            res["Rule-Based Buy"] = "✅" if rule_based_buy else "❌"

            # Model proba + band
            prob = _predict_proba_for_last_row(symbol, df)
            res["Model Probability"] = round(float(prob), 2)
            res["Model-Driven Buy"] = "✅" if prob >= BUY_THRESH else "❌"
            res["Model-Driven Strong Buy"] = "✅" if prob >= STRONG_BUY_THRESH else "❌"
            res["Confidence Band"] = get_confidence_band(prob)

            # Exit signals (entry = Refined Buy Price)
            entry = res.get("Refined Buy Price", None)
            try:
                entry_price = float(entry) if entry is not None and not (isinstance(entry, float) and np.isnan(entry)) else float(df["close"].iloc[-1])
            except Exception:
                entry_price = float(df["close"].iloc[-1])

            exit_info = compute_exit_signals(df, entry_price=entry_price)
            res["Exit Now"] = bool(exit_info.get("ExitNow", False))
            res["Atr Trailing Stop"] = exit_info.get("AtrTrailingStop", None)
            res["Exit Reasons"] = exit_info.get("ExitReasons", "")

            # Safety
            if bool(res.get("Exit Now", False)):
                res["Signal"] = "HOLD (Exit risk)"
                res["Model-Driven Buy"] = "❌"
                res["Model-Driven Strong Buy"] = "❌"
                res["Confidence Band"] = "NO TRADE"

            # Tech fallback
            if prob < BUY_THRESH:
                tech_score = _tech_fallback_score(snap, df, smc, mean_rev)
                res["Tech Fallback Score"] = tech_score
                if (not bool(res.get("Exit Now", False))) and snap.get("EMA_Uptrend") and tech_score >= TECH_BUY_FALLBACK and signal_count >= 2:
                    res["Signal"] = "BUY (Tech Fallback)"
            else:
                res["Tech Fallback Score"] = 0.0

            # Backtest (90D)
            hit, gain, days_to_peak = evaluate_backtest_accuracy(symbol, df, entry_price, gain_thresh=0.04, use_close=True)

        except Exception as e:
            print(f"⚠️ {symbol}: enrichment/backtest failed ({e})")
            hit, gain, days_to_peak = False, 0.0, -1
            res.setdefault("Model Probability", 0.0)
            res.setdefault("Model-Driven Buy", "❌")
            res.setdefault("Model-Driven Strong Buy", "❌")
            res.setdefault("Confidence Band", "WATCH")

        hit_list.append("✅" if hit else "❌")
        gain_list.append(round(float(gain), 2) if np.isfinite(float(gain)) else None)
        days_list.append(int(days_to_peak) if isinstance(days_to_peak, int) and days_to_peak >= 0 else "N/A")

    df_summary = pd.DataFrame(summary)

    # Add derived user-friendly columns (Breakout/Undervalued/Reversal/Patterns/DipReclaim)
    df_summary = add_excel_derived_columns(df_summary)

    # Add actionable trade-management columns
    df_summary = add_trade_management_columns(df_summary)

    # Populate Momentum* columns (breakout vs pullback rubric)
    df_summary = add_momentum_columns(df_summary)

    # Ensure Exit Reasons is never blank (keeps Excel column populated)
    if "Exit Reasons" in df_summary.columns:
        er = df_summary["Exit Reasons"].astype(str).fillna("").str.strip()
        df_summary["Exit Reasons"] = np.where(er.eq(""), "No exit signals", er)


    # Ensure Candle Entry columns exist
    for col in [
        "Candle Entry 2w", "Candle Entry 4w", "Candle Entry 6w", "Candle Entry 8w",
        "Candle Entry 12w", "Candle Entry 18w", "Candle Entry 30w"
    ]:
        if col not in df_summary.columns:
            df_summary[col] = None

    # Price Reversal aggregate
    rev_cols = ["Mean_Reversion", "Bullish_Engulfing", "Hammer"]
    for c in rev_cols:
        if c not in df_summary.columns:
            df_summary[c] = False
    has_rev = df_summary[rev_cols].fillna(False).astype(bool).any(axis=1)
    df_summary["Price Reversal"] = np.where(has_rev, "✅", "❌")

    # 90D results
    df_summary["90D Hit"] = hit_list
    df_summary["90D Gain (%)"] = gain_list
    df_summary["Days to Peak"] = days_list

    # Sort
    if "Model Probability" in df_summary.columns:
        df_summary.sort_values(by="Model Probability", ascending=False, inplace=True)

    columns_to_display = [
        # Action & execution
        "Final_Action", "Best_Risk_Reward", "Buy_Window_Status", "Position_Size_Class",
        "Primary_Entry_Price", "Primary_Entry_Source", "Add_On_Dip_Price", "Invalidation_Level",
        "Risk_%", "Reward_%", "Risk_Reward_Ratio",

        # Price context
        "Current Price",

        "Symbol", "Refined Buy Price", "Candle Entry 2w", "Candle Entry 4w",
        "Candle Entry 6w", "Candle Entry 8w", "Candle Entry 12w", "Candle Entry 18w", "Candle Entry 30w",
        "VWAP Support", "ADX", "Institutional Score",
        "Volume Weight", "Confidence Score", "Sector Correlation",
        "Trend", "Recommendation", "Decision Reason", "Confidence Grade", "Expected Holding Period",
        "Momentum Recommendation", "Momentum Decision Reason", "Momentum Confidence Grade", "Momentum Expected Holding Period", "Darvas Breakout %", "Darvas Signal",
        "Rule-Based Buy", "Model-Driven Buy", "Model-Driven Strong Buy", "Model Probability",
        "Confidence Band", "Tech Fallback Score", "Signal",
                "Breakout", "Undervalued", "Trend Reversal", "Pattern Detected", "DipReclaim",
        "90D Hit", "90D Gain (%)", "Days to Peak",
        "EMA Uptrend", "EMA21 Slope", "ADX Strength", "MACD Cross", "RSI", "RSI State",
        "OBV Trend", "At BB Lower",
        "Volume Surge", "Near Support", "Signal Score",
        "Price Reversal",
        "SMC_Breakout", "Mean_Reversion", "Bullish_Engulfing", "Hammer", "Trend_Strength",
        "Market Stage", "Market Sub-Stage",
        "Sym Vol Regime", "VIX Vol Regime", "Volume Pressure",
        "Exit Now", "Atr Trailing Stop", "Exit Reasons",
        "EPS Increase 2Q", "EPS Increase 3Q", "EPS Increase 4Q",
        "News Sentiment Score", "News Positive Ratio", "News Article Count", "Sentiment Confidence",
    ]

    # Ensure optional cols exist
    for col in ["Volume Surge", "Near Support", "Signal Score"]:
        if col not in df_summary.columns:
            df_summary[col] = None

    # Build output frame in that order; create any missing display cols as None so Excel gets filled
    for col in columns_to_display:
        if col not in df_summary.columns:
            df_summary[col] = None
    out_df = df_summary[columns_to_display].copy()

    # Ensure Exit Reasons always populated in the exported sheet
    if "Exit Reasons" in out_df.columns:
        er = out_df["Exit Reasons"].astype(str).fillna("").str.strip()
        out_df["Exit Reasons"] = np.where(er.eq(""), "No exit signals", er)


    # ------------------------------------------------------------------
    # FINAL SAFETY PASS: Force-fill Risk_% and Reward_% right before writing.
    # This guarantees the Excel columns are never blank even if upstream
    # logic produced NaN/None due to earlier fallbacks.
    # ------------------------------------------------------------------
    def _sf(v, default=np.nan):
        try:
            if isinstance(v, pd.Series):
                v = v.iloc[-1] if len(v) else default
            v = float(v)
            return v if np.isfinite(v) else default
        except Exception:
            return default

    if "Risk_%" in out_df.columns:
        entry = pd.to_numeric(out_df.get("Primary_Entry_Price"), errors="coerce")
        if entry is None:
            entry = pd.Series([np.nan] * len(out_df))
        stop = pd.to_numeric(out_df.get("Invalidation_Level"), errors="coerce")
        if stop is None:
            stop = pd.Series([np.nan] * len(out_df))

        # If stop missing, fall back to ATR / VWAP / candle supports, then 3% below entry
        if stop.isna().all():
            for c in ["Atr Trailing Stop", "VWAP Support", "Candle Entry 8w", "Candle Entry 12w"]:
                if c in out_df.columns:
                    cand = pd.to_numeric(out_df.get(c), errors="coerce")
                    stop = stop.fillna(cand)
            stop = stop.fillna(entry * 0.97)

        # Ensure stop < entry
        stop = stop.where(stop < entry, entry * 0.97)

        risk = ((entry - stop) / entry) * 100.0
        risk = risk.clip(lower=0.0)
        out_df["Risk_%"] = risk.round(2)

        # Fill any remaining missing Risk_% using Refined Buy Price and Atr Trailing Stop (last-resort)
        if out_df["Risk_%"].isna().any():
            e2 = pd.to_numeric(out_df.get("Refined Buy Price"), errors="coerce")
            s2 = pd.to_numeric(out_df.get("Atr Trailing Stop"), errors="coerce")
            alt = ((e2 - s2) / e2) * 100.0
            out_df.loc[out_df["Risk_%"].isna(), "Risk_%"] = alt.round(2)


    if "Reward_%" in out_df.columns:
        # Prefer 90D Gain (%) if present; else use proxy based on Recommendation
        if "90D Gain (%)" in out_df.columns:
            rew = pd.to_numeric(out_df.get("90D Gain (%)"), errors="coerce")
        else:
            rew = pd.Series([np.nan] * len(out_df))

        recs = out_df.get("Recommendation", pd.Series([""] * len(out_df))).astype(str).str.upper()
        proxy = recs.map(lambda x: 15.0 if x in ("STRONG_BUY","STRONG BUY") else 10.0 if x == "BUY" else 5.0 if x == "WATCH" else 0.0)

        rew = rew.where(rew.notna() & (rew != 0), proxy)
        out_df["Reward_%"] = pd.to_numeric(rew, errors="coerce").round(2)


    # ------------------------------------------------------------------
    # Write Excel (single authoritative write)
    # Convert NaN/inf -> None so Excel cells are truly populated/blank as intended.
    # ------------------------------------------------------------------
    out_df = out_df.replace([np.inf, -np.inf], np.nan)
    out_df = out_df.where(pd.notna(out_df), None)

    write_template_excel(out_df, template_path=template_path, out_path=out_path, sheet_name=args.sheet)
    print(f"\n📊 Excel saved → {out_path}")

if __name__ == "__main__":
    main()
